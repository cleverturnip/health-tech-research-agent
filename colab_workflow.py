"""
Health Tech Research Agent - Colab Workflow

This file mirrors the active Google Colab workflow.

Current execution environment:

* Google Colab
* Google Drive for persistent storage
* OpenAI API for fit synthesis
* CSV/XLSX exports for dashboard outputs

Important:

* This file is currently a source-of-truth reference for copy/paste Colab cells.
* It is not yet a fully modular script.
* As the workflow stabilizes, cells should be converted into functions and eventually into a runnable pipeline.
  """

# =============================================================================

# STEP 1 - Environment / imports / setup

# =============================================================================

!pip install -q openai pandas

# =============================================================================

# STEP 2 - Company/source config

# =============================================================================

import os
import time
import json
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openai import OpenAI, RateLimitError, APIError
from google.colab import userdata

# --- Make the version-controlled package importable. Same precedent the fit-brief
# --- taxonomy block already used (sys.path -> repo/src). The research runner now
# --- lives in the package: src/health_tech_research_agent/research_runner.py
REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
if SRC_DIR.exists() and str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from health_tech_research_agent import research_runner as _research_runner
from health_tech_research_agent.research_runner import run_research_batch

# Pull your API key securely from Colab Secrets
os.environ["OPENAI_API_KEY"] = userdata.get("OPENAI_API_KEY")

client = OpenAI()

# Use the same model you are using in the Playground.
# If this errors, copy the exact model name from the Playground "Code" button.
MODEL = "gpt-5.4-mini"

# Rate-limit protection
WAIT_BETWEEN_WEB_SEARCHES = 120  # seconds
MAX_RETRIES = 3

# =============================================================================

# STEP 3 - Search / research helpers

# =============================================================================

def call_openai(prompt, use_web_search=False, max_output_tokens=500):
    """Notebook shim -> package call_openai (src/.../research_runner.py).

    Binds the live OpenAI client, MODEL, and MAX_RETRIES to the package
    implementation. Behavior is unchanged (retry on RateLimitError with
    90*attempt waits; APIError re-raises immediately); only the source moved.
    """
    return _research_runner.call_openai(
        prompt,
        client=client,
        model=MODEL,
        use_web_search=use_web_search,
        max_output_tokens=max_output_tokens,
        max_retries=MAX_RETRIES,
    )

# =============================================================================

# STEP 4 - Raw research functions

# =============================================================================

# Purpose:

# - Funding research

# - Payer / institutional distribution research

# - Outcomes research

# - Commercial scale / revenue-quality research

# 4 - Prompt 1 research functions
# Purpose:
# - Collect structured latest-status research inputs for each company
# - Funding / stage
# - Payer, employer, provider, institutional distribution
# - Outcomes / clinical / engagement evidence
# - Commercial scale / revenue quality
#
# Depends on:
# - call_openai(prompt, use_web_search=True, max_output_tokens=...)

def search_funding(research_query):
    """Notebook shim -> package search_funding (web search ON, 400 tokens)."""
    return _research_runner.search_funding(research_query, client=client, model=MODEL)


def search_payer_signal(research_query):
    """Notebook shim -> package search_payer_signal (web search ON, 350 tokens)."""
    return _research_runner.search_payer_signal(research_query, client=client, model=MODEL)


def search_outcomes(research_query):
    """Notebook shim -> package search_outcomes (web search ON, 350 tokens)."""
    return _research_runner.search_outcomes(research_query, client=client, model=MODEL)


def search_commercial_scale(research_query):
    """Notebook shim -> package search_commercial_scale (web search ON, 700 tokens)."""
    return _research_runner.search_commercial_scale(research_query, client=client, model=MODEL)


def search_org_events(research_query):
    """Notebook shim -> package search_org_events (web search ON, 800 tokens; Slice 3.7).

    Recency-bounded org / leadership events -> reset_evidence.
    """
    return _research_runner.search_org_events(research_query, client=client, model=MODEL)


def search_operating_characteristics(research_query):
    """Notebook shim -> package search_operating_characteristics (web search ON, 800 tokens; Slice 3.7).

    Product-engagement + operational-strain evidence -> capability-fit (scored in Slice 4).
    """
    return _research_runner.search_operating_characteristics(research_query, client=client, model=MODEL)

# =============================================================================

# STEP 5 - Company fit synthesis prompt

# =============================================================================

# Purpose:

# - Convert raw research findings into structured JSON

# - Score thesis fit, PMF/scale, evidence confidence, role fit, timing

# - Output scale signal fields

# - Use priority gate logic

# =============================================================================
# STEP 5 - Company fit synthesis prompt
# =============================================================================
# Purpose:
# - Convert raw research findings into structured company fit JSON
# - Score thesis fit, PMF/scale, evidence confidence, Katelynd role fit, and operator timing
# - Classify commercial, institutional, and outcomes scale signals explicitly
# - Use native P0/P1/P2/P3/P4 priority labels
# - Use scale-engine logic:
#   revenue quality/commercial traction OR institutional distribution can each be a primary scale engine
# - Do not let high role fit or thesis relevance override weak PMF/scale evidence


def load_taxonomy_prompt_block_for_fit_brief():
    """Notebook shim -> package taxonomy-block loader, pointed at the repo taxonomy dir."""
    return _research_runner.load_taxonomy_prompt_block_for_fit_brief(REPO_DIR / "taxonomy")


def run_company_fit_brief(company_name, latest_status_findings):
    """Notebook shim -> package run_company_fit_brief (src/.../research_runner.py).

    Binds the live client, MODEL, and repo taxonomy dir. The synthesis prompt, JSON
    schema, model, web-search-off, and 6500-token cap are unchanged -- the package
    builds the byte-identical prompt (verified against this file in Commit 1).
    """
    return _research_runner.run_company_fit_brief(
        company_name,
        latest_status_findings,
        client=client,
        model=MODEL,
        taxonomy_dir=REPO_DIR / "taxonomy",
    )

# =============================================================================

# STEP 6 - Batch config / paths / company list

# =============================================================================

# 6 - Batch config / file paths
# Run this every time:
# - you start a new batch
# - runtime reconnects/restarts
# - before Step 7

from pathlib import Path

# CHANGE THIS FOR EACH NEW BATCH
BATCH_NAME = "pmf_scale_rubric_d2c_rerun_1"

# Local runtime folder
research_batches_folder = Path("research_batches")
research_batches_folder.mkdir(parents=True, exist_ok=True)

# Standard batch paths
batch_checkpoint_path = research_batches_folder / f"{BATCH_NAME}_checkpoint.csv"
batch_raw_export_path = research_batches_folder / f"{BATCH_NAME}_raw.csv"
batch_summary_export_path = research_batches_folder / f"{BATCH_NAME}_summary.csv"

print("Batch config set.")
print("BATCH_NAME:", BATCH_NAME)
print("batch_checkpoint_path:", batch_checkpoint_path)
print("batch_raw_export_path:", batch_raw_export_path)
print("batch_summary_export_path:", batch_summary_export_path)

# =============================================================================

# STEP 6B - Standard cross-batch checkpoint recovery

# =============================================================================

# 6B - Seed current batch checkpoint from all prior research sources
# Purpose:
# - Define the current batch companies
# - Reuse prior completed research by company, even if saved under a different batch name
# - Search exact current checkpoint, all Drive research_batches files, and raw archive
# - Seed the current local + Drive checkpoint so Step 7 only runs missing companies
#
# Edit the companies list in this cell for each new batch.
# Step 7 should no longer define companies.

import pandas as pd
import shutil
from pathlib import Path
from google.colab import drive

drive.mount("/content/drive")

# -----------------------------
# Safety checks
# -----------------------------

if "BATCH_NAME" not in globals():
    raise NameError("STOP: BATCH_NAME is not defined. Run Step 6 first.")

if "batch_checkpoint_path" not in globals():
    local_batches_folder = Path("research_batches")
    local_batches_folder.mkdir(parents=True, exist_ok=True)
    batch_checkpoint_path = local_batches_folder / f"{BATCH_NAME}_checkpoint.csv"

# -----------------------------
# Current batch company list
# -----------------------------
# Commercial-scale backfill batch:
# Focus: D2C / cash-pay / hybrid companies where revenue quality,
# retention, subscriber growth, acquisition efficiency, and margin structure
# could materially change fit assessment.

companies = [
    {
        "company": "Oura",
        "research_query": "Oura ring revenue growth subscribers paid members retention renewal rate D2C employer payer insurance partnerships outcomes funding"
    },
    {
        "company": "Function Health",
        "research_query": "Function Health revenue growth memberships subscribers retention D2C employer partnerships funding outcomes commercial scale"
    },
    {
        "company": "ZOE",
        "research_query": "ZOE personalized nutrition revenue growth subscribers retention D2C employer payer partnerships outcomes funding commercial scale"
    },
    {
        "company": "Signos",
        "research_query": "Signos CGM metabolic health revenue growth subscribers retention D2C employer payer insurance outcomes funding commercial scale"
    },
    {
        "company": "Levels Health",
        "research_query": "Levels Health metabolic health CGM revenue growth subscribers retention D2C funding outcomes commercial scale"
    },
    {
        "company": "InsideTracker",
        "research_query": "InsideTracker revenue growth subscribers retention memberships D2C employer partnerships funding outcomes commercial scale"
    },
    {
        "company": "Noom Med",
        "research_query": "Noom Med revenue growth subscribers retention employer payer insurance GLP-1 weight loss outcomes funding commercial scale"
    },
    {
        "company": "Oova",
        "research_query": "Oova fertility hormone testing revenue growth subscribers retention D2C employer payer insurance partnerships outcomes funding commercial scale"
    },
    {
        "company": "Outcomes4Me",
        "research_query": "Outcomes4Me cancer patient platform revenue growth users retention pharma partnerships payer provider outcomes funding commercial scale"
    }
]

current_batch_company_names = [
    item["company"] if isinstance(item, dict) else item
    for item in companies
]

# -----------------------------
# Paths
# -----------------------------

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_batches_folder = drive_folder / "research_batches"
drive_batches_folder.mkdir(parents=True, exist_ok=True)

drive_archive_path = drive_folder / "health_tech_raw_research_ARCHIVE.csv"
drive_checkpoint_path = drive_batches_folder / f"{BATCH_NAME}_checkpoint.csv"

required_current_schema_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "org_events_finding",
    "operating_characteristics_finding",
    "fit_brief_json"
]

print("Batch name:", BATCH_NAME)
print("Current batch companies:", current_batch_company_names)
print("Local checkpoint:", batch_checkpoint_path)
print("Drive checkpoint:", drive_checkpoint_path)
print("Drive batches folder:", drive_batches_folder)
print("Raw archive:", drive_archive_path)

# -----------------------------
# Helpers
# -----------------------------

def nonblank(value):
    return pd.notna(value) and str(value).strip() != ""

def source_priority_score(source_name):
    """
    Higher = preferred if duplicate complete rows exist.
    """
    priority = {
        "local_exact_batch_checkpoint": 4,
        "drive_exact_batch_checkpoint": 3,
        "drive_research_batches_any_batch": 2,
        "raw_archive_any_batch": 1
    }
    return priority.get(source_name, 0)

def load_csv_safely(path):
    try:
        temp = pd.read_csv(path)
        if "company" not in temp.columns:
            return None
        return temp
    except Exception as e:
        print(f"Skipping unreadable CSV: {path} | {e}")
        return None

def prepare_source_df(source_df, source_name, source_file):
    """
    Adds completeness scoring and filters to current batch companies.
    """
    if source_df is None or source_df.empty:
        return pd.DataFrame()

    df_source = source_df.copy()

    for col in required_current_schema_cols:
        if col not in df_source.columns:
            df_source[col] = ""

    df_source = df_source[
        df_source["company"].isin(current_batch_company_names)
    ].copy()

    if df_source.empty:
        return pd.DataFrame()

    for col in required_current_schema_cols:
        df_source[f"has_{col}"] = df_source[col].apply(nonblank)

    completeness_cols = [f"has_{col}" for col in required_current_schema_cols]

    df_source["current_schema_completeness"] = df_source[completeness_cols].sum(axis=1)
    df_source["is_current_schema_complete"] = df_source[completeness_cols].all(axis=1)
    df_source["seed_source"] = source_name
    df_source["source_file"] = str(source_file)
    df_source["source_priority"] = source_priority_score(source_name)

    return df_source

def missing_fields_for_row(row):
    return [
        col for col in required_current_schema_cols
        if not nonblank(row.get(col, ""))
    ]

# -----------------------------
# Gather all possible recovery sources
# -----------------------------

candidate_frames = []

# 1. Local exact checkpoint
if batch_checkpoint_path.exists():
    local_df = load_csv_safely(batch_checkpoint_path)
    prepared = prepare_source_df(
        local_df,
        source_name="local_exact_batch_checkpoint",
        source_file=batch_checkpoint_path
    )
    if not prepared.empty:
        candidate_frames.append(prepared)
else:
    print("\nNo local exact-batch checkpoint found.")

# 2. Drive exact checkpoint
if drive_checkpoint_path.exists():
    drive_checkpoint_df = load_csv_safely(drive_checkpoint_path)
    prepared = prepare_source_df(
        drive_checkpoint_df,
        source_name="drive_exact_batch_checkpoint",
        source_file=drive_checkpoint_path
    )
    if not prepared.empty:
        candidate_frames.append(prepared)
else:
    print("\nNo Drive exact-batch checkpoint found.")

# 3. All Drive research batch CSVs, any batch name
drive_batch_csvs = sorted(drive_batches_folder.glob("*.csv"))

print(f"\nScanning Drive research_batches CSVs: {len(drive_batch_csvs)} files found")

for path in drive_batch_csvs:
    # Skip exact current checkpoint here because already handled above.
    if path == drive_checkpoint_path:
        continue

    temp_df = load_csv_safely(path)

    if temp_df is None:
        continue

    prepared = prepare_source_df(
        temp_df,
        source_name="drive_research_batches_any_batch",
        source_file=path
    )

    if not prepared.empty:
        candidate_frames.append(prepared)

# 4. Raw archive, any prior batch
if drive_archive_path.exists():
    archive_df = load_csv_safely(drive_archive_path)
    prepared = prepare_source_df(
        archive_df,
        source_name="raw_archive_any_batch",
        source_file=drive_archive_path
    )
    if not prepared.empty:
        candidate_frames.append(prepared)
else:
    print("\nNo raw archive found in Drive.")

# -----------------------------
# Evaluate candidates
# -----------------------------

if not candidate_frames:
    print("\nNo prior rows found for current batch companies.")
    print("Step 7 will run all current batch companies from scratch.")
else:
    candidates = pd.concat(candidate_frames, ignore_index=True)

    # For reporting: best available row per company, even if incomplete
    best_available = (
        candidates
        .sort_values(
            by=[
                "company",
                "is_current_schema_complete",
                "current_schema_completeness",
                "source_priority",
                "date_researched"
            ],
            ascending=[True, False, False, False, False]
        )
        .drop_duplicates(subset=["company"], keep="first")
        .reset_index(drop=True)
    )

    print("\nBest prior row found per company:")
    display(
        best_available[
            [
                col for col in [
                    "company",
                    "date_researched",
                    "seed_source",
                    "source_file",
                    "current_schema_completeness",
                    "is_current_schema_complete",
                    "funding_finding",
                    "payer_institutional_finding",
                    "outcomes_finding",
                    "commercial_scale_finding"
                ]
                if col in best_available.columns
            ]
        ]
    )

    incomplete_best = best_available[
        ~best_available["is_current_schema_complete"]
    ].copy()

    if not incomplete_best.empty:
        missing_report = []

        for _, row in incomplete_best.iterrows():
            missing_report.append({
                "company": row["company"],
                "seed_source": row.get("seed_source", ""),
                "source_file": row.get("source_file", ""),
                "date_researched": row.get("date_researched", ""),
                "missing_fields": ", ".join(missing_fields_for_row(row))
            })

        print("\nPrior rows exist but are NOT reusable under the current schema:")
        display(pd.DataFrame(missing_report))

    # Reusable = complete current-schema rows only
    reusable = candidates[
        candidates["is_current_schema_complete"]
    ].copy()

    if reusable.empty:
        print("\nNo complete current-schema prior rows found.")
        print("Step 7 will run all current batch companies from scratch.")
    else:
        # Choose best complete row per company.
        reusable_best = (
            reusable
            .sort_values(
                by=[
                    "company",
                    "source_priority",
                    "date_researched"
                ],
                ascending=[True, False, False]
            )
            .drop_duplicates(subset=["company"], keep="first")
            .reset_index(drop=True)
        )

        reusable_checkpoint_rows = reusable_best[required_current_schema_cols].copy()

        # Save seeded checkpoint locally and to Drive
        batch_checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
        reusable_checkpoint_rows.to_csv(batch_checkpoint_path, index=False)
        shutil.copy(batch_checkpoint_path, drive_checkpoint_path)

        print("\nSeeded checkpoint created from complete prior research.")
        print("Seeded companies:", reusable_checkpoint_rows["company"].tolist())
        print("Checkpoint rows:", reusable_checkpoint_rows.shape[0])
        print("Local checkpoint:", batch_checkpoint_path)
        print("Drive checkpoint:", drive_checkpoint_path)

# -----------------------------
# Show what Step 7 will still run
# -----------------------------

if batch_checkpoint_path.exists():
    seeded_df = pd.read_csv(batch_checkpoint_path)
    seeded_companies = set(seeded_df["company"].tolist())
else:
    seeded_df = pd.DataFrame(columns=["company"])
    seeded_companies = set()

companies_to_run = [
    company for company in current_batch_company_names
    if company not in seeded_companies
]

print("\nCompanies already complete / reusable:")
print(sorted(seeded_companies) if seeded_companies else "None")

print("\nCompanies Step 7 still needs to run:")
print(companies_to_run if companies_to_run else "None")

if not seeded_df.empty:
    print("\nSeeded checkpoint preview:")
    display(seeded_df)

# =============================================================================

# STEP 7 - Run research + fit briefs

# =============================================================================

# Step 7 - Run companies
# Purpose:
# - Run research only for companies missing from the seeded/current checkpoint
# - Save local + Drive checkpoint after each company
# - Includes commercial_scale_finding

from pathlib import Path
import pandas as pd
from datetime import datetime
import time
import shutil
from google.colab import drive

# -----------------------------
# Safety checks
# -----------------------------

if "BATCH_NAME" not in globals():
    raise NameError("STOP: BATCH_NAME is not defined. Run Step 6 first.")

if "companies" not in globals():
    raise NameError(
        "STOP: companies list is not defined. Run Step 6B first."
    )

if "batch_checkpoint_path" not in globals():
    local_batches_folder = Path("research_batches")
    local_batches_folder.mkdir(parents=True, exist_ok=True)
    batch_checkpoint_path = local_batches_folder / f"{BATCH_NAME}_checkpoint.csv"

checkpoint_path = batch_checkpoint_path

required_current_schema_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "org_events_finding",
    "operating_characteristics_finding",
    "fit_brief_json"
]

def nonblank(value):
    return pd.notna(value) and str(value).strip() != ""

def row_is_current_schema_complete(row):
    return all(nonblank(row.get(col, "")) for col in required_current_schema_cols)

# -----------------------------
# Drive checkpoint path
# -----------------------------

drive.mount("/content/drive")

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_batches_folder = drive_folder / "research_batches"
drive_batches_folder.mkdir(parents=True, exist_ok=True)

drive_checkpoint_path = drive_batches_folder / f"{BATCH_NAME}_checkpoint.csv"

# =============================================================================
# COLAB VERIFICATION CHECKLIST - Slice 1 research-runner rewire
# =============================================================================
# The offline test suite proves the PACKAGE functions are faithful, but it cannot
# verify this notebook's wiring to them. After pulling this branch in Colab, run
# these checks in order. Each line says what you should SEE if the rewire works.
#
# 1. Package import resolves (run STEP 1-2):
#    EXPECT: no ModuleNotFoundError; print(_research_runner.DEFAULT_MODEL) -> "gpt-5.4-mini".
#    If it fails: confirm the repo is cloned at REPO_DIR and REPO_DIR/src exists.
#
# 2. Shims bind with the OLD signatures (run STEP 3-5, then a scratch cell):
#    EXPECT: callable(call_openai) and callable(run_company_fit_brief) -> True.
#    EXPECT: call_openai("Reply with the single word OK.") -> short text, no error.
#    EXPECT: run_company_fit_brief("Test Co", "Funding: seed round.") -> a JSON
#            string that json.loads() parses (keys include scores, scale_signal_assessment).
#
# 3. STEP 7 runs a small batch and produces the SAME 7-column checkpoint:
#    Set `companies` to ONE real company; run STEP 6 / 6B / 7.
#    EXPECT: console shows "Researched this run: [<company>]" and "Failed ...: {}".
#    EXPECT: pd.read_csv(batch_checkpoint_path).columns.tolist() ==
#            ["company","date_researched","funding_finding",
#             "payer_institutional_finding","outcomes_finding",
#             "commercial_scale_finding","fit_brief_json"]   (same schema + order),
#            exactly one row, every cell non-blank.
#    EXPECT: the Drive mirror at drive_checkpoint_path matches the local file.
#    Resume: re-run STEP 7 -> EXPECT "Reused (already complete): [<company>]" and
#            the company is skipped (not re-researched).
#
# 4. Per-company recovery (optional - to watch recovery happen live):
#    Add one nonsense company name alongside a real one and run STEP 7.
#    EXPECT: the nonsense one in "Failed (retried on resume)", the real one in
#            "Researched this run", the batch does NOT abort, and the checkpoint
#            holds only the real company.
#
# 5. STEP 26 still rescores (its exec()-based globals contract):
#    Run STEP 26 (DRY_RUN) on a couple of companies that have archived evidence.
#    EXPECT: "PASS: loaded latest Step 5 scoring rubric." and MODEL printed -- i.e.
#            its assertion that client / MODEL / call_openai / run_company_fit_brief
#            exist as globals PASSES (now package-backed shims; the names persist).
#    EXPECT: rescore produces fit briefs with no NameError.
#
# 6. STEP 9 backstop still fires on bad JSON (kept as the batch-wide safety net):
#    EXPECT (normal): STEP 9 reports all fit_brief_json parse OK.
#    To confirm it still STOPS: hand-edit one checkpoint fit_brief_json cell to
#    invalid JSON and run STEP 9 -> EXPECT it raises
#    "Invalid fit_brief_json found. Do not archive until repaired."
# =============================================================================

# -----------------------------
# Run the batch via the package research runner
# -----------------------------
# The loop, per-company checkpointing, and the (new) per-company error recovery now
# live in the package (src/health_tech_research_agent/research_runner.py), proven
# faithful by the offline suite. This cell only wires the live client and the
# Colab/Drive paths into it.
#
# Resume is unchanged: companies already complete (all 7 columns non-blank) in the
# checkpoint are skipped. New: one company's failure (API error, network, or an
# unparseable fit brief) is caught and recorded; the batch continues, and the failed
# company is retried on the next run instead of aborting the whole batch.

batch_result = run_research_batch(
    companies,
    client=client,
    model=MODEL,
    checkpoint_path=batch_checkpoint_path,
    mirror_checkpoint_path=drive_checkpoint_path,
    taxonomy_dir=REPO_DIR / "taxonomy",
    wait_between_searches=WAIT_BETWEEN_WEB_SEARCHES,
)

# Final df for downstream steps (Step 8 / 9 / 10 consume `df`), re-read from the
# checkpoint the runner just wrote -- faithful to the prior end-of-Step-7 behavior.
if batch_checkpoint_path.exists():
    df = pd.read_csv(batch_checkpoint_path)
else:
    df = pd.DataFrame(columns=required_current_schema_cols)

print("\nStep 7 complete.")
print("Researched this run:", batch_result.completed)
print("Reused (already complete):", batch_result.reused)
print("Failed (retried on resume):", batch_result.failed)
print("df shape:", df.shape)
print("companies in df:", df["company"].tolist() if "company" in df.columns else [])

df

# =============================================================================

# STEP 8 - Save current batch checkpoint

# =============================================================================

print("About to save CURRENT BATCH ONLY.")
print("df shape:", df.shape)
print("companies:", df["company"].tolist())

df.to_csv(batch_checkpoint_path, index=False)

print("Saved current batch checkpoint to:")
print(batch_checkpoint_path)

# =============================================================================

# STEP 8A - Save raw research archive

# =============================================================================

# 8A - Save raw research archive
# Purpose:
# - Preserve the full raw research layer permanently
# - Save current batch raw research to Google Drive
# - Append current batch raw research to a permanent archive
# - Download local copies so Colab runtime loss does not wipe the work
# - Prevent incomplete raw records from being archived
# - Includes commercial_scale_finding for revenue quality / D2C scale mechanics

import pandas as pd
from pathlib import Path
from datetime import datetime
import shutil
from google.colab import drive, files

# -----------------------------
# Safety checks
# -----------------------------

if "df" not in globals():
    raise NameError("STOP: df does not exist. Run Step 7 / Step 8 before running 8A.")

if "company" not in df.columns:
    raise ValueError("STOP: df does not have a company column. This does not look like the raw research dataframe.")

if "BATCH_NAME" not in globals():
    raise NameError("STOP: BATCH_NAME is not defined. Run Step 6 before running 8A.")

required_raw_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "org_events_finding",
    "operating_characteristics_finding",
    "fit_brief_json"
]

missing_cols = [col for col in required_raw_cols if col not in df.columns]

if missing_cols:
    raise ValueError(
        f"STOP: df is missing required raw research columns: {missing_cols}. "
        "If this is an old checkpoint, rerun the companies under the new Step 7 schema."
    )

if df.shape[0] == 0:
    raise ValueError("STOP: df has zero rows. Nothing to archive.")

# -----------------------------
# Blank-field protection for current batch
# -----------------------------

blank_issues = []

for col in required_raw_cols:
    blank_mask = df[col].isna() | (df[col].astype(str).str.strip() == "")
    if blank_mask.any():
        for company in df.loc[blank_mask, "company"].tolist():
            blank_issues.append({
                "company": company,
                "blank_field": col
            })

if blank_issues:
    blank_df = pd.DataFrame(blank_issues)
    print("STOP: Blank raw research fields found in current batch. Repair these before archiving.")
    display(blank_df)
    raise ValueError("Blank raw research fields found. Do not archive until repaired.")

# -----------------------------
# Validate fit_brief_json parses for current batch
# -----------------------------

import json

json_issues = []

for idx, row in df.iterrows():
    company = row["company"]
    value = row["fit_brief_json"]

    try:
        json.loads(str(value))
    except Exception as e:
        json_issues.append({
            "company": company,
            "issue": str(e)
        })

if json_issues:
    json_issue_df = pd.DataFrame(json_issues)
    print("STOP: Some fit_brief_json values are not valid JSON. Repair before archiving.")
    display(json_issue_df)
    raise ValueError("Invalid fit_brief_json found. Do not archive until repaired.")

# -----------------------------
# Add archive metadata
# -----------------------------

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

raw_batch_df = df.copy()

raw_batch_df["batch_name"] = BATCH_NAME
raw_batch_df["archive_saved_at"] = timestamp
raw_batch_df["archive_record_id"] = (
    raw_batch_df["batch_name"].astype(str)
    + "__"
    + raw_batch_df["company"].astype(str).str.replace(" ", "_", regex=False)
    + "__"
    + raw_batch_df["archive_saved_at"].astype(str)
)

# -----------------------------
# Local runtime paths
# -----------------------------

local_batches_folder = Path("research_batches")
local_batches_folder.mkdir(parents=True, exist_ok=True)

current_batch_raw_path = local_batches_folder / f"{BATCH_NAME}_raw_{timestamp}.csv"
current_batch_checkpoint_path = local_batches_folder / f"{BATCH_NAME}_checkpoint.csv"
raw_archive_path = Path("health_tech_raw_research_ARCHIVE.csv")

# -----------------------------
# Mount Google Drive and pull latest archive if it exists
# -----------------------------

drive.mount("/content/drive")

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_batches_folder = drive_folder / "research_batches"

drive_folder.mkdir(parents=True, exist_ok=True)
drive_batches_folder.mkdir(parents=True, exist_ok=True)

drive_raw_archive_path = drive_folder / "health_tech_raw_research_ARCHIVE.csv"

# If Drive already has an archive, use that as the starting source of truth.
# This prevents a fresh Colab runtime from accidentally creating a tiny archive.
if drive_raw_archive_path.exists():
    shutil.copy(drive_raw_archive_path, raw_archive_path)
    print("Loaded existing raw archive from Google Drive.")
else:
    print("No existing raw archive found in Google Drive. Creating a new one.")

# -----------------------------
# Save current batch raw files locally
# -----------------------------

raw_batch_df.to_csv(current_batch_raw_path, index=False)
raw_batch_df.to_csv(current_batch_checkpoint_path, index=False)

# -----------------------------
# Append to archive
# -----------------------------

if raw_archive_path.exists():
    existing_archive = pd.read_csv(raw_archive_path)

    # Backward-compatible archive schema migration.
    # Older archived rows may not have commercial_scale_finding.
    # Keep them, but add the column blank so the archive schema is consistent.
    for col in required_raw_cols:
        if col not in existing_archive.columns:
            existing_archive[col] = ""

    for col in ["batch_name", "archive_saved_at", "archive_record_id"]:
        if col not in existing_archive.columns:
            existing_archive[col] = ""

    combined_archive = pd.concat([existing_archive, raw_batch_df], ignore_index=True)
else:
    combined_archive = raw_batch_df.copy()

# Avoid duplicate company records within the same batch if you rerun 8A after a repair.
# Keeps the latest version for each company + batch_name.
combined_archive = combined_archive.drop_duplicates(
    subset=["company", "batch_name"],
    keep="last"
).reset_index(drop=True)

combined_archive.to_csv(raw_archive_path, index=False)

# -----------------------------
# Google Drive save
# -----------------------------

drive_current_batch_raw_path = drive_batches_folder / f"{BATCH_NAME}_raw_{timestamp}.csv"
drive_current_batch_checkpoint_path = drive_batches_folder / f"{BATCH_NAME}_checkpoint.csv"
drive_raw_archive_backup_path = drive_folder / f"health_tech_raw_research_ARCHIVE_backup_{timestamp}.csv"

# Save current batch to Drive
shutil.copy(current_batch_raw_path, drive_current_batch_raw_path)
shutil.copy(current_batch_checkpoint_path, drive_current_batch_checkpoint_path)

# Save updated archive to Drive
shutil.copy(raw_archive_path, drive_raw_archive_path)
shutil.copy(raw_archive_path, drive_raw_archive_backup_path)

# -----------------------------
# Validation output
# -----------------------------

print("Raw research archive saved successfully.")
print()
print("Current batch rows:", raw_batch_df.shape[0])
print("Current batch companies:", raw_batch_df["company"].tolist())
print()
print("Local current batch raw:")
print(current_batch_raw_path)
print()
print("Local raw archive:")
print(raw_archive_path)
print("Archive shape:", combined_archive.shape)
print("Archive company count:", combined_archive["company"].nunique())
print()
print("Google Drive current batch raw:")
print(drive_current_batch_raw_path)
print()
print("Google Drive raw archive:")
print(drive_raw_archive_path)
print()
print("Google Drive archive backup:")
print(drive_raw_archive_backup_path)

print()
print("Current batch raw field check:")
display(
    raw_batch_df[
        [
            "company",
            "funding_finding",
            "payer_institutional_finding",
            "outcomes_finding",
            "commercial_scale_finding"
        ]
    ]
)

# Show archive company counts by batch
print()
print("Archive batch summary:")
display(
    combined_archive
    .groupby("batch_name")
    .agg(
        company_count=("company", "nunique"),
        row_count=("company", "count")
    )
    .reset_index()
)

# -----------------------------
# Download local copies
# -----------------------------

files.download(str(current_batch_raw_path))
files.download(str(raw_archive_path))

# =============================================================================

# STEP 9 - Print raw batch results

# =============================================================================

pd.set_option("display.max_colwidth", None)

print("Printing CURRENT BATCH raw results only.")
print("df shape:", df.shape)
print("companies:", df["company"].tolist())

for idx, row in df.iterrows():
    print("\n" + "="*80)
    print(row["company"])
    print("="*80)
    print("FUNDING:")
    print(row["funding_finding"])
    print("\nPAYER / INSTITUTIONAL:")
    print(row["payer_institutional_finding"])
    print("\nOUTCOMES:")
    print(row["outcomes_finding"])
    print("\nFIT BRIEF:")
    print(row["fit_brief_json"])

# =============================================================================

# STEP 10 - Parse fit brief JSON into score/summary table

# =============================================================================

# Step 10 - Validation summary
# Purpose:
# - Parse fit_brief_json into summary_df
# - Extract scores, recommendation, priority, business model, and commercial-scale assessment
# - Normalize nested/dict outputs into clean text
# - Add calibration flags before export / master update

import json
import re
import pandas as pd

# -----------------------------
# Helpers
# -----------------------------

def clean_json_text(text):
    text = str(text).strip()
    text = re.sub(r"^```json", "", text)
    text = re.sub(r"^```", "", text)
    text = re.sub(r"```$", "", text)
    return text.strip()

def parse_first_json_object(text):
    """
    Parses the first valid JSON object from a text string.
    This handles cases where the model returns valid JSON plus extra text afterward.
    """
    raw = clean_json_text(text)
    start = raw.find("{")

    if start == -1:
        raise ValueError("No JSON object found")

    decoder = json.JSONDecoder()
    parsed, end = decoder.raw_decode(raw[start:])

    return parsed

def extract_score(scores, key):
    value = scores.get(key)

    if isinstance(value, dict):
        return value.get("score")

    return value

def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

# -----------------------------
# Company identity normalization
# -----------------------------
# Prevent duplicate rows when the same company appears under a variant name.
# Keep this intentionally conservative; only add aliases when we are confident.

COMPANY_ALIASES = {
    "fay nutrition": "Fay",
}

def normalize_company_key(value):
    text = safe_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def canonical_company_name(value):
    original = safe_text(value)
    if original == "":
        return original
    return COMPANY_ALIASES.get(normalize_company_key(original), original)

def as_number(value):
    try:
        return float(value)
    except Exception:
        return None

def normalize_jsonish(value):
    """
    Converts nested dict/list fields into readable text for export/display.
    If a dict has a summary field, use that summary.
    Otherwise, convert to a clean JSON string.
    """
    if isinstance(value, dict):
        if "summary" in value and str(value["summary"]).strip() != "":
            return str(value["summary"]).strip()
        return json.dumps(value, ensure_ascii=False)

    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)

    if pd.isna(value):
        return ""

    return str(value).strip()

def is_d2c_specific_business_model(business_model_text):
    """
    Detects true D2C / cash-pay companies without over-flagging
    hybrid payer/employer/provider-enabled consumer platforms.
    """
    business_model = safe_text(business_model_text).lower()

    institutional_terms = [
        "b2b2c",
        "payer",
        "employer",
        "provider",
        "benefits",
        "insurance",
        "health plan",
        "government",
        "pharma",
        "medicaid",
        "medicare"
    ]

    has_institutional_path = any(term in business_model for term in institutional_terms)

    appears_d2c = (
        "d2c" in business_model
        or "cash-pay" in business_model
        or "cash pay" in business_model
        or "primarily d2c" in business_model
        or "primarily cash" in business_model
    )

    # Do not treat generic "consumer health platform" as D2C if the classification
    # is clearly hybrid/B2B2C/payer/employer/provider-enabled.
    d2c_specific_review_needed = appears_d2c and not (
        has_institutional_path and "hybrid" in business_model
    )

    return d2c_specific_review_needed

def commercial_signal_is_weak(commercial_scale_finding, commercial_scale_assessment):
    text = (
        safe_text(commercial_scale_finding)
        + " "
        + safe_text(commercial_scale_assessment)
    ).lower()

    weak_phrases = [
        "no strong public commercial scale evidence found",
        "weak evidence",
        "commercial scale evidence is weak",
        "weak-to-moderate",
        "not enough hard evidence",
        "not strongly substantiated",
        "not revenue quality",
        "no credible public revenue",
        "no public revenue",
        "no verified current arr",
        "no public arr",
        "no subscriber count",
        "no retention",
        "no renewal",
        "no cac",
        "no margin"
    ]

    if safe_text(commercial_scale_finding) == "":
        return True

    return any(phrase in text for phrase in weak_phrases)

def build_calibration_flag(row):
    flags = []

    priority = safe_text(row.get("priority_level"))
    business_model = safe_text(row.get("business_model_classification"))

    commercial_scale_finding = safe_text(row.get("commercial_scale_finding"))
    commercial_scale_assessment = safe_text(row.get("commercial_scale_assessment"))

    thesis = as_number(row.get("thesis_fit_score"))
    pmf = as_number(row.get("pmf_scale_score"))
    evidence = as_number(row.get("evidence_confidence_score"))
    role_fit = as_number(row.get("katelynd_role_fit_score"))
    operator_timing = as_number(row.get("operator_timing_score"))

    # P2 evidence / PMF checks
    if priority == "P2: Worth deeper diligence":
        if evidence is not None and evidence < 50:
            flags.append("CHECK: P2 with low evidence")

        if pmf is not None and pmf < 60:
            flags.append("CHECK: P2 with weak PMF/scale")

        if (
            role_fit is not None
            and operator_timing is not None
            and role_fit < 70
            and operator_timing < 70
        ):
            flags.append("CHECK: P2 with weak role/timing fit")

        if (
            evidence is not None
            and pmf is not None
            and (evidence < 65 or pmf < 70)
        ):
            flags.append("REVIEW: P2 with moderate evidence/PMF")

    # Possible P1 under-promotion
    if priority != "P1: High-priority target":
        if (
            thesis is not None
            and pmf is not None
            and evidence is not None
            and role_fit is not None
            and operator_timing is not None
            and thesis >= 90
            and pmf >= 80
            and evidence >= 70
            and role_fit >= 80
            and operator_timing >= 75
        ):
            flags.append("CHECK: possible P1 under-promotion")

    # D2C / commercial-scale checks
    d2c_specific_review_needed = is_d2c_specific_business_model(business_model)
    weak_commercial_signal = commercial_signal_is_weak(
        commercial_scale_finding,
        commercial_scale_assessment
    )

    if (
        d2c_specific_review_needed
        and priority in ["P1: High-priority target", "P2: Worth deeper diligence"]
        and weak_commercial_signal
    ):
        flags.append("REVIEW: D2C priority with weak commercial-scale evidence")

    if (
        d2c_specific_review_needed
        and pmf is not None
        and evidence is not None
        and pmf >= 75
        and evidence < 60
    ):
        flags.append("REVIEW: D2C scale claim with moderate/weak evidence confidence")

    return " | ".join(flags)

# -----------------------------
# Parse fit briefs
# -----------------------------


def step10_taxonomy_join(value):
    """Flatten taxonomy tag arrays or strings into semicolon-separated strings."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join([str(item).strip() for item in value if str(item).strip()])
    return str(value).strip()


summary_rows = []

for _, row in df.iterrows():
    company = row["company"]
    raw = clean_json_text(row["fit_brief_json"])

    try:
        parsed = parse_first_json_object(raw)
        scores = parsed.get("scores", {})
        taxonomy_classification = parsed.get("taxonomy_classification", {})
        if not isinstance(taxonomy_classification, dict):
            taxonomy_classification = {}

        summary_rows.append({
            "company": company,
            "thesis_fit_score": extract_score(scores, "thesis_fit_score"),
            "pmf_scale_score": extract_score(scores, "pmf_scale_score"),
            "evidence_confidence_score": (
                extract_score(scores, "evidence_confidence_score")
                or extract_score(scores, "overall_confidence")
            ),
            "katelynd_role_fit_score": extract_score(scores, "katelynd_role_fit_score"),
            "operator_timing_score": extract_score(scores, "operator_timing_score"),
            "final_recommendation": parsed.get("final_recommendation"),
            "priority_level": parsed.get("priority_level"),
            "business_model_classification": parsed.get("business_model_classification"),
            "primary_market_segment_code": taxonomy_classification.get("primary_market_segment", ""),
            "primary_market_segment": taxonomy_classification.get("primary_market_segment", ""),
            "subsegment_tags": step10_taxonomy_join(taxonomy_classification.get("subsegment_tags", "")),
            "product_model_tags": step10_taxonomy_join(taxonomy_classification.get("product_model_tags", "")),
            "distribution_model_tags": step10_taxonomy_join(taxonomy_classification.get("distribution_model_tags", "")),
            "data_input_tags": step10_taxonomy_join(taxonomy_classification.get("data_input_tags", "")),
            "taxonomy_assignment_method": "llm_fit_brief" if taxonomy_classification.get("primary_market_segment", "") else "",
            "taxonomy_assignment_basis": taxonomy_classification.get("classification_rationale", "") if taxonomy_classification.get("primary_market_segment", "") else "",
            "commercial_scale_assessment": parsed.get("commercial_scale_assessment"),
            "final_takeaway": parsed.get("final_takeaway"),
            "commercial_scale_finding": row.get("commercial_scale_finding", "")
        })

    except Exception as e:
        summary_rows.append({
            "company": company,
            "error": f"Could not parse JSON: {e}",
            "raw_preview": raw[:500],
            "commercial_scale_finding": row.get("commercial_scale_finding", "")
        })

summary_df = pd.DataFrame(summary_rows)

# -----------------------------
# Role/timing maturity assessment + timing caps
# -----------------------------
# Purpose:
# - Separate "good company" from "right timing for Katelynd."
# - Preserve raw timing score and apply deterministic caps for mature companies.
# - Surface maturity/agency fields in summary exports and review packets.

ROLE_TIMING_FIELDS = [
    "company_maturity_read",
    "likely_agency_level",
    "stage_timing_fit",
    "why_now_or_why_not",
    "timing_penalty_applied",
]

for _col in ROLE_TIMING_FIELDS + [
    "operator_timing_score_raw",
    "operator_timing_score_cap",
    "operator_timing_calibration_flag",
    "maturity_needs_review",
]:
    if _col not in summary_df.columns:
        summary_df[_col] = ""

def _rt_safe_text(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()

def _rt_company_key(value):
    return normalize_company_key(value)

_rt_raw_lookup = {}
if "df" in globals() and isinstance(df, pd.DataFrame) and "company" in df.columns:
    for _, _raw_row in df.iterrows():
        _rt_raw_lookup[_rt_company_key(_raw_row.get("company", ""))] = _raw_row

def _rt_get_parsed_for_company(company):
    _raw_row = _rt_raw_lookup.get(_rt_company_key(company))
    if _raw_row is None:
        return {}
    try:
        return parse_first_json_object(_raw_row.get("fit_brief_json", ""))
    except Exception:
        return {}

def _rt_get_raw_text_for_company(company, parsed=None):
    parsed = parsed or {}
    _raw_row = _rt_raw_lookup.get(_rt_company_key(company))

    pieces = [
        company,
        parsed.get("business_model_classification", ""),
        parsed.get("commercial_scale_assessment", ""),
        parsed.get("pmf_scale_assessment", ""),
        parsed.get("final_takeaway", ""),
        parsed.get("calibration_flag", ""),
    ]

    if _raw_row is not None:
        pieces.extend([
            _raw_row.get("funding_finding", ""),
            _raw_row.get("payer_institutional_finding", ""),
            _raw_row.get("outcomes_finding", ""),
            _raw_row.get("commercial_scale_finding", ""),
        ])

    return " ".join([_rt_safe_text(piece).lower() for piece in pieces])

# _rt_infer_maturity_read RETIRED (Slice 2): company_maturity_read is now derived
# deterministically from the researched funding_stage / ipo_status by
# structured_evidence.derive_maturity (revenue/scale never touch maturity), not from a
# keyword text-scan over the findings.
from health_tech_research_agent.structured_evidence import derive_maturity

def _rt_has_high_agency_exception(text):
    exception_terms = [
        "new business line", "new line of business", "operating rebuild", "rebuild",
        "turnaround", "operationally immature", "founder-led", "founder led",
        "needs operating system", "build the operating system", "high-agency",
        "high agency", "white space", "whitespace", "major scaling phase",
        "scaling from early traction", "pre-professionalized", "not yet professionalized",
    ]
    return any(term in text for term in exception_terms)

def _rt_operator_timing_cap(maturity_read, high_agency_exception=False):
    maturity = _rt_safe_text(maturity_read).lower()

    if maturity == "public":
        return 65 if high_agency_exception else 50
    if maturity == "late-stage":
        return 70 if high_agency_exception else 60
    if maturity == "scale-up":
        return 90 if high_agency_exception else 80
    return None

def _rt_default_agency_level(maturity_read, high_agency_exception):
    maturity = _rt_safe_text(maturity_read).lower()
    if high_agency_exception:
        return "role-dependent"
    if maturity in ["public", "late-stage"]:
        return "low"
    if maturity == "scale-up":
        return "medium"
    if maturity == "early-growth":
        return "high"
    return "unclear"

def _rt_default_stage_timing_fit(maturity_read, high_agency_exception):
    maturity = _rt_safe_text(maturity_read).lower()
    if maturity == "early-growth":
        return "ideal"
    if maturity == "scale-up":
        return "good" if high_agency_exception else "borderline"
    if maturity in ["late-stage", "public"]:
        return "borderline" if high_agency_exception else "too late"
    return "unclear"

def _rt_bool_text(value):
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    text = _rt_safe_text(value).lower()
    if text in ["true", "yes", "1"]:
        return "TRUE"
    if text in ["false", "no", "0"]:
        return "FALSE"
    return ""

for _idx, _summary_row in summary_df.iterrows():
    _company = _summary_row.get("company", "")
    _parsed = _rt_get_parsed_for_company(_company)
    _role_timing = _parsed.get("role_timing_assessment", {})
    if not isinstance(_role_timing, dict):
        _role_timing = {}

    _raw_text = _rt_get_raw_text_for_company(_company, _parsed)
    _maturity_evidence = _parsed.get("maturity_evidence", {})
    if not isinstance(_maturity_evidence, dict):
        _maturity_evidence = {}
    _maturity_read, _maturity_needs_review = derive_maturity(_maturity_evidence)
    _high_agency_exception = _rt_has_high_agency_exception(_raw_text)
    _agency_level = _rt_safe_text(_role_timing.get("likely_agency_level", "")) or _rt_default_agency_level(_maturity_read, _high_agency_exception)
    _stage_fit = _rt_safe_text(_role_timing.get("stage_timing_fit", "")) or _rt_default_stage_timing_fit(_maturity_read, _high_agency_exception)
    _why_now = _rt_safe_text(_role_timing.get("why_now_or_why_not", ""))
    _penalty_from_model = _rt_bool_text(_role_timing.get("timing_penalty_applied", ""))

    if not _why_now:
        if _maturity_read in ["late-stage", "public"] and not _high_agency_exception:
            _why_now = "Mature company; likely lower agency unless a specific broad mandate exists."
        elif _maturity_read == "early-growth":
            _why_now = "Likely closer to Katelynd's preferred post-PMF / pre-professionalized scaling window."
        elif _maturity_read == "scale-up":
            _why_now = "Potentially relevant, but role scope and remaining operating whitespace need diligence."
        else:
            _why_now = "Timing unclear from public evidence."

    _cap = _rt_operator_timing_cap(_maturity_read, _high_agency_exception)

    _raw_score = summary_df.at[_idx, "operator_timing_score"] if "operator_timing_score" in summary_df.columns else None
    summary_df.at[_idx, "operator_timing_score_raw"] = _raw_score
    summary_df.at[_idx, "operator_timing_score_cap"] = "" if _cap is None else _cap

    _penalty_applied = (_penalty_from_model == "TRUE")

    try:
        _numeric_raw_score = float(_raw_score)
    except Exception:
        _numeric_raw_score = None

    _timing_flag = ""

    if _cap is not None and _numeric_raw_score is not None and _numeric_raw_score > _cap:
        summary_df.at[_idx, "operator_timing_score"] = _cap
        _penalty_applied = True
        _timing_flag = (
            f"CHECK: timing score may be inflated by company quality; "
            f"{_maturity_read} maturity capped operator timing from {int(_numeric_raw_score)} to {_cap}."
        )
    elif _maturity_read in ["late-stage", "public"] and _numeric_raw_score is not None and _numeric_raw_score >= 75:
        _timing_flag = "CHECK: mature company may be too late for high-agency operator entry."

    summary_df.at[_idx, "company_maturity_read"] = _maturity_read
    summary_df.at[_idx, "maturity_needs_review"] = "TRUE" if _maturity_needs_review else "FALSE"
    summary_df.at[_idx, "likely_agency_level"] = _agency_level
    summary_df.at[_idx, "stage_timing_fit"] = _stage_fit
    summary_df.at[_idx, "why_now_or_why_not"] = _why_now
    summary_df.at[_idx, "timing_penalty_applied"] = "TRUE" if _penalty_applied else "FALSE"
    summary_df.at[_idx, "operator_timing_calibration_flag"] = _timing_flag

print("PASS: role/timing maturity assessment and timing caps applied.")


# -----------------------------
# Normalize nested text fields
# -----------------------------

for col in ["commercial_scale_assessment", "final_takeaway"]:
    if col in summary_df.columns:
        summary_df[col] = summary_df[col].apply(normalize_jsonish)

# -----------------------------
# Convert scores to numeric
# -----------------------------

score_cols = [
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]

for col in score_cols:
    if col in summary_df.columns:
        summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce")

# -----------------------------
# Parse warning
# -----------------------------

if "error" in summary_df.columns and summary_df["error"].notna().any():
    print("WARNING: Some rows could not be parsed.")
    display(summary_df[summary_df["error"].notna()])

# -----------------------------
# Required summary columns check
# -----------------------------

required_summary_cols = [
    "company",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "priority_level",
    "business_model_classification",
    "commercial_scale_assessment",
    "commercial_scale_finding"
]

missing_summary_cols = [col for col in required_summary_cols if col not in summary_df.columns]

if missing_summary_cols:
    raise ValueError(f"STOP: summary_df missing required columns: {missing_summary_cols}")

# -----------------------------
# Calibration flags
# -----------------------------

summary_df["calibration_flag"] = summary_df.apply(build_calibration_flag, axis=1)

if "operator_timing_calibration_flag" in summary_df.columns:
    def _append_operator_timing_flag(row):
        base = safe_text(row.get("calibration_flag", ""))
        extra = safe_text(row.get("operator_timing_calibration_flag", ""))

        if not extra:
            return base
        if not base:
            return extra
        if extra in base:
            return base
        return base + " | " + extra

    summary_df["calibration_flag"] = summary_df.apply(_append_operator_timing_flag, axis=1)

# -----------------------------
# Display summary
# -----------------------------

display_cols = [
    "company",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "priority_level",
    "business_model_classification",
    "commercial_scale_assessment",
    "calibration_flag"
]

display_cols = [col for col in display_cols if col in summary_df.columns]

summary_df[display_cols]

# =============================================================================


# -----------------------------
# SCALE-SIGNAL FLATTENING BLOCK
# -----------------------------
# Purpose:
# - Preserve scale_signal_assessment fields from fit_brief_json as flat columns.
# - This supports downstream QA/dashboard fields without changing model scoring.

import json as _scale_signal_json

def _scale_signal_safe_text(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()

def _scale_signal_parse_fit_json(raw_json):
    if _scale_signal_safe_text(raw_json) == "":
        return {}

    try:
        return _scale_signal_json.loads(raw_json)
    except Exception:
        try:
            return parse_first_json_object(raw_json)
        except Exception:
            return {}

def _scale_signal_fields(raw_json):
    parsed_json = _scale_signal_parse_fit_json(raw_json)

    scale = parsed_json.get("scale_signal_assessment", {})

    if not isinstance(scale, dict):
        scale = {}

    return {
        "commercial_scale_signal": scale.get("commercial_scale_signal", ""),
        "commercial_scale_signal_reason": scale.get("commercial_scale_signal_reason", ""),
        "institutional_distribution_signal": scale.get("institutional_distribution_signal", ""),
        "institutional_distribution_signal_reason": scale.get("institutional_distribution_signal_reason", ""),
        "outcomes_signal": scale.get("outcomes_signal", ""),
        "outcomes_signal_reason": scale.get("outcomes_signal_reason", ""),
        "plausible_near_term_scale_path": scale.get("plausible_near_term_scale_path", ""),
        "priority_gate_preliminary_result": scale.get("priority_gate_preliminary_result", ""),
        "priority_gate_reason": scale.get("priority_gate_reason", ""),
        "scale_engine_type": scale.get("scale_engine_type", ""),
        "strong_scale_engine_present": scale.get("strong_scale_engine_present", ""),
    }

if (
    "summary_df" in globals()
    and isinstance(summary_df, pd.DataFrame)
    and "df" in globals()
    and isinstance(df, pd.DataFrame)
    and "fit_brief_json" in df.columns
):
    _scale_signal_lookup = {}

    for _, _raw_row in df.iterrows():
        _company = _scale_signal_safe_text(_raw_row.get("company", ""))

        if _company:
            _scale_signal_lookup[_company] = _scale_signal_fields(
                _raw_row.get("fit_brief_json", "")
            )

    _scale_signal_cols = [
        "commercial_scale_signal",
        "commercial_scale_signal_reason",
        "institutional_distribution_signal",
        "institutional_distribution_signal_reason",
        "outcomes_signal",
        "outcomes_signal_reason",
        "plausible_near_term_scale_path",
        "priority_gate_preliminary_result",
        "priority_gate_reason",
        "scale_engine_type",
        "strong_scale_engine_present",
    ]

    for _col in _scale_signal_cols:
        if _col not in summary_df.columns:
            summary_df[_col] = ""

    for _idx, _summary_row in summary_df.iterrows():
        _company = _scale_signal_safe_text(_summary_row.get("company", ""))

        if _company not in _scale_signal_lookup:
            continue

        for _col, _value in _scale_signal_lookup[_company].items():
            if (
                _scale_signal_safe_text(summary_df.at[_idx, _col]) == ""
                and _scale_signal_safe_text(_value) != ""
            ):
                summary_df.at[_idx, _col] = _value

    print("PASS: scale-signal flattening applied.")
else:
    print("WARNING: scale-signal flattening skipped; summary_df/df/fit_brief_json unavailable.")

# -----------------------------
# SLICE 2 STRUCTURED-EVIDENCE BLOCK
# -----------------------------
# Flatten maturity_evidence + commercial_evidence from fit_brief_json into persisted
# columns, and derive the commercial signal deterministically from them (package:
# structured_evidence). company_maturity_read / maturity_needs_review are set in the
# role/timing block above via derive_maturity. Persisting the COMPONENTS lets the
# labels/signals be recalibrated later WITHOUT re-research (spec Part C).

from health_tech_research_agent.structured_evidence import (
    derive_commercial_signal,
    commercial_signal_to_text,
    flatten_slice2_fields,
    derive_reset_signal,
    reset_basis_for,
    reset_needs_review,
    flatten_reset_fields,
    MATURITY_EVIDENCE_FIELDS,
    COMMERCIAL_EVIDENCE_FIELDS,
    RESET_PERSIST_FIELDS,
)

if (
    "summary_df" in globals()
    and isinstance(summary_df, pd.DataFrame)
    and "df" in globals()
    and isinstance(df, pd.DataFrame)
    and "fit_brief_json" in df.columns
):
    _slice2_cols = (
        list(MATURITY_EVIDENCE_FIELDS)
        + list(COMMERCIAL_EVIDENCE_FIELDS)
        + ["commercial_scale_signal", "commercial_scale_signal_inferred"]
        + list(RESET_PERSIST_FIELDS)
        + ["reset_or_restructure_signal", "reset_or_restructure_basis", "reset_needs_review"]
    )
    for _col in _slice2_cols:
        if _col not in summary_df.columns:
            summary_df[_col] = ""

    _slice2_lookup = {}
    for _, _raw_row in df.iterrows():
        _company = _scale_signal_safe_text(_raw_row.get("company", ""))
        if not _company:
            continue
        _parsed_s2 = _scale_signal_parse_fit_json(_raw_row.get("fit_brief_json", ""))
        _flat = flatten_slice2_fields(_parsed_s2)
        _commercial_evidence = _parsed_s2.get("commercial_evidence", {}) if isinstance(_parsed_s2, dict) else {}
        if not isinstance(_commercial_evidence, dict):
            _commercial_evidence = {}
        _signal_int = derive_commercial_signal(_commercial_evidence)
        _flat["commercial_scale_signal_inferred"] = _signal_int
        _flat["commercial_scale_signal"] = commercial_signal_to_text(_signal_int)
        # Slice 3.5: reset/restructure — flatten the per-event list (reset_events_json +
        # reset_event_types) and derive the signal PER EVENT. strategic-pivot / ma-integration
        # never fire (recorded but not scored); an unrecognized event type is surfaced via
        # reset_needs_review rather than fired blind.
        _flat.update(flatten_reset_fields(_parsed_s2))
        _reset_evidence = _parsed_s2.get("reset_evidence", {}) if isinstance(_parsed_s2, dict) else {}
        if not isinstance(_reset_evidence, dict):
            _reset_evidence = {}
        _flat["reset_or_restructure_signal"] = derive_reset_signal(_reset_evidence)
        _flat["reset_or_restructure_basis"] = reset_basis_for(_reset_evidence)
        _flat["reset_needs_review"] = reset_needs_review(_reset_evidence)
        _slice2_lookup[_company] = _flat

    for _idx, _summary_row in summary_df.iterrows():
        _company = _scale_signal_safe_text(_summary_row.get("company", ""))
        if _company not in _slice2_lookup:
            continue
        for _col, _value in _slice2_lookup[_company].items():
            summary_df.at[_idx, _col] = _value  # authoritative derived values

    print("PASS: Slice 2/3 structured-evidence flatten + commercial-signal + reset derivation applied.")
else:
    print("WARNING: Slice 2 structured-evidence block skipped; summary_df/df/fit_brief_json unavailable.")

# =============================================================================
# COLAB VERIFICATION CHECKLIST - Slice 2 (structured evidence) notebook wiring
# =============================================================================
# Offline tests cover the package functions; this wiring is Colab-verified. After
# pulling this branch, research ONE real company (small batch) through STEP 7 -> 10 ->
# 10A, then check, in order:
# 1. New component columns exist + non-empty on a real company:
#    cols = MATURITY_EVIDENCE_FIELDS + COMMERCIAL_EVIDENCE_FIELDS
#    EXPECT summary_df[cols] populated (funding_stage, q1..q4, etc.) for the company.
# 2. Derived maturity is from funding stage, NOT a text-scan:
#    EXPECT company_maturity_read matches the derive_maturity rule for its funding_stage
#    (e.g. a Series B with big revenue -> "early-growth", never "late-stage"); and
#    maturity_needs_review == "TRUE" only when funding_stage is unknown.
# 3. Derived commercial signal:
#    EXPECT commercial_scale_signal (text) + commercial_scale_signal_inferred (0-3) match
#    derive_commercial_signal on the company's commercial_evidence (funding-dependent -> not strong).
# 4. Retired inferences are gone:
#    EXPECT no _rt_infer_maturity_read / infer_commercial_signal in the live module
#    namespace (NameError if referenced); no stale text-scan maturity.
# 5. STEP 26 still rescores with derived maturity:
#    Run STEP 26 (DRY_RUN=False) on one company w/ archived evidence -> EXPECT it completes
#    and cap_info["company_maturity_read"] equals derive_maturity for that company.
# 6. New columns land on the master via STEP 12:
#    After the master update, EXPECT the master CSV has the 8 maturity + 10 commercial
#    component columns + maturity_needs_review + commercial_scale_signal_inferred.
# =============================================================================

# =============================================================================
# COLAB VERIFICATION CHECKLIST - Slice 3.5 (multi-event reset) notebook wiring
# =============================================================================
# After researching ONE real company through STEP 7 -> 10, check, in order:
# 1. The reset_events LIST populates from real output:
#    parsed["reset_evidence"]["reset_events"] is a list of {event_type, basis,
#    creates_high_agency_opening}; summary_df has reset_events_json (the full list) +
#    reset_event_types (comma-joined). Empty list when no event is found.
# 2. ZOE comes back with TWO events and fires on the restructuring:
#    EXPECT ZOE's reset_event_types includes BOTH strategic-pivot AND restructuring-layoffs,
#    and reset_or_restructure_signal == True (the restructuring's opening=yes fires; the louder
#    pivot no longer buries it). EXPECT it == derive_reset_signal(reset_evidence).
#    (Noom -> [strategic-pivot] only -> False, unchanged.)
# 3. The MATERIALIZED signal is what the engine reads:
#    from health_tech_research_agent.candidate_priority import reset_signal
#    EXPECT reset_signal(row) (reads reset_or_restructure_signal) == the derived bool.
# 4. Scan reset_needs_review:
#    EXPECT reset_needs_review == True only where an event_type is outside the seven recognized
#    types — those rows are findable (not fired blind, not silently dropped).
# 5. Columns land on the master via STEP 12:
#    After the master update, EXPECT the master CSV has reset_events_json, reset_event_types,
#    reset_or_restructure_signal, reset_or_restructure_basis, reset_needs_review.
# =============================================================================

# =============================================================================
# COLAB VERIFICATION CHECKLIST - Slice 3.7 (search-layer redesign) notebook wiring
# =============================================================================
# After pulling this branch, run ONE real company through STEP 7 (fresh research),
# then STEP 10, then a STEP 26 rescore. Check, in order:
#
# 1. STEP 4 exposes the two new operator search shims:
#    EXPECT callable(search_org_events) and callable(search_operating_characteristics) -> True.
#
# 2. STEP 7 runs SIX web searches/company and persists a NINE-column checkpoint:
#    EXPECT the checkpoint columns ==
#      ["company","date_researched","funding_finding","payer_institutional_finding",
#       "outcomes_finding","commercial_scale_finding","org_events_finding",
#       "operating_characteristics_finding","fit_brief_json"]   (schema + order),
#    and org_events_finding / operating_characteristics_finding are POPULATED (non-blank)
#    for the researched company -- the two new searches returned evidence. (A genuine
#    "No qualifying recent org/leadership events found." is itself a valid populated result.)
#
# 3. Multi-event reset now SURFACES on ZOE (the Slice 3.5 finding that motivated 3.7):
#    Research ZOE (a strategic-pivot AND a restructuring-toward-expansion). EXPECT
#    org_events_finding lists BOTH events, and after STEP 10 reset_or_restructure_signal
#    == True (fires on the restructuring's opening even alongside the pivot). Under pre-3.7
#    research ZOE returned an EMPTY events list. Also scan reset_needs_review == TRUE rows.
#
# 4. STEP 26 rescore uses the new evidence:
#    Run STEP 26 (DRY_RUN) on a company with archived org/operating findings. EXPECT
#    step26_build_status_findings includes the "Recent org / leadership events" and
#    "Operating characteristics" sections, and the rescore completes with no NameError.
#
# 5. Capability A1/A2/A3 are NOT scored yet (that is Slice 4): the fit brief gathers the
#    operating-characteristics evidence but emits no capability_a*_score fields.
# =============================================================================

# STEP 10A - Deterministic priority adjudication

# =============================================================================

# Purpose:

# - Enforce hard priority rules after LLM scoring

# - Prevent role fit / thesis interest from incorrectly promoting weak-scale companies

# - Update fit_brief_json, checkpoint, and archive rows in place

# =============================================================================
# STEP 10A - Deterministic priority adjudication
# =============================================================================
# Purpose:
# - Enforce hard priority rules after LLM scoring
# - Use native P0/P1/P2/P3/P4 priority labels
# - Prevent role fit / thesis interest from incorrectly promoting weak-scale companies
# - Treat incorrect outputs as decision-logic calibration data
# - Update fit_brief_json in df, checkpoint, and current batch raw/archive rows
#
# Required flow:
# 1. Run Step 10
# 2. Run this Step 10A cell
# 3. Rerun Step 10
# 4. Run Step 10B

import json
import re
import pandas as pd
import shutil
from pathlib import Path

# -----------------------------
# Safety checks
# -----------------------------

if "df" not in globals() or not isinstance(df, pd.DataFrame) or df.empty:
    raise NameError("STOP: df is not available or empty.")

if "BATCH_NAME" not in globals():
    raise NameError("STOP: BATCH_NAME is not defined.")

if "batch_checkpoint_path" not in globals() or "drive_checkpoint_path" not in globals():
    raise NameError("STOP: checkpoint paths are not defined.")

required_current_schema_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "org_events_finding",
    "operating_characteristics_finding",
    "fit_brief_json"
]

missing_cols = [col for col in required_current_schema_cols if col not in df.columns]

if missing_cols:
    raise ValueError(f"STOP: df missing required columns: {missing_cols}")

# -----------------------------
# JSON helpers
# -----------------------------

def clean_json_text(text):
    text = str(text).strip()
    text = re.sub(r"^```json", "", text)
    text = re.sub(r"^```", "", text)
    text = re.sub(r"```$", "", text)
    return text.strip()

def parse_first_json_object(text):
    raw = clean_json_text(text)
    start = raw.find("{")

    if start == -1:
        raise ValueError("No JSON object found")

    decoder = json.JSONDecoder()
    parsed, end = decoder.raw_decode(raw[start:])
    return parsed

def get_score(parsed, key):
    scores = parsed.get("scores", {})
    value = scores.get(key)

    if isinstance(value, dict):
        return value.get("score")

    return value

def safe_int(value, default=0):
    try:
        if pd.isna(value):
            return default
        return int(float(value))
    except Exception:
        return default

def normalize_signal(value):
    text = str(value).strip().lower()

    if text in ["strong", "moderate", "weak", "none"]:
        return text

    if "strong" in text:
        return "strong"
    if "moderate" in text or "medium" in text:
        return "moderate"
    if "weak" in text:
        return "weak"
    if "none" in text or text in ["no", "n/a", "na", ""]:
        return "none"

    return "none"

def text_contains_any(text, terms):
    text = str(text).lower()
    return any(term.lower() in text for term in terms)

def bool_from_value(value, default=False):
    if isinstance(value, bool):
        return value

    text = str(value).strip().lower()

    if text in ["true", "yes", "1"]:
        return True

    if text in ["false", "no", "0"]:
        return False

    return default

def append_flag(existing, new_flag):
    existing = str(existing or "").strip()

    if not existing:
        return new_flag

    if new_flag in existing:
        return existing

    return existing + " | " + new_flag

def priority_code(value):
    text = str(value or "").upper()
    match = re.search(r"\bP[0-4]\b", text)
    return match.group(0) if match else ""

def normalize_priority_label(value):
    text = str(value or "").strip()
    lower = text.lower()

    if lower.startswith("p0") or "highest-priority" in lower or "highest priority" in lower:
        return "P0: Highest-priority target"

    if (
        lower.startswith("p1: near-priority")
        or lower.startswith("p1: near priority")
        or "near-priority" in lower
        or "near priority" in lower
        or "p1-border" in lower
        or "p1 border" in lower
        or "strong p2" in lower
    ):
        return "P1: High-priority diligence"

    # Backward compatibility: old P1 means new P0.
    if lower.startswith("p1: high-priority") or lower.startswith("p1: high priority"):
        return "P0: Highest-priority target"

    if lower.startswith("p2") or "review p2" in lower or "worth deeper diligence" in lower:
        return "P2: Worth deeper diligence"

    if lower.startswith("p3") or "watch list" in lower or "watchlist" in lower:
        return "P3: Watch list"

    if lower.startswith("p4") or "low priority" in lower or "likely reject" in lower or "reject" in lower:
        return "P4: Low priority / likely reject"

    return text

def recommendation_for_priority(priority_level):
    code = priority_code(priority_level)

    return {
        "P0": "Strong fit, active pursuit",
        "P1": "Strong fit, near-priority diligence",
        "P2": "Possible fit, pending diligence",
        "P3": "Watch list",
        "P4": "Weak fit"
    }.get(code, "Watch list")

# -----------------------------
# Signal inference helpers
# -----------------------------

def get_scale_assessment(parsed):
    scale_assessment = parsed.get("scale_signal_assessment", {})

    if isinstance(scale_assessment, dict):
        return scale_assessment

    return {}

# infer_commercial_signal RETIRED (Slice 2): the commercial signal is now derived
# deterministically from the researched commercial facts + four red-flags by
# structured_evidence.derive_commercial_signal (funding structurally excluded; the old
# dollar-token strong_markers, which could not tell revenue from funding, are gone).
from health_tech_research_agent.structured_evidence import (
    derive_commercial_signal,
    commercial_signal_to_text,
)

def infer_institutional_signal(row, parsed):
    scale_assessment = get_scale_assessment(parsed)
    explicit = normalize_signal(scale_assessment.get("institutional_distribution_signal", ""))

    if explicit in ["strong", "moderate", "weak", "none"] and str(scale_assessment.get("institutional_distribution_signal", "")).strip():
        return explicit

    text = " ".join([
        str(row.get("payer_institutional_finding", "")),
        str(parsed.get("pmf_scale_assessment", "")),
        str(parsed.get("business_model_classification", "")),
    ]).lower()

    weak_markers = [
        "no strong public institutional",
        "no strong institutional signal",
        "mostly d2c",
        "mostly d2c/cash-pay",
        "no public payer contracts",
        "no clear payer contracts",
        "no medicare",
        "no medicaid",
        "not broad payer",
        "not broad institutional",
        "does not surface clear payer contracts",
        "no clear evidence of payer contracts"
    ]

    strong_markers = [
        "strong institutional",
        "top health plans",
        "health plans and employers",
        "employer/health-plan",
        "employers/health plans",
        "covered lives",
        "health-plan channels",
        "enterprise clients",
        "employer customers",
        "lives covered",
        "payer contracts",
        "provider network",
        "health system"
    ]

    moderate_markers = [
        "some institutional",
        "limited partner",
        "employer-facing",
        "provider-facing",
        "b2b2c",
        "partnership",
        "employer page",
        "provider referral",
        "benefits-consultant",
        "one named plan partnership",
        "emerging payer-linked distribution"
    ]

    if text_contains_any(text, weak_markers):
        return "weak"

    if text_contains_any(text, strong_markers):
        return "strong"

    if text_contains_any(text, moderate_markers):
        return "moderate"

    return "none"

def infer_outcomes_signal(row, parsed):
    scale_assessment = get_scale_assessment(parsed)
    explicit = normalize_signal(scale_assessment.get("outcomes_signal", ""))

    if explicit in ["strong", "moderate", "weak", "none"] and str(scale_assessment.get("outcomes_signal", "")).strip():
        return explicit

    text = " ".join([
        str(row.get("outcomes_finding", "")),
        str(parsed.get("pmf_scale_assessment", "")),
    ]).lower()

    weak_markers = [
        "no strong public outcomes",
        "no peer-reviewed",
        "no clinical trials",
        "not a clinical trial",
        "no control group",
        "mostly marketing",
        "limited public evidence",
        "no meaningful outcomes",
        "not independently validated outcomes"
    ]

    strong_markers = [
        "randomized controlled trial",
        "nature medicine",
        "peer-reviewed",
        "published real-world evidence",
        "claims-based study",
        "clinical efficacy",
        "real-world evidence",
        "health improvement",
        "reduced costs",
        "improved outcomes"
    ]

    moderate_markers = [
        "mixed",
        "weak-to-moderate",
        "moderate evidence",
        "company-reported",
        "retrospective",
        "engagement-linked",
        "validation studies",
        "measurement studies",
        "behavior-change"
    ]

    if text_contains_any(text, weak_markers):
        return "weak"

    if text_contains_any(text, strong_markers):
        return "strong"

    if text_contains_any(text, moderate_markers):
        return "moderate"

    return "none"

# -----------------------------
# Deterministic adjudication logic
# -----------------------------

adjudication_rows = []

for idx, row in df.iterrows():
    company = str(row["company"]).strip()
    parsed = parse_first_json_object(row["fit_brief_json"])

    thesis = safe_int(get_score(parsed, "thesis_fit_score"))
    pmf = safe_int(get_score(parsed, "pmf_scale_score"))
    evidence = safe_int(get_score(parsed, "evidence_confidence_score"))
    role_fit = safe_int(get_score(parsed, "katelynd_role_fit_score"))
    timing = safe_int(get_score(parsed, "operator_timing_score"))

    scale_assessment = get_scale_assessment(parsed)

    _commercial_evidence = parsed.get("commercial_evidence", {})
    if not isinstance(_commercial_evidence, dict):
        _commercial_evidence = {}
    commercial_signal = commercial_signal_to_text(derive_commercial_signal(_commercial_evidence))
    institutional_signal = infer_institutional_signal(row, parsed)
    outcomes_signal = infer_outcomes_signal(row, parsed)

    plausible_near_term_scale_path = bool_from_value(
        scale_assessment.get("plausible_near_term_scale_path"),
        default=False
    )

    if not plausible_near_term_scale_path:
        plausible_near_term_scale_path = (
            commercial_signal == "strong" or
            institutional_signal == "strong" or
            (
                outcomes_signal == "strong" and
                (
                    commercial_signal in ["moderate", "strong"] or
                    institutional_signal in ["moderate", "strong"]
                )
            )
        )

    commercial_strong_gate = commercial_signal == "strong" and evidence >= 50
    institutional_strong_gate = institutional_signal == "strong" and evidence >= 50

    outcomes_plus_scale_path_gate = (
        outcomes_signal == "strong" and
        plausible_near_term_scale_path is True and
        evidence >= 55 and
        pmf >= 70 and
        (
            commercial_signal in ["moderate", "strong"] or
            institutional_signal in ["moderate", "strong"]
        )
    )

    score_qualifies_for_p2 = pmf >= 70 and evidence >= 50

    qualifies_for_p2 = (
        score_qualifies_for_p2 or
        commercial_strong_gate or
        institutional_strong_gate or
        outcomes_plus_scale_path_gate
    )

    strong_signal_count = sum([
        commercial_signal == "strong",
        institutional_signal == "strong",
        outcomes_signal == "strong"
    ])

    moderate_or_strong_signal_count = sum([
        commercial_signal in ["moderate", "strong"],
        institutional_signal in ["moderate", "strong"],
        outcomes_signal in ["moderate", "strong"]
    ])

    qualifies_for_p1 = (
        qualifies_for_p2 and
        thesis >= 85 and
        pmf >= 74 and
        evidence >= 55 and
        role_fit >= 78 and
        timing >= 72 and
        (
            strong_signal_count >= 1 or
            moderate_or_strong_signal_count >= 2
        )
    )

    qualifies_for_p0 = (
        qualifies_for_p1 and
        thesis >= 88 and
        pmf >= 82 and
        evidence >= 75 and
        role_fit >= 80 and
        timing >= 78 and
        (
            strong_signal_count >= 3 or
            (
                institutional_signal == "strong" and
                outcomes_signal in ["moderate", "strong"] and
                commercial_signal in ["moderate", "strong"]
            ) or
            (
                commercial_signal == "strong" and
                institutional_signal == "strong"
            )
        )
    )


    qualifies_for_p4 = (
        not qualifies_for_p2 and
        (
            thesis < 55 or
            pmf < 45 or
            role_fit < 55 or
            timing < 55
        ) and
        commercial_signal in ["weak", "none"] and
        institutional_signal in ["weak", "none"]
    )

    original_priority = normalize_priority_label(parsed.get("priority_level", ""))
    original_flag = str(parsed.get("calibration_flag", "") or "").strip()

    if qualifies_for_p0:
        adjudicated_priority = "P0: Highest-priority target"
        adjudication_action = "assigned_P0"

    elif qualifies_for_p1:
        adjudicated_priority = "P1: High-priority diligence"
        adjudication_action = "assigned_P1"

    elif qualifies_for_p2:
        adjudicated_priority = "P2: Worth deeper diligence"
        adjudication_action = "assigned_P2"

    elif qualifies_for_p4:
        adjudicated_priority = "P4: Low priority / likely reject"
        adjudication_action = "assigned_P4"

    else:
        adjudicated_priority = "P3: Watch list"
        adjudication_action = "assigned_P3"

    adjudicated_recommendation = recommendation_for_priority(adjudicated_priority)
    adjudicated_flag = original_flag

    if original_priority != "" and original_priority != adjudicated_priority:
        adjudicated_flag = append_flag(
            adjudicated_flag,
            f"Deterministic gate: reassigned from {original_priority} to {adjudicated_priority}."
        )

    if adjudicated_priority.startswith("P1"):
        adjudicated_flag = append_flag(
            adjudicated_flag,
            "Deterministic gate: P1 indicates near-priority / former P1-border; validate before active pursuit."
        )

    if adjudicated_priority.startswith("P2") and (evidence < 65 or pmf < 70):
        adjudicated_flag = append_flag(
            adjudicated_flag,
            "Deterministic gate: P2 has moderate evidence and/or PMF; diligence required."
        )

    parsed["priority_adjudication"] = {
        "decision_logic_version": "p0_p4_scale_engine_gate_v2",
        "commercial_scale_signal": commercial_signal,
        "institutional_distribution_signal": institutional_signal,
        "outcomes_signal": outcomes_signal,
        "plausible_near_term_scale_path": plausible_near_term_scale_path,
        "strong_signal_count": strong_signal_count,
        "moderate_or_strong_signal_count": moderate_or_strong_signal_count,
        "score_qualifies_for_p2": score_qualifies_for_p2,
        "commercial_strong_gate": commercial_strong_gate,
        "institutional_strong_gate": institutional_strong_gate,
        "outcomes_plus_scale_path_gate": outcomes_plus_scale_path_gate,
        "qualifies_for_p2": qualifies_for_p2,
        "qualifies_for_p1": qualifies_for_p1,
        "qualifies_for_p0": qualifies_for_p0,
        "qualifies_for_p4": qualifies_for_p4,
        "original_priority_level": original_priority,
        "adjudicated_priority_level": adjudicated_priority,
        "adjudication_action": adjudication_action
    }

    parsed["priority_level"] = adjudicated_priority
    parsed["final_recommendation"] = adjudicated_recommendation
    parsed["calibration_flag"] = adjudicated_flag

    if adjudicated_priority.startswith("P3"):
        parsed["final_takeaway"] = (
            f"{company} remains strategically interesting, but the current public evidence does not clear the P2 priority gate. "
            "Keep on watch list until stronger commercial traction, institutional distribution, or outcomes-plus-scale evidence emerges."
        )

    elif adjudicated_priority.startswith("P4"):
        parsed["final_takeaway"] = (
            f"{company} does not currently clear the job-search priority gate based on available evidence. "
            "Revisit only if stronger scale, outcomes, or role-fit evidence emerges."
        )

    df.loc[idx, "fit_brief_json"] = json.dumps(parsed, indent=2, ensure_ascii=False)

    adjudication_rows.append({
        "company": company,
        "thesis_fit_score": thesis,
        "pmf_scale_score": pmf,
        "evidence_confidence_score": evidence,
        "katelynd_role_fit_score": role_fit,
        "operator_timing_score": timing,
        "commercial_scale_signal": commercial_signal,
        "institutional_distribution_signal": institutional_signal,
        "outcomes_signal": outcomes_signal,
        "plausible_near_term_scale_path": plausible_near_term_scale_path,
        "strong_signal_count": strong_signal_count,
        "moderate_or_strong_signal_count": moderate_or_strong_signal_count,
        "score_qualifies_for_p2": score_qualifies_for_p2,
        "commercial_strong_gate": commercial_strong_gate,
        "institutional_strong_gate": institutional_strong_gate,
        "outcomes_plus_scale_path_gate": outcomes_plus_scale_path_gate,
        "qualifies_for_p2": qualifies_for_p2,
        "qualifies_for_p1": qualifies_for_p1,
        "qualifies_for_p0": qualifies_for_p0,
        "qualifies_for_p4": qualifies_for_p4,
        "original_priority": original_priority,
        "adjudicated_priority": adjudicated_priority,
        "adjudication_action": adjudication_action
    })

adjudication_df = pd.DataFrame(adjudication_rows)

# -----------------------------
# Save corrected checkpoint
# -----------------------------

df = df[required_current_schema_cols].copy().reset_index(drop=True)
checkpoint_df = df.copy()

Path(batch_checkpoint_path).parent.mkdir(parents=True, exist_ok=True)
df.to_csv(batch_checkpoint_path, index=False)

Path(drive_checkpoint_path).parent.mkdir(parents=True, exist_ok=True)
shutil.copy(batch_checkpoint_path, drive_checkpoint_path)

print("PASS: Deterministic priority adjudication applied.")
print("Local checkpoint:", batch_checkpoint_path)
print("Drive checkpoint:", drive_checkpoint_path)

print("\nAdjudication summary:")
display(adjudication_df)

# -----------------------------
# Update already-written raw/archive rows in place
# -----------------------------

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_batches_folder = drive_folder / "research_batches"
local_batches_folder = Path("research_batches")

archive_paths = [
    Path("health_tech_raw_research_ARCHIVE.csv"),
    drive_folder / "health_tech_raw_research_ARCHIVE.csv"
]

archive_paths.extend(sorted(local_batches_folder.glob(f"{BATCH_NAME}_raw_*.csv")))
archive_paths.extend(sorted(drive_batches_folder.glob(f"{BATCH_NAME}_raw_*.csv")))

updated_files = []

for archive_path in archive_paths:
    archive_path = Path(archive_path)

    if not archive_path.exists():
        continue

    archive_df = pd.read_csv(archive_path)

    if "batch_name" not in archive_df.columns or "company" not in archive_df.columns:
        continue

    batch_mask = archive_df["batch_name"].astype(str).eq(BATCH_NAME)

    if not batch_mask.any():
        continue

    for _, source_row in df.iterrows():
        company = source_row["company"]
        row_mask = batch_mask & archive_df["company"].astype(str).eq(company)

        if not row_mask.any():
            continue

        for col in required_current_schema_cols:
            if col in archive_df.columns:
                archive_df.loc[row_mask, col] = source_row[col]

    archive_df.to_csv(archive_path, index=False)
    updated_files.append(str(archive_path))

print("\nUpdated archive/current-batch files:")
for path in updated_files:
    print("-", path)

print("\nNext: rerun Step 10, then Step 10B.")


# =============================================================================

# STEP 10B - Batch QA checks before export/master update

# =============================================================================

# 10B - Batch QA checks before export / master update
# Purpose:
# - Check current batch raw fields
# - Check obvious wrong-company contamination
# - Surface priority/evidence calibration issues
# - Use narrower D2C detection so hybrid B2B2C/payer/employer companies are not incorrectly flagged as D2C

import pandas as pd

print("Running batch QA checks...")
print("df shape:", df.shape)
print("companies:", df["company"].tolist())

qa_flags = []

# -----------------------------
# Helpers
# -----------------------------

def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def is_d2c_specific_business_model(business_model_text):
    """
    Detects true D2C / cash-pay companies without over-flagging
    hybrid payer/employer/provider-enabled consumer platforms.
    """
    business_model = safe_text(business_model_text).lower()

    institutional_terms = [
        "b2b2c",
        "payer",
        "employer",
        "provider",
        "benefits",
        "insurance",
        "health plan",
        "government",
        "pharma",
        "medicaid",
        "medicare"
    ]

    has_institutional_path = any(term in business_model for term in institutional_terms)

    appears_d2c = (
        "d2c" in business_model
        or "cash-pay" in business_model
        or "cash pay" in business_model
        or "primarily d2c" in business_model
        or "primarily cash" in business_model
    )

    # Do not treat generic "consumer health platform" as D2C if the classification
    # is clearly hybrid/B2B2C/payer/employer/provider-enabled.
    d2c_specific_review_needed = appears_d2c and not (
        has_institutional_path and "hybrid" in business_model
    )

    return d2c_specific_review_needed

def commercial_signal_is_weak(commercial_text):
    text = safe_text(commercial_text).lower()

    weak_phrases = [
        "no strong public commercial scale evidence found",
        "weak evidence",
        "commercial scale evidence is weak",
        "weak-to-moderate",
        "not enough hard evidence",
        "not strongly substantiated",
        "not revenue quality",
        "no credible public revenue",
        "no public revenue",
        "no verified current arr",
        "no public arr",
        "no subscriber count",
        "no retention",
        "no renewal",
        "no cac",
        "no margin"
    ]

    if text == "":
        return True

    return any(phrase in text for phrase in weak_phrases)

def add_flag(company, issue):
    qa_flags.append({
        "company": company,
        "issue": issue
    })

# -----------------------------
# 1. Check for blank raw research fields
# -----------------------------

raw_fields = [
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "fit_brief_json"
]

missing_raw_fields = [field for field in raw_fields if field not in df.columns]

if missing_raw_fields:
    for field in missing_raw_fields:
        add_flag("BATCH", f"Missing raw field column: {field}")
else:
    for _, row in df.iterrows():
        company = row["company"]

        for field in raw_fields:
            value = row.get(field, "")

            if pd.isna(value) or str(value).strip() == "":
                add_flag(company, f"Blank field: {field}")

# -----------------------------
# 2. Check for likely wrong-company contamination
# -----------------------------

suspicious_terms = [
    "wealth management",
    "edtech",
    "unrelated",
    "different company"
]

for _, row in df.iterrows():
    company = row["company"]

    combined_text = " ".join([
        str(row.get("funding_finding", "")),
        str(row.get("payer_institutional_finding", "")),
        str(row.get("outcomes_finding", "")),
        str(row.get("commercial_scale_finding", "")),
        str(row.get("fit_brief_json", ""))
    ]).lower()

    for term in suspicious_terms:
        if term in combined_text:
            add_flag(company, f"Possible wrong-company contamination: '{term}'")

# -----------------------------
# 3. Check P2s with moderate/weak evidence
# -----------------------------

if "summary_df" in globals():
    for _, row in summary_df.iterrows():
        company = row["company"]
        priority = safe_text(row.get("priority_level"))
        evidence = row.get("evidence_confidence_score", 0)
        pmf = row.get("pmf_scale_score", 0)

        if priority == "P2: Worth deeper diligence" and (evidence < 65 or pmf < 70):
            add_flag(company, "Review P2: moderate evidence and/or PMF")

# -----------------------------
# 4. D2C commercial-scale sanity check
# -----------------------------

if "summary_df" in globals():
    for _, row in summary_df.iterrows():
        company = row["company"]
        business_model = row.get("business_model_classification", "")
        priority = safe_text(row.get("priority_level"))
        evidence = row.get("evidence_confidence_score", 0)
        pmf = row.get("pmf_scale_score", 0)

        raw_match = df[df["company"] == company]

        if not raw_match.empty:
            commercial_text = str(raw_match.iloc[0].get("commercial_scale_finding", ""))
        else:
            commercial_text = ""

        d2c_specific_review_needed = is_d2c_specific_business_model(business_model)
        weak_commercial_signal = commercial_signal_is_weak(commercial_text)

        if (
            d2c_specific_review_needed
            and priority in ["P1: High-priority target", "P2: Worth deeper diligence"]
            and weak_commercial_signal
        ):
            add_flag(company, "Review D2C priority: weak commercial-scale evidence")

        if (
            d2c_specific_review_needed
            and pmf >= 75
            and evidence < 60
        ):
            add_flag(company, "Review D2C scale claim: strong PMF score but moderate/weak evidence confidence")

# -----------------------------
# 5. Surface calibration flags from Step 10
# -----------------------------

if "summary_df" in globals() and "calibration_flag" in summary_df.columns:
    flagged_rows = summary_df[
        summary_df["calibration_flag"].notna()
        & (summary_df["calibration_flag"].astype(str).str.strip() != "")
    ]

    for _, row in flagged_rows.iterrows():
        add_flag(row["company"], f"Calibration flag: {row['calibration_flag']}")

# -----------------------------
# 6. Dedupe QA flags
# -----------------------------

qa_df = pd.DataFrame(qa_flags)

if not qa_df.empty:
    qa_df = qa_df.drop_duplicates().reset_index(drop=True)

if qa_df.empty:
    print("QA passed. No issues flagged.")
else:
    print("QA flags found. Review before updating master.")
    display(qa_df)

# =============================================================================

# STEP 11 - Export current batch only

# =============================================================================

# Step 11 - Export current batch only
# Purpose:
# - Export raw current-batch research
# - Export parsed current-batch summary
# - Confirm commercial-scale fields are included

print("Exporting CURRENT BATCH ONLY.")
print("Raw df shape:", df.shape)
print("Summary df shape:", summary_df.shape)

required_raw_export_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "fit_brief_json"
]

required_summary_export_cols = [
    "company",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "priority_level",
    "business_model_classification",
    "commercial_scale_assessment",
    "commercial_scale_finding",
    "calibration_flag"
]

missing_raw_cols = [col for col in required_raw_export_cols if col not in df.columns]
missing_summary_cols = [col for col in required_summary_export_cols if col not in summary_df.columns]

if missing_raw_cols:
    raise ValueError(f"STOP: Raw df missing export columns: {missing_raw_cols}")

if missing_summary_cols:
    raise ValueError(f"STOP: summary_df missing export columns: {missing_summary_cols}")

df.to_csv(batch_raw_export_path, index=False)
summary_df.to_csv(batch_summary_export_path, index=False)

print("Saved batch raw export to:")
print(batch_raw_export_path)

print("Saved batch summary export to:")
print(batch_summary_export_path)

print("\nCommercial-scale fields confirmed in exports:")
print("- df includes commercial_scale_finding")
print("- summary_df includes commercial_scale_finding and commercial_scale_assessment")

# =============================================================================

# STEP 11A - Final raw archive QA

# =============================================================================

# 11A - Final raw archive QA
# Purpose:
# - Confirm raw archive exists and is structurally clean
# - Confirm core raw fields are complete
# - Confirm fit_brief_json parses
# - Confirm no duplicate company+batch records
# - Support commercial_scale_finding while staying backward-compatible with older archived rows

import pandas as pd
import json
from pathlib import Path
from google.colab import drive
import shutil

drive.mount("/content/drive")

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_archive_path = drive_folder / "health_tech_raw_research_ARCHIVE.csv"
local_archive_path = Path("health_tech_raw_research_ARCHIVE.csv")

if not drive_archive_path.exists():
    raise FileNotFoundError(f"Archive not found in Drive: {drive_archive_path}")

shutil.copy(drive_archive_path, local_archive_path)

archive_df = pd.read_csv(local_archive_path)

# -----------------------------
# Required columns
# -----------------------------

core_required_cols = [
    "company",
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "fit_brief_json",
    "batch_name",
    "archive_saved_at",
    "archive_record_id"
]

# New field added during commercial-scale upgrade.
# Older archived rows may be blank, but all new rows should have it.
commercial_scale_col = "commercial_scale_finding"

all_expected_cols = core_required_cols + [commercial_scale_col]

print("ARCHIVE SHAPE:", archive_df.shape)
print("COMPANY COUNT:", archive_df["company"].nunique())

missing_core_cols = [col for col in core_required_cols if col not in archive_df.columns]
missing_expected_cols = [col for col in all_expected_cols if col not in archive_df.columns]

print("MISSING CORE COLUMNS:", missing_core_cols)
print("MISSING EXPECTED COLUMNS:", missing_expected_cols)

if missing_core_cols:
    raise ValueError(f"STOP: Archive is missing core required columns: {missing_core_cols}")

# Backward-compatible schema repair in memory only.
# This does not save the archive; it just allows QA to run cleanly.
if commercial_scale_col not in archive_df.columns:
    archive_df[commercial_scale_col] = ""
    print("NOTE: commercial_scale_finding column was missing and added in memory for QA compatibility.")

# -----------------------------
# Blank-field check: core fields
# -----------------------------

blank_issues = []

for col in core_required_cols:
    blank_mask = archive_df[col].isna() | (archive_df[col].astype(str).str.strip() == "")
    for company in archive_df.loc[blank_mask, "company"].tolist():
        blank_issues.append({
            "company": company,
            "blank_field": col,
            "severity": "FAIL"
        })

blank_df = pd.DataFrame(blank_issues)

if blank_df.empty:
    print("CORE BLANK FIELD CHECK: PASS")
else:
    print("CORE BLANK FIELD CHECK: FAIL")
    display(blank_df)

# -----------------------------
# Commercial-scale field check
# -----------------------------
# Older rows may have blank commercial_scale_finding because the field did not exist yet.
# New rows should not be blank.
#
# Rule:
# - If the column exists and is populated for a batch, good.
# - If a batch has blanks in commercial_scale_finding, flag as WARNING, not FAIL.
# - Future new batches should be enforced by Step 8A before archiving.

commercial_blank_mask = (
    archive_df[commercial_scale_col].isna()
    | (archive_df[commercial_scale_col].astype(str).str.strip() == "")
)

commercial_blank_df = archive_df.loc[
    commercial_blank_mask,
    ["company", "batch_name", "date_researched"]
].copy()

if commercial_blank_df.empty:
    print("COMMERCIAL SCALE FIELD CHECK: PASS")
else:
    print("COMMERCIAL SCALE FIELD CHECK: WARNING")
    print(
        "Some archived rows have blank commercial_scale_finding. "
        "This is expected for older research batches created before the commercial-scale field was added. "
        "New batches should not have blanks because Step 8A now blocks them."
    )
    display(commercial_blank_df)

# -----------------------------
# Validate fit_brief_json parses
# -----------------------------

json_issues = []

for _, row in archive_df.iterrows():
    try:
        json.loads(str(row["fit_brief_json"]))
    except Exception as e:
        json_issues.append({
            "company": row["company"],
            "batch_name": row.get("batch_name", ""),
            "json_issue": str(e)
        })

json_issue_df = pd.DataFrame(json_issues)

if json_issue_df.empty:
    print("JSON PARSE CHECK: PASS")
else:
    print("JSON PARSE CHECK: FAIL")
    display(json_issue_df)

# -----------------------------
# Duplicate company+batch check
# -----------------------------

duplicate_df = archive_df[
    archive_df.duplicated(subset=["company", "batch_name"], keep=False)
]

if duplicate_df.empty:
    print("DUPLICATE COMPANY+BATCH CHECK: PASS")
else:
    print("DUPLICATE COMPANY+BATCH CHECK: FAIL")
    display(duplicate_df[["company", "batch_name", "archive_saved_at"]])

# -----------------------------
# Batch summary
# -----------------------------

print("\nBATCH SUMMARY:")
display(
    archive_df
    .groupby("batch_name")
    .agg(
        company_count=("company", "nunique"),
        row_count=("company", "count"),
        commercial_scale_populated=(
            commercial_scale_col,
            lambda x: x.notna().sum() - (x.astype(str).str.strip() == "").sum()
        )
    )
    .reset_index()
)

print("\nALL ARCHIVED COMPANIES:")
display(
    archive_df[
        [
            "company",
            "batch_name",
            "date_researched",
            commercial_scale_col
        ]
    ]
    .sort_values(["batch_name", "company"])
)

# -----------------------------
# Final pass/fail gate
# -----------------------------

has_core_blank_fail = not blank_df.empty
has_json_fail = not json_issue_df.empty
has_duplicate_fail = not duplicate_df.empty

if has_core_blank_fail or has_json_fail or has_duplicate_fail:
    raise ValueError(
        "11A archive QA failed. Review failed checks above before continuing."
    )

print("\n11A raw archive QA complete.")
print("PASS: Core archive checks passed.")
print("NOTE: commercial_scale_finding blanks are warnings for older archived rows only.")

# =============================================================================

# STEP 12 - Add/update current batch in master

# =============================================================================

# 12 - Add current batch to master as review-needed
# Purpose:
# - Add/update the current batch summary_df into the active master
# - Preserve existing human-reviewed priority/status/notes for companies already in master
# - For brand-new companies, initialize reviewed_priority_level from model priority_level
# - Mark brand-new companies as "New batch - needs review"
# - Backup active master before saving
# - Save updated master locally + to Google Drive
# - Includes commercial_scale_finding and commercial_scale_assessment

import pandas as pd
import numpy as np
import shutil
import re
from pathlib import Path
from datetime import datetime
from google.colab import drive, files

# -----------------------------
# Config
# -----------------------------

drive.mount("/content/drive")

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_folder.mkdir(parents=True, exist_ok=True)

MASTER_FILENAME = "health_tech_market_research_summary_MASTER.csv"
drive_master_path = drive_folder / MASTER_FILENAME
local_master_path = Path(MASTER_FILENAME)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

backup_drive_path = drive_folder / f"health_tech_market_research_summary_MASTER_backup_before_12_{timestamp}.csv"
updated_snapshot_drive_path = drive_folder / f"health_tech_market_research_summary_MASTER_after_12_{timestamp}.csv"
change_log_drive_path = drive_folder / f"health_tech_market_research_summary_MASTER_12_change_log_{timestamp}.csv"

# Try to capture batch name if your notebook has one
if "BATCH_NAME" in globals():
    batch_label = str(BATCH_NAME)
elif "batch_name" in globals():
    batch_label = str(batch_name)
elif "RUN_ID" in globals():
    batch_label = str(RUN_ID)
else:
    batch_label = f"unknown_batch_{timestamp}"

# -----------------------------
# Safety checks
# -----------------------------

if "summary_df" not in globals():
    raise NameError(
        "STOP: summary_df is not defined. Run Step 10 before Step 12."
    )

if not isinstance(summary_df, pd.DataFrame):
    raise TypeError("STOP: summary_df exists but is not a pandas DataFrame.")

if summary_df.empty:
    raise ValueError("STOP: summary_df is empty. Do not update master.")

required_summary_cols = [
    "company",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "priority_level",
    "business_model_classification",
    "commercial_scale_finding",
    "commercial_scale_assessment",
    "calibration_flag"
]

missing_summary_cols = [col for col in required_summary_cols if col not in summary_df.columns]

if missing_summary_cols:
    raise ValueError(f"STOP: summary_df is missing required columns: {missing_summary_cols}")

if summary_df["company"].duplicated().any():
    dupes = summary_df[summary_df["company"].duplicated(keep=False)]["company"].tolist()
    raise ValueError(f"STOP: Duplicate companies found in summary_df: {dupes}")

if not drive_master_path.exists():
    raise FileNotFoundError(
        f"STOP: Active master not found in Google Drive: {drive_master_path}. "
        "Restore or create the master before running Step 12."
    )

# Load master from Drive, not local, to avoid stale local files
shutil.copy(drive_master_path, local_master_path)
master_df = pd.read_csv(local_master_path)

if "company" not in master_df.columns:
    raise ValueError("STOP: Master is missing company column.")

if master_df["company"].duplicated().any():
    dupes = master_df[master_df["company"].duplicated(keep=False)]["company"].tolist()
    raise ValueError(f"STOP: Duplicate companies found in master: {dupes}")

print("Loaded active master from Drive.")
print("Master shape before:", master_df.shape)
print("Master company count before:", master_df["company"].nunique())
print()
print("Current batch shape:", summary_df.shape)
print("Current batch companies:", summary_df["company"].tolist())

# -----------------------------
# Prepare columns
# -----------------------------

review_cols = [
    "reviewed_priority_level",
    "review_status",
    "review_notes"
]

model_cols_to_update = [
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "priority_level",
    "business_model_classification",
    "calibration_flag"
]

taxonomy_model_cols = [
    "primary_market_segment_code",
    "primary_market_segment",
    "market_segment",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_assignment_method",
    "taxonomy_assignment_basis",
]

optional_model_cols = [
    "final_takeaway",
    "commercial_scale_finding",
    "commercial_scale_assessment",
    "company_maturity_read",
    "likely_agency_level",
    "stage_timing_fit",
    "why_now_or_why_not",
    "timing_penalty_applied",
    "operator_timing_score_raw",
    "operator_timing_score_cap",
    "operator_timing_calibration_flag"
]

for col in taxonomy_model_cols:
    if col in summary_df.columns and col not in model_cols_to_update:
        model_cols_to_update.append(col)

for col in optional_model_cols:
    if col in summary_df.columns and col not in model_cols_to_update:
        model_cols_to_update.append(col)

# Ensure master has all necessary columns
for col in ["company"] + model_cols_to_update + review_cols:
    if col not in master_df.columns:
        master_df[col] = ""

# Ensure summary has review columns
batch_df = summary_df.copy()

# Canonicalize company names before matching batch rows to master rows.
# This prevents aliases like "Fay Nutrition" from creating duplicate companies.
master_df["company"] = master_df["company"].apply(canonical_company_name)
batch_df["company"] = batch_df["company"].apply(canonical_company_name)

duplicate_master_companies = (
    master_df.loc[master_df["company"].duplicated(keep=False), "company"]
    .dropna()
    .astype(str)
    .unique()
    .tolist()
)

if duplicate_master_companies:
    raise ValueError(
        "STOP: Duplicate canonical companies found in master before Step 12 update: "
        + ", ".join(sorted(duplicate_master_companies))
        + ". Run the cleanup repair cell first."
    )

duplicate_batch_companies = (
    batch_df.loc[batch_df["company"].duplicated(keep=False), "company"]
    .dropna()
    .astype(str)
    .unique()
    .tolist()
)

if duplicate_batch_companies:
    raise ValueError(
        "STOP: Duplicate canonical companies found in current batch: "
        + ", ".join(sorted(duplicate_batch_companies))
    )

for col in review_cols:
    if col not in batch_df.columns:
        batch_df[col] = ""

# Keep review fields text-safe after CSV loads.
# Some blank CSV columns can be inferred as float64 by pandas.
for col in review_cols:
    master_df[col] = master_df[col].astype("object").fillna("")
    batch_df[col] = batch_df[col].astype("object").fillna("")

# Clean calibration flags
if "calibration_flag" in master_df.columns:
    master_df["calibration_flag"] = master_df["calibration_flag"].fillna("")

if "calibration_flag" in batch_df.columns:
    batch_df["calibration_flag"] = batch_df["calibration_flag"].fillna("")

def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

# -----------------------------
# Backup before modifying
# -----------------------------

master_df.to_csv(local_master_path, index=False)
shutil.copy(local_master_path, backup_drive_path)

print("\nBackup saved before Step 12 update:")
print(backup_drive_path)

# -----------------------------
# Apply updates
# -----------------------------

master_before = master_df.copy()
existing_companies = set(master_df["company"].tolist())

change_log = []
rows_to_append = []

def log_change(company, field, old_value, new_value, change_type):
    if safe_text(old_value) != safe_text(new_value):
        change_log.append({
            "company": company,
            "field": field,
            "old_value": old_value,
            "new_value": new_value,
            "change_type": change_type,
            "batch_label": batch_label,
            "updated_at": timestamp
        })

for _, batch_row in batch_df.iterrows():
    company = batch_row["company"]

    if company in existing_companies:
        # Existing company: update model-generated fields only.
        # Preserve human-reviewed fields unless they are blank.
        idx = master_df.index[master_df["company"] == company].tolist()[0]

        for col in model_cols_to_update:
            if col in batch_df.columns:
                old_value = master_df.loc[idx, col]
                new_value = batch_row[col]
                log_change(company, col, old_value, new_value, "existing_company_model_field_update")
                master_df.loc[idx, col] = new_value

        # Preserve existing human-review columns.
        # If blank, initialize them gently from the current model output.
        if safe_text(master_df.loc[idx, "reviewed_priority_level"]) == "":
            new_reviewed_priority = safe_text(batch_row.get("priority_level"))
            log_change(
                company,
                "reviewed_priority_level",
                master_df.loc[idx, "reviewed_priority_level"],
                new_reviewed_priority,
                "existing_company_initialize_blank_review_field"
            )
            master_df.loc[idx, "reviewed_priority_level"] = new_reviewed_priority

        if safe_text(master_df.loc[idx, "review_status"]) == "":
            new_status = "Existing company - needs review"
            log_change(
                company,
                "review_status",
                master_df.loc[idx, "review_status"],
                new_status,
                "existing_company_initialize_blank_review_field"
            )
            master_df.loc[idx, "review_status"] = new_status

        if safe_text(master_df.loc[idx, "review_notes"]) == "":
            takeaway = safe_text(batch_row.get("final_takeaway"))
            if takeaway:
                new_note = f"Updated from {batch_label} on {timestamp}; needs human review. Model takeaway: {takeaway}"
            else:
                new_note = f"Updated from {batch_label} on {timestamp}; needs human review."

            log_change(
                company,
                "review_notes",
                master_df.loc[idx, "review_notes"],
                new_note,
                "existing_company_initialize_blank_review_field"
            )
            master_df.loc[idx, "review_notes"] = new_note

    else:
        # New company: append model fields and initialize human-review fields.
        new_row = {}

        # Preserve all master columns
        for col in master_df.columns:
            new_row[col] = ""

        new_row["company"] = company

        for col in model_cols_to_update:
            if col in batch_df.columns:
                new_row[col] = batch_row[col]

        model_priority = safe_text(batch_row.get("priority_level"))
        calibration_flag = safe_text(batch_row.get("calibration_flag"))
        takeaway = safe_text(batch_row.get("final_takeaway"))

        new_row["reviewed_priority_level"] = model_priority
        new_row["review_status"] = "New batch - needs review"

        review_note_parts = [
            f"Added from {batch_label} on {timestamp}; needs human review."
        ]

        if calibration_flag:
            review_note_parts.append(f"Calibration flag: {calibration_flag}.")

        if takeaway:
            review_note_parts.append(f"Model takeaway: {takeaway}")

        new_row["review_notes"] = " ".join(review_note_parts)

        rows_to_append.append(new_row)

        for col, new_value in new_row.items():
            if safe_text(new_value) != "":
                log_change(company, col, "", new_value, "new_company_added")

# Append new companies
if rows_to_append:
    master_df = pd.concat([master_df, pd.DataFrame(rows_to_append)], ignore_index=True)

# -----------------------------
# Validation
# -----------------------------

if master_df["company"].duplicated().any():
    dupes = master_df[master_df["company"].duplicated(keep=False)]["company"].tolist()
    raise ValueError(f"STOP: Duplicate companies created after update: {dupes}")

expected_company_count = master_before["company"].nunique() + len([
    c for c in batch_df["company"].tolist()
    if c not in existing_companies
])

actual_company_count = master_df["company"].nunique()

if actual_company_count != expected_company_count:
    raise ValueError(
        "STOP: Unexpected company count after update. "
        f"Expected {expected_company_count}, got {actual_company_count}."
    )

# Confirm all batch companies are now in master
missing_after_update = sorted(set(batch_df["company"].tolist()) - set(master_df["company"].tolist()))

if missing_after_update:
    raise ValueError(f"STOP: Batch companies missing from updated master: {missing_after_update}")

# Confirm commercial-scale fields populated for this batch in master
commercial_issues = []

for field in ["commercial_scale_finding", "commercial_scale_assessment"]:
    if field not in master_df.columns:
        commercial_issues.append({
            "company": "MASTER",
            "issue": f"Missing master column: {field}"
        })
    else:
        batch_rows = master_df[master_df["company"].isin(batch_df["company"].tolist())]
        blank_mask = batch_rows[field].isna() | (batch_rows[field].astype(str).str.strip() == "")
        for company in batch_rows.loc[blank_mask, "company"].tolist():
            commercial_issues.append({
                "company": company,
                "issue": f"Blank commercial-scale field after update: {field}"
            })

if commercial_issues:
    print("STOP: Commercial-scale fields were not correctly carried into the master.")
    display(pd.DataFrame(commercial_issues))
    raise ValueError("Commercial-scale master update validation failed.")

# Confirm human-reviewed fields are nonblank for batch companies
batch_master_rows = master_df[master_df["company"].isin(batch_df["company"].tolist())].copy()

blank_review_issues = []

for col in review_cols:
    blank_mask = batch_master_rows[col].isna() | (batch_master_rows[col].astype(str).str.strip() == "")
    for company in batch_master_rows.loc[blank_mask, "company"].tolist():
        blank_review_issues.append({
            "company": company,
            "blank_review_field": col
        })

if blank_review_issues:
    print("STOP: Some batch companies have blank review fields after update.")
    display(pd.DataFrame(blank_review_issues))
    raise ValueError("Blank review fields after Step 12 update.")

# -----------------------------
# Save updated master + change log
# -----------------------------

change_log_df = pd.DataFrame(change_log)

master_df.to_csv(local_master_path, index=False)

# Save active official master to Drive
shutil.copy(local_master_path, drive_master_path)

# Save timestamped snapshot to Drive
master_df.to_csv("health_tech_market_research_summary_MASTER_after_12.csv", index=False)
shutil.copy("health_tech_market_research_summary_MASTER_after_12.csv", updated_snapshot_drive_path)

# Save change log
local_change_log_path = Path(f"health_tech_market_research_summary_MASTER_12_change_log_{timestamp}.csv")
change_log_df.to_csv(local_change_log_path, index=False)
shutil.copy(local_change_log_path, change_log_drive_path)

print("\nStep 12 master update complete.")
print("Batch label:", batch_label)
print("Master shape after:", master_df.shape)
print("Master company count after:", master_df["company"].nunique())

print("\nActive master saved to:")
print(drive_master_path)

print("\nTimestamped updated master snapshot:")
print(updated_snapshot_drive_path)

print("\nStep 12 change log saved to:")
print(change_log_drive_path)

# -----------------------------
# Display outputs
# -----------------------------

new_companies_added = sorted([c for c in batch_df["company"].tolist() if c not in existing_companies])
existing_companies_updated = sorted([c for c in batch_df["company"].tolist() if c in existing_companies])

print("\nNew companies added:")
print(new_companies_added if new_companies_added else "None")

print("\nExisting companies updated:")
print(existing_companies_updated if existing_companies_updated else "None")

display_cols = [
    "company",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "business_model_classification",
    "commercial_scale_assessment",
    "commercial_scale_finding",
    "calibration_flag",
    "review_notes"
]

display_cols = [col for col in display_cols if col in master_df.columns]

print("\nBatch companies in updated master:")
display(
    master_df[
        master_df["company"].isin(batch_df["company"].tolist())
    ][display_cols].sort_values("company")
)

print("\nChange log preview:")
display(change_log_df)

print("\nReviewed priority summary after Step 12:")
display(
    master_df
    .groupby("reviewed_priority_level")
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .sort_values("company_count", ascending=False)
)

files.download(str(local_master_path))
files.download(str(local_change_log_path))


# =============================================================================
# STEP 12B - Priority field helper
# =============================================================================

# =============================================================================
# STEP 12B - Priority field helper
# =============================================================================
# Purpose:
# - Import shared P0-P4 priority utilities from GitHub package
# - Keep priority logic centralized in src/health_tech_research_agent/priority.py
# - Create final_priority_level, priority_source, final_priority_code, final_priority_rank
#
# Run after pulling latest GitHub repo in Colab.

from pathlib import Path
import sys

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"

if not SRC_DIR.exists():
    raise FileNotFoundError(
        f"STOP: src directory not found at {SRC_DIR}. "
        "Pull the GitHub repo / correct branch first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

from health_tech_research_agent.priority import (
    apply_priority_fields,
    extract_priority_code,
    normalize_priority_level,
    priority_code,
    priority_rank,
    safe_text,
    is_blank_value,
)

print("PASS: Step 12B priority helper imported from GitHub package.")
print("Priority model:")
print("- P0 = Highest-priority target / old P1")
print("- P1 = High-priority diligence / old P1-border")
print("- P2 = Worth deeper diligence")
print("- P3 = Watch list")
print("- P4 = Low priority / likely reject")

# =============================================================================

# STEP 13 - Load master dashboard using final priority

# =============================================================================

# 13 - Load master dashboard using final priority
# Purpose:
# - Load active master
# - Apply priority helper from Step 12B
# - Create final_priority_level using reviewed_priority_level first, then priority_level
# - Create priority_source for Auto Adjudicated vs Human Reviewed
# - Preserve decision_priority / decision_priority_sort as backward-compatible aliases for older dashboard steps
# - Does not modify or save the master

import pandas as pd
import shutil
from pathlib import Path
from google.colab import drive

drive.mount("/content/drive")

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_master_path = drive_folder / "health_tech_market_research_summary_MASTER.csv"
local_master_path = Path("health_tech_market_research_summary_MASTER.csv")

if not drive_master_path.exists():
    raise FileNotFoundError(f"Master not found: {drive_master_path}")

# Step 13 now depends on Step 12B.
# This prevents multiple cells from maintaining conflicting priority logic.
if "apply_priority_fields" not in globals():
    raise NameError(
        "STOP: apply_priority_fields is not defined. "
        "Run Step 12B - Priority field helper before Step 13."
    )

if "priority_rank" not in globals():
    raise NameError(
        "STOP: priority_rank is not defined. "
        "Run Step 12B - Priority field helper before Step 13."
    )

shutil.copy(drive_master_path, local_master_path)

master_df = pd.read_csv(local_master_path)

# -----------------------------
# Required baseline columns
# -----------------------------

if "company" not in master_df.columns:
    raise ValueError("STOP: Master is missing required column: company")

if "priority_level" not in master_df.columns:
    master_df["priority_level"] = ""

if "reviewed_priority_level" not in master_df.columns:
    master_df["reviewed_priority_level"] = ""

if "priority_review_note" not in master_df.columns:
    master_df["priority_review_note"] = ""

# -----------------------------
# Apply final priority logic from Step 12B
# -----------------------------

master_df = apply_priority_fields(master_df)

# Backward-compatible aliases for steps that still reference old names.
# These should eventually be replaced downstream with final_priority_level / final_priority_rank.
master_df["decision_priority"] = master_df["final_priority_level"]
master_df["decision_priority_sort"] = master_df["final_priority_rank"]

# -----------------------------
# Sort dashboard-ready master
# -----------------------------

sort_cols = []
ascending = []

if "final_priority_rank" in master_df.columns:
    sort_cols.append("final_priority_rank")
    ascending.append(True)

for col in [
    "pmf_scale_score",
    "thesis_fit_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "evidence_confidence_score"
]:
    if col in master_df.columns:
        sort_cols.append(col)
        ascending.append(False)

if sort_cols:
    master_df = master_df.sort_values(
        by=sort_cols,
        ascending=ascending
    ).reset_index(drop=True)

# -----------------------------
# Output summary
# -----------------------------

print("Master loaded for dashboard use.")
print("Master shape:", master_df.shape)
print("Company count:", master_df["company"].nunique())

print("\nFinal priority summary:")
priority_summary = (
    master_df
    .groupby(["final_priority_level", "priority_source"], dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
)

priority_summary["priority_sort"] = priority_summary["final_priority_level"].apply(priority_rank)

priority_summary = priority_summary.sort_values(
    by=["priority_sort", "priority_source"],
    ascending=[True, True]
).drop(columns=["priority_sort"])

display(priority_summary)

print("\nPriority source summary:")
display(
    master_df
    .groupby("priority_source", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .sort_values("priority_source")
)

human_reviewed_df = master_df[
    master_df["priority_source"].astype(str).str.lower().eq("human reviewed")
].copy()

if not human_reviewed_df.empty:
    print("\nHuman-reviewed priority overrides:")
    override_cols = [
        "company",
        "priority_level",
        "reviewed_priority_level",
        "final_priority_level",
        "priority_review_note"
    ]
    override_cols = [col for col in override_cols if col in human_reviewed_df.columns]

    display(
        human_reviewed_df[override_cols]
        .sort_values("final_priority_level")
        .reset_index(drop=True)
    )
else:
    print("\nNo human-reviewed priority overrides found.")

print("\nDashboard priority fields available:")
print("- priority_level = auto/adjudicated system priority")
print("- reviewed_priority_level = optional human override")
print("- final_priority_level = dashboard priority")
print("- priority_source = Auto Adjudicated or Human Reviewed")
print("- final_priority_rank = hidden/helper sort field")

# =============================================================================

# STEP 14 - Build market map view

# =============================================================================

# 14 - Build market map view
# Purpose:
# - Build a dashboard-ready market_map_df from master_df
# - Use final_priority_level / final_priority_rank from Step 12B as the priority source of truth
# - Restore market segment mapping
# - Create P0-aware strategic_bucket values
# - Preserve backward-compatible aliases for older downstream cells
#
# Run after:
# 12B -> 13
#
# Then continue:
# 15 -> 16 -> 17 -> 18 -> 19 -> 19A

import pandas as pd
import numpy as np
import re
import sys
from pathlib import Path
from datetime import datetime

# -----------------------------
# Import shared priority helper
# -----------------------------

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"

if SRC_DIR.exists():
    src_path = str(SRC_DIR)
    if src_path not in sys.path:
        sys.path.insert(0, src_path)

try:
    from health_tech_research_agent.priority import (
        apply_priority_fields,
        extract_priority_code,
        priority_rank,
        safe_text,
    )
except Exception as e:
    raise ImportError(
        "STOP: Could not import shared priority helpers. "
        "Run the GitHub pull cell and Step 12B first."
    ) from e

# Backward-compatible aliases used inside this cell.
extract_final_priority_code = extract_priority_code
final_priority_rank_from_level = priority_rank

# -----------------------------
# Validate inputs
# -----------------------------

if "master_df" not in globals() or not isinstance(master_df, pd.DataFrame) or master_df.empty:
    raise NameError("STOP: master_df not found or empty. Run Step 13 first.")

market_map_df = master_df.copy()
market_map_df = apply_priority_fields(market_map_df)

# -----------------------------
# Helpers
# -----------------------------

from health_tech_research_agent.dashboard import (
    existing_cols,
    normalize_name,
    map_market_segment,
    map_strategic_bucket,
)

# -----------------------------
# Required fields
# -----------------------------

default_columns = {
    "company": "",
    "market_segment": "",
    "primary_market_segment_code": "",
    "primary_market_segment": "",
    "subsegment_tags": "",
    "product_model_tags": "",
    "distribution_model_tags": "",
    "data_input_tags": "",
    "taxonomy_assignment_method": "",
    "taxonomy_assignment_basis": "",
    "strategic_bucket": "",
    "final_priority_level": "",
    "priority_source": "",
    "final_priority_code": "",
    "final_priority_rank": 99,
    "priority_level": "",
    "reviewed_priority_level": "",
    "priority_review_note": "",
    "thesis_fit_score": np.nan,
    "pmf_scale_score": np.nan,
    "evidence_confidence_score": np.nan,
    "katelynd_role_fit_score": np.nan,
    "operator_timing_score": np.nan,
    "business_model_classification": "",
    "commercial_scale_assessment": "",
    "pmf_scale_assessment": "",
    "final_recommendation": "",
    "final_takeaway": "",
    "calibration_flag": ""
}

for col_name, default_value in default_columns.items():
    if col_name not in market_map_df.columns:
        market_map_df[col_name] = default_value

for score_col in [
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]:
    market_map_df[score_col] = pd.to_numeric(
        market_map_df[score_col],
        errors="coerce"
    )

# Force clean P0-P4 code/rank from the shared priority helper.
market_map_df["final_priority_code"] = market_map_df["final_priority_level"].apply(extract_priority_code)
market_map_df["final_priority_rank"] = market_map_df["final_priority_level"].apply(priority_rank)

# Backward-compatible aliases.
market_map_df["decision_priority"] = market_map_df["final_priority_level"]
market_map_df["decision_priority_sort"] = market_map_df["final_priority_rank"]
market_map_df["reviewed_priority_rank"] = market_map_df["final_priority_rank"]

# -----------------------------
# Market segment mapping
# -----------------------------
# First preserve an existing good market_segment.
# Then use exact known-company mapping.
# Then infer from company/business model/evidence text.
# This prevents new researched companies from silently landing in Unmapped.

segment_map = {
    "pomelo care": "Women’s and family health",
    "thyme care": "Oncology / cancer navigation",
    "waymark": "Medicaid / value-based care",

    # Nutrition, metabolic health, obesity, food as medicine
    "nourish": "Nutrition / food as medicine",
    "fay": "Nutrition / food as medicine",
    "fay nutrition": "Nutrition / food as medicine",
    "berry street": "Nutrition / food as medicine",
    "season health": "Nutrition / food as medicine",
    "culina health": "Nutrition / food as medicine",
    "summer health": "Women’s and family health",
    "9amhealth": "Metabolic health / virtual care",
    "noom med": "Metabolic health / obesity care",
    "omada health": "Metabolic health / digital therapeutics",
    "levels health": "Metabolic health / D2C behavior change",
    "signos": "Metabolic health / D2C behavior change",
    "zoe": "Metabolic health / D2C behavior change",

    # GI / specialty care
    "oshi health": "GI / specialty virtual care",

    # Women’s health / family health
    "maven clinic": "Women’s and family health",
    "midi health": "Women’s and family health",
    "allara health": "Women’s and family health",
    "visana health": "Women’s and family health",
    "familywell health": "Women’s and family health",
    "mae health": "Women’s and family health",
    "oula": "Women’s and family health",
    "tia": "Women’s and family health",
    "oova": "Women’s health / fertility",

    # MSK
    "hinge health": "MSK / digital physical therapy",
    "sword health": "MSK / digital physical therapy",

    # Mental and behavioral health
    "equip health": "Behavioral health / eating disorder care",
    "grow therapy": "Mental health / provider marketplace",
    "headway mental health insurance network": "Mental health / insurance network",
    "headway": "Mental health / insurance network",
    "affect therapeutics": "Behavioral health / substance use treatment",

    # Navigation, advocacy, hybrid care
    "transcarent": "Care navigation / hybrid care",
    "solace health patient advocacy": "Care navigation / advocacy",
    "solace health": "Care navigation / advocacy",
    "angle health": "Health insurance / benefits infrastructure",
    "included health": "Care navigation / hybrid care",

    # Oncology / serious illness
    "jasper health": "Oncology / cancer navigation",
    "outcomes4me": "Oncology / cancer navigation",

    # Preventive health / diagnostics / wearables / clinical intelligence
    "function health": "Preventive health / diagnostics",
    "insidetracker": "Preventive health / diagnostics",
    "oura": "Wearables / consumer health",
    "openevidence": "Clinical AI / provider intelligence",
}

UNMAPPED_SEGMENT_VALUES = {
    "",
    "nan",
    "none",
    "null",
    "unmapped",
    "unassigned",
    "needs segment review",
}

def step14_segment_text(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()

def step14_row_text_for_segment(row):
    fields_for_inference = [
        "company",
        "business_model_classification",
        "final_recommendation",
        "final_takeaway",
        "commercial_scale_assessment",
        "pmf_scale_assessment",
        "commercial_scale_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "review_notes",
        "priority_review_note",
        "calibration_flag",
    ]

    return " ".join(
        step14_segment_text(row.get(col, "")).lower()
        for col in fields_for_inference
        if col in row.index
    )

def infer_market_segment(row, segment_map):
    existing_segment = step14_segment_text(row.get("market_segment", ""))
    existing_segment_key = existing_segment.lower()

    # Preserve manually assigned / already valid segment labels.
    if existing_segment_key not in UNMAPPED_SEGMENT_VALUES:
        return existing_segment

    company_key = normalize_name(row.get("company", ""))

    # Exact company-name mapping still wins.
    if company_key in segment_map:
        return segment_map[company_key]

    text = step14_row_text_for_segment(row)

    # Women’s / family / maternal health.
    # Put this before generic metabolic/Medicaid rules because women’s health companies
    # often mention metabolic, Medicaid, payer, or virtual care.
    maternal_terms = [
        "maternal",
        "maternity",
        "pregnancy",
        "postpartum",
        "doula",
        "midwife",
        "midwives",
        "ob/gyn",
        "obgyn",
        "newborn",
        "birth",
        "perinatal",
    ]

    if any(term in text for term in maternal_terms):
        return "Women’s and family health"

    fertility_terms = [
        "fertility",
        "ivf",
        "ovulation",
        "egg freezing",
        "hormone testing",
        "hormone test",
    ]

    if any(term in text for term in fertility_terms):
        return "Women’s health / fertility"

    womens_health_terms = [
        "women’s health",
        "women's health",
        "womens health",
        "female health",
        "gynecologic",
        "gynecology",
        "menopause",
        "perimenopause",
        "pcos",
        "endometriosis",
        "hormonal",
        "women’s healthcare",
        "women's healthcare",
        "womens healthcare",
    ]

    if any(term in text for term in womens_health_terms):
        return "Women’s and family health"

    # Nutrition / food as medicine.
    nutrition_terms = [
        "nutrition",
        "dietitian",
        "dietician",
        "food as medicine",
        "food-as-medicine",
        "registered dietitian",
        "medical nutrition",
    ]

    if any(term in text for term in nutrition_terms):
        return "Nutrition / food as medicine"

    # Oncology / cancer navigation.
    oncology_terms = [
        "oncology",
        "cancer",
        "tumor",
        "serious illness",
    ]

    if any(term in text for term in oncology_terms):
        return "Oncology / cancer navigation"

    # MSK / physical therapy.
    msk_terms = [
        "msk",
        "musculoskeletal",
        "physical therapy",
        "digital physical therapy",
        "pain management",
    ]

    if any(term in text for term in msk_terms):
        return "MSK / digital physical therapy"

    # Mental / behavioral health.
    eating_disorder_terms = [
        "eating disorder",
        "eating disorders",
    ]

    if any(term in text for term in eating_disorder_terms):
        return "Behavioral health / eating disorder care"

    substance_use_terms = [
        "substance use",
        "addiction",
        "alcohol use",
        "opioid",
    ]

    if any(term in text for term in substance_use_terms):
        return "Behavioral health / substance use treatment"

    mental_health_terms = [
        "mental health",
        "behavioral health",
        "therapy",
        "therapist",
        "psychiatry",
        "depression",
        "anxiety",
    ]

    if any(term in text for term in mental_health_terms):
        if "insurance" in text or "in-network" in text or "payer" in text:
            return "Mental health / insurance network"
        return "Mental health / provider marketplace"

    # Medicaid / value-based care.
    if "medicaid" in text or "value-based care" in text or "vbc" in text:
        return "Medicaid / value-based care"

    # Care navigation / advocacy.
    if "advocacy" in text or "patient advocacy" in text:
        return "Care navigation / advocacy"

    if "care navigation" in text or "navigation" in text:
        return "Care navigation / hybrid care"

    # Metabolic / obesity / diabetes / CGM.
    if "digital therapeutics" in text or "digital therapeutic" in text:
        return "Metabolic health / digital therapeutics"

    obesity_terms = [
        "obesity",
        "weight loss",
        "weight management",
        "glp-1",
        "glp1",
    ]

    if any(term in text for term in obesity_terms):
        return "Metabolic health / obesity care"

    metabolic_terms = [
        "metabolic",
        "diabetes",
        "cardiometabolic",
        "cgm",
        "glucose",
    ]

    if any(term in text for term in metabolic_terms):
        return "Metabolic health / virtual care"

    # Preventive health / diagnostics / longevity.
    diagnostics_terms = [
        "diagnostics",
        "biomarker",
        "blood testing",
        "lab testing",
        "preventive health",
        "preventative health",
        "longevity",
    ]

    if any(term in text for term in diagnostics_terms):
        return "Preventive health / diagnostics"

    # Wearables / consumer health.
    wearable_terms = [
        "wearable",
        "wearables",
        "smart ring",
        "sleep tracking",
        "recovery tracking",
    ]

    if any(term in text for term in wearable_terms):
        return "Wearables / consumer health"

    # Clinical AI / provider workflow.
    clinical_ai_terms = [
        "clinical ai",
        "provider intelligence",
        "clinical decision support",
        "medical search",
        "medical ai",
    ]

    if any(term in text for term in clinical_ai_terms):
        return "Clinical AI / provider intelligence"

    # Do not silently call this Unmapped. Make the needed human action explicit.
    return "Needs segment review"


# -----------------------------
# Controlled taxonomy classification
# -----------------------------
# This replaces loose market-segment inference with the controlled taxonomy.
# Existing downstream dashboard logic still uses market_segment as a label alias,
# but the source-of-truth fields are primary_market_segment_code and primary_market_segment.

from health_tech_research_agent.taxonomy import classify_dataframe

market_map_df = classify_dataframe(
    market_map_df,
    taxonomy_dir=REPO_DIR / "taxonomy",
    hard_stop=True,
)

taxonomy_cols = existing_cols(market_map_df, [
    "company",
    "primary_market_segment_code",
    "primary_market_segment",
    "market_segment",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_assignment_method",
    "taxonomy_assignment_basis",
    "business_model_classification",
    "final_takeaway",
])

print("")
print("CONTROLLED TAXONOMY CLASSIFICATION COMPLETE")
print("=" * 80)
print("Taxonomy columns added:")
for col in taxonomy_cols:
    print("-", col)

print("")
print("Taxonomy assignment method summary:")
display(
    market_map_df
    .groupby("taxonomy_assignment_method", dropna=False)["company"]
    .nunique()
    .reset_index(name="company_count")
    .sort_values(["company_count", "taxonomy_assignment_method"], ascending=[False, True])
)

print("")
print("Primary market segment summary:")
display(
    market_map_df
    .groupby(["primary_market_segment", "primary_market_segment_code"], dropna=False)["company"]
    .nunique()
    .reset_index(name="company_count")
    .sort_values(["company_count", "primary_market_segment"], ascending=[False, True])
)

# -----------------------------
# Strategic bucket mapping
# -----------------------------
# P0-aware replacement for old logic that treated P0 as unmapped/unprioritized.

market_map_df["strategic_bucket"] = market_map_df.apply(map_strategic_bucket, axis=1)

# -----------------------------
# Sort market map
# -----------------------------

sort_cols = []
ascending = []

for col_name, ascending_value in [
    ("final_priority_rank", True),
    ("thesis_fit_score", False),
    ("pmf_scale_score", False),
    ("katelynd_role_fit_score", False),
    ("operator_timing_score", False),
    ("evidence_confidence_score", False),
    ("company", True)
]:
    if col_name in market_map_df.columns:
        sort_cols.append(col_name)
        ascending.append(ascending_value)

if sort_cols:
    market_map_df = market_map_df.sort_values(
        by=sort_cols,
        ascending=ascending
    ).reset_index(drop=True)

# -----------------------------
# Save snapshot
# -----------------------------

snapshot_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
market_map_snapshot_path = Path(f"health_tech_market_map_snapshot_{snapshot_timestamp}.csv")
market_map_df.to_csv(market_map_snapshot_path, index=False)

# -----------------------------
# Output checks
# -----------------------------

print("Market map view built.")
print("Shape:", market_map_df.shape)
print("Snapshot:", market_map_snapshot_path)

print("\nFinal priority summary:")
priority_summary = (
    market_map_df
    .groupby(["final_priority_level", "priority_source"], dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
)

priority_summary["final_priority_rank"] = priority_summary["final_priority_level"].apply(priority_rank)

priority_summary = priority_summary.sort_values(
    by=["final_priority_rank", "priority_source"],
    ascending=[True, True]
).drop(columns=["final_priority_rank"])

display(priority_summary)

print("\nStrategic bucket summary:")
display(
    market_map_df
    .groupby("strategic_bucket", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .sort_values("company_count", ascending=False)
)

print("\nMarket segment summary:")
display(
    market_map_df
    .groupby("market_segment", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .sort_values(["company_count", "market_segment"], ascending=[False, True])
)

print("\nPreview:")
preview_cols = existing_cols(market_map_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "market_segment",
    "primary_market_segment_code",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_assignment_method",
    "strategic_bucket",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_takeaway"
])

display(market_map_df[preview_cols].head(20))

# =============================================================================

# STEP 15 - Segment-level summary

# =============================================================================

# 15 - Segment-level summary
# Purpose:
# - Summarize market segments using final priority logic
# - Use final_priority_level / final_priority_rank as dashboard source of truth
# - Preserve visibility into priority_source and human-reviewed overrides

import pandas as pd
import numpy as np

if "market_map_df" not in globals():
    raise NameError("STOP: market_map_df is not defined. Run Step 14 first.")

required_cols = [
    "company",
    "market_segment",
    "final_priority_level",
    "final_priority_rank",
    "priority_source",
    "strategic_bucket",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]

missing_cols = [col for col in required_cols if col not in market_map_df.columns]

if missing_cols:
    raise ValueError(f"STOP: market_map_df is missing required columns: {missing_cols}")

summary_df = market_map_df.copy()

# Optional columns for display
for col in ["review_status", "review_notes", "priority_review_note", "priority_level", "reviewed_priority_level"]:
    if col not in summary_df.columns:
        summary_df[col] = ""

# -----------------------------
# Segment summary
# -----------------------------

segment_summary = (
    summary_df
    .groupby("market_segment", dropna=False)
    .agg(
        company_count=("company", "nunique"),
        p1_count=("final_priority_level", lambda x: x.astype(str).str.contains("P1", case=False, na=False).sum()),
        p2_count=("final_priority_level", lambda x: x.astype(str).str.contains("P2", case=False, na=False).sum()),
        p3_count=("final_priority_level", lambda x: x.astype(str).str.contains("P3", case=False, na=False).sum()),
        p4_count=("final_priority_level", lambda x: x.astype(str).str.contains("P4", case=False, na=False).sum()),
        human_reviewed_count=("priority_source", lambda x: x.astype(str).str.contains("Human Reviewed", case=False, na=False).sum()),
        avg_thesis_fit=("thesis_fit_score", "mean"),
        avg_pmf_scale=("pmf_scale_score", "mean"),
        avg_evidence_confidence=("evidence_confidence_score", "mean"),
        avg_katelynd_role_fit=("katelynd_role_fit_score", "mean"),
        avg_operator_timing=("operator_timing_score", "mean"),
        best_final_priority_rank=("final_priority_rank", "min")
    )
    .reset_index()
)

# Round averages for readability
score_cols = [
    "avg_thesis_fit",
    "avg_pmf_scale",
    "avg_evidence_confidence",
    "avg_katelynd_role_fit",
    "avg_operator_timing"
]

for col in score_cols:
    segment_summary[col] = segment_summary[col].round(1)

segment_summary = segment_summary.sort_values(
    by=[
        "best_final_priority_rank",
        "p1_count",
        "p2_count",
        "avg_thesis_fit",
        "avg_operator_timing",
        "avg_pmf_scale"
    ],
    ascending=[True, False, False, False, False, False]
).reset_index(drop=True)

print("SEGMENT SUMMARY")
display(segment_summary)

# -----------------------------
# Companies by segment
# -----------------------------

company_by_segment_cols = [
    "market_segment",
    "company",
    "strategic_bucket",
    "final_priority_level",
    "priority_source",
    "priority_review_note",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "review_notes",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]

company_by_segment_cols = [col for col in company_by_segment_cols if col in summary_df.columns]

company_by_segment = (
    summary_df
    .sort_values(
        by=[
            "market_segment",
            "final_priority_rank",
            "thesis_fit_score",
            "operator_timing_score",
            "pmf_scale_score"
        ],
        ascending=[True, True, False, False, False]
    )
    [company_by_segment_cols]
    .reset_index(drop=True)
)

print("COMPANIES BY SEGMENT")
display(company_by_segment)

# -----------------------------
# Top / diligence targets
# -----------------------------

top_target_mask = (
    summary_df["final_priority_level"]
    .astype(str)
    .str.contains("P1|P2|Strong P2|P1-border|P1 border|Worth deeper diligence", case=False, na=False)
)

top_targets_cols = [
    "company",
    "market_segment",
    "strategic_bucket",
    "final_priority_level",
    "priority_source",
    "priority_review_note",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "review_notes",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]

top_targets_cols = [col for col in top_targets_cols if col in summary_df.columns]

top_targets = (
    summary_df[top_target_mask]
    .sort_values(
        by=[
            "final_priority_rank",
            "thesis_fit_score",
            "operator_timing_score",
            "pmf_scale_score"
        ],
        ascending=[True, False, False, False]
    )
    [top_targets_cols]
    .reset_index(drop=True)
)

print("TOP / DILIGENCE TARGETS")
display(top_targets)

# -----------------------------
# Human-reviewed overrides
# -----------------------------

human_reviewed = summary_df[
    summary_df["priority_source"].astype(str).str.contains("Human Reviewed", case=False, na=False)
].copy()

human_reviewed_cols = [
    "company",
    "market_segment",
    "priority_level",
    "reviewed_priority_level",
    "final_priority_level",
    "priority_review_note",
    "review_status",
    "review_notes"
]

human_reviewed_cols = [col for col in human_reviewed_cols if col in human_reviewed.columns]

print("HUMAN-REVIEWED PRIORITY OVERRIDES")
if human_reviewed.empty:
    print("No human-reviewed priority overrides found.")
else:
    display(
        human_reviewed[human_reviewed_cols]
        .sort_values(["final_priority_level", "company"])
        .reset_index(drop=True)
    )

# -----------------------------
# Unassigned companies
# -----------------------------

unassigned = summary_df[summary_df["market_segment"] == "Unassigned"].copy()

unassigned_cols = [
    "company",
    "final_priority_level",
    "priority_source",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "review_notes"
]

unassigned_cols = [col for col in unassigned_cols if col in unassigned.columns]

print("UNASSIGNED COMPANIES")
if unassigned.empty:
    print("No unassigned companies.")
else:
    display(
        unassigned[unassigned_cols]
        .sort_values(["final_priority_level", "company"])
        .reset_index(drop=True)
    )

# =============================================================================

# STEP 16 - Segment priority summary

# =============================================================================

# 16 - Segment priority summary
# Purpose:
# - Build segment-level priority summary using final_priority_level / final_priority_rank
# - Use strategic_bucket from Step 14
# - Use segment_summary from Step 15
# - Identify best companies per segment using final priority + scores

import pandas as pd
import numpy as np

if "market_map_df" not in globals():
    raise NameError("STOP: market_map_df is not defined. Run Step 14 first.")

if "segment_summary" not in globals():
    raise NameError("STOP: segment_summary is not defined. Run Step 15 first.")

required_market_cols = [
    "company",
    "market_segment",
    "strategic_bucket",
    "final_priority_level",
    "final_priority_rank",
    "thesis_fit_score",
    "pmf_scale_score",
    "operator_timing_score"
]

missing_market_cols = [col for col in required_market_cols if col not in market_map_df.columns]

if missing_market_cols:
    raise ValueError(f"STOP: market_map_df is missing required columns: {missing_market_cols}")

if "market_segment" not in segment_summary.columns:
    raise ValueError("STOP: segment_summary is missing required column: market_segment")

summary_source_df = market_map_df.copy()

# Optional fields
for col in [
    "priority_source",
    "priority_level",
    "reviewed_priority_level",
    "priority_review_note",
    "evidence_confidence_score",
    "katelynd_role_fit_score"
]:
    if col not in summary_source_df.columns:
        summary_source_df[col] = ""

# -----------------------------
# Strategic bucket counts by segment
# -----------------------------

priority_counts = (
    summary_source_df
    .pivot_table(
        index="market_segment",
        columns="strategic_bucket",
        values="company",
        aggfunc="nunique",
        fill_value=0
    )
    .reset_index()
)

# Flatten pivot column names just in case
priority_counts.columns = [
    str(col).strip() if not isinstance(col, tuple) else "_".join([str(x) for x in col if x])
    for col in priority_counts.columns
]

# -----------------------------
# Final priority counts by segment
# -----------------------------

final_priority_counts = (
    summary_source_df
    .pivot_table(
        index="market_segment",
        columns="final_priority_level",
        values="company",
        aggfunc="nunique",
        fill_value=0
    )
    .reset_index()
)

final_priority_counts.columns = [
    str(col).strip() if col == "market_segment" else f"count_{str(col).split(':')[0].strip()}"
    for col in final_priority_counts.columns
]

# -----------------------------
# Human-reviewed count by segment
# -----------------------------

human_reviewed_counts = (
    summary_source_df
    .assign(
        is_human_reviewed=lambda df: (
            df["priority_source"]
            .astype(str)
            .str.contains("Human Reviewed", case=False, na=False)
        )
    )
    .groupby("market_segment", dropna=False)
    .agg(
        human_reviewed_count=("is_human_reviewed", "sum")
    )
    .reset_index()
)

# -----------------------------
# Best companies per segment
# -----------------------------

best_companies = (
    summary_source_df
    .sort_values(
        by=[
            "market_segment",
            "final_priority_rank",
            "thesis_fit_score",
            "operator_timing_score",
            "pmf_scale_score"
        ],
        ascending=[True, True, False, False, False]
    )
    .groupby("market_segment", dropna=False)
    .agg(
        best_companies=(
            "company",
            lambda x: ", ".join(x.head(4))
        )
    )
    .reset_index()
)

# -----------------------------
# Best priority label per segment
# -----------------------------

best_priority_by_segment = (
    summary_source_df
    .sort_values(
        by=[
            "market_segment",
            "final_priority_rank",
            "thesis_fit_score",
            "operator_timing_score",
            "pmf_scale_score"
        ],
        ascending=[True, True, False, False, False]
    )
    .groupby("market_segment", dropna=False)
    .agg(
        best_final_priority_level=("final_priority_level", "first"),
        best_final_priority_rank=("final_priority_rank", "min")
    )
    .reset_index()
)

# -----------------------------
# Merge summaries
# -----------------------------

segment_priority_summary = (
    segment_summary
    .merge(
        priority_counts,
        on="market_segment",
        how="left"
    )
    .merge(
        final_priority_counts,
        on="market_segment",
        how="left"
    )
    .merge(
        human_reviewed_counts,
        on="market_segment",
        how="left"
    )
    .merge(
        best_companies,
        on="market_segment",
        how="left"
    )
    .merge(
        best_priority_by_segment,
        on="market_segment",
        how="left",
        suffixes=("", "_from_market_map")
    )
)

# Prefer best_final_priority_rank from Step 15 if already present; otherwise use merged version
if "best_final_priority_rank" not in segment_priority_summary.columns:
    if "best_final_priority_rank_from_market_map" in segment_priority_summary.columns:
        segment_priority_summary["best_final_priority_rank"] = segment_priority_summary["best_final_priority_rank_from_market_map"]

if "best_final_priority_rank_from_market_map" in segment_priority_summary.columns:
    segment_priority_summary = segment_priority_summary.drop(columns=["best_final_priority_rank_from_market_map"])

# Clean numeric count columns
count_cols = [
    col for col in segment_priority_summary.columns
    if col.startswith("count_")
    or col in [
        "Priority target",
        "High-priority diligence",
        "Diligence target",
        "Evidence / role-fit review",
        "Watch list",
        "Low priority",
        "Unprioritized",
        "human_reviewed_count"
    ]
]

for col in count_cols:
    if col in segment_priority_summary.columns:
        segment_priority_summary[col] = (
            segment_priority_summary[col]
            .fillna(0)
            .astype(int)
        )

# -----------------------------
# Sort for dashboard usefulness
# -----------------------------

sort_cols = []
ascending = []

if "best_final_priority_rank" in segment_priority_summary.columns:
    sort_cols.append("best_final_priority_rank")
    ascending.append(True)

for col in [
    "p1_count",
    "p2_count",
    "avg_thesis_fit",
    "avg_operator_timing",
    "avg_pmf_scale"
]:
    if col in segment_priority_summary.columns:
        sort_cols.append(col)
        ascending.append(False)

if sort_cols:
    segment_priority_summary = segment_priority_summary.sort_values(
        by=sort_cols,
        ascending=ascending
    ).reset_index(drop=True)

print("SEGMENT PRIORITY SUMMARY")
display(segment_priority_summary)

# -----------------------------
# Optional quick view
# -----------------------------

quick_view_cols = [
    "market_segment",
    "company_count",
    "best_final_priority_level",
    "best_final_priority_rank",
    "p1_count",
    "p2_count",
    "p3_count",
    "p4_count",
    "Priority target",
    "High-priority diligence",
    "Diligence target",
    "Watch list",
    "Low priority",
    "human_reviewed_count",
    "avg_thesis_fit",
    "avg_pmf_scale",
    "avg_katelynd_role_fit",
    "avg_operator_timing",
    "best_companies"
]

quick_view_cols = [col for col in quick_view_cols if col in segment_priority_summary.columns]

print("SEGMENT PRIORITY SUMMARY - QUICK VIEW")
display(segment_priority_summary[quick_view_cols])

# =============================================================================

# STEP 17 - Company data depth audit

# =============================================================================

# =============================================================================
# STEP 17 - Company data depth audit
# =============================================================================
# Purpose:
# - Audit whether each company has recoverable raw research evidence
# - Use final_priority_level / final_priority_rank as dashboard source of truth
# - Import shared P0-P4 priority logic from src/health_tech_research_agent/priority.py
# - Check raw evidence completeness across funding, payer/institutional, outcomes,
#   commercial scale, and fit brief JSON
#
# Run after:
# 12B -> 13 -> 14 -> 15 -> 16
#
# Then continue:
# 18 -> 19 -> 19A

import pandas as pd
import json
import re
import sys
from pathlib import Path

# -----------------------------
# Import shared priority helper
# -----------------------------

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"

if SRC_DIR.exists():
    src_path = str(SRC_DIR)
    if src_path not in sys.path:
        sys.path.insert(0, src_path)

try:
    from health_tech_research_agent.priority import (
        apply_priority_fields,
        extract_priority_code,
        priority_rank,
        safe_text,
    )
except Exception as e:
    raise ImportError(
        "STOP: Could not import shared priority helpers. "
        "Run the GitHub pull cell and Step 12B first."
    ) from e

# -----------------------------
# Safety checks
# -----------------------------

if "master_df" not in globals():
    raise NameError("STOP: master_df is not defined. Run Step 13 first.")

master_summary_df = master_df.copy()

# -----------------------------
# Helpers
# -----------------------------

def nonblank(value):
    return pd.notna(value) and str(value).strip() != ""

def clean_json_text(text):
    text = str(text).strip()
    text = re.sub(r"^```json", "", text)
    text = re.sub(r"^```", "", text)
    text = re.sub(r"```$", "", text)
    return text.strip()

def parseable_json(value):
    if not nonblank(value):
        return False

    try:
        raw = clean_json_text(value)
        start = raw.find("{")

        if start == -1:
            return False

        decoder = json.JSONDecoder()
        parsed, end = decoder.raw_decode(raw[start:])

        return isinstance(parsed, dict)

    except Exception:
        return False

def is_priority_or_diligence_target(value):
    """
    Priority/diligence target means P0, P1, or P2 in the native model.
    This replaces the old P1/P2-only helper.
    """
    code = extract_priority_code(value)
    return code in ["P0", "P1", "P2"]

# -----------------------------
# Build audit base
# -----------------------------

if "market_map_df" in globals():
    base_cols = [
        "company",
        "market_segment",
        "strategic_bucket",
        "final_priority_level",
        "final_priority_rank",
        "priority_source",
        "priority_review_note",
        "priority_level",
        "reviewed_priority_level",
        "review_status",
        "review_notes",
        "thesis_fit_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "katelynd_role_fit_score",
        "operator_timing_score",
        "calibration_flag"
    ]

    base_cols = [col for col in base_cols if col in market_map_df.columns]
    audit_base = market_map_df[base_cols].copy()

else:
    audit_base = master_summary_df.copy()
    audit_base = apply_priority_fields(audit_base)

    if "market_segment" not in audit_base.columns:
        audit_base["market_segment"] = ""

    if "strategic_bucket" not in audit_base.columns:
        audit_base["strategic_bucket"] = ""

# Ensure priority fields exist / are fresh.
audit_base = apply_priority_fields(audit_base)

if "market_segment" not in audit_base.columns:
    audit_base["market_segment"] = ""

if "strategic_bucket" not in audit_base.columns:
    audit_base["strategic_bucket"] = ""

for col in [
    "priority_source",
    "priority_review_note",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "review_notes",
    "calibration_flag"
]:
    if col not in audit_base.columns:
        audit_base[col] = ""

# Force rank from shared helper so P0 is handled correctly.
audit_base["final_priority_rank"] = audit_base["final_priority_level"].apply(priority_rank)

# -----------------------------
# Find raw/checkpoint files that may contain full research evidence
# -----------------------------

candidate_paths = []

local_patterns = [
    "research_batches/*.csv",
    "health_tech_market_research_full_results.csv",
    "health_tech_market_research_results_checkpoint.csv",
    "health_tech_raw_research_ARCHIVE.csv"
]

for pattern in local_patterns:
    candidate_paths.extend(Path(".").glob(pattern))

# Include Drive archive/batch files if available
drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_batches_folder = drive_folder / "research_batches"

if drive_folder.exists():
    candidate_paths.append(drive_folder / "health_tech_raw_research_ARCHIVE.csv")
    candidate_paths.append(drive_folder / "health_tech_market_research_summary_MASTER.csv")

if drive_batches_folder.exists():
    candidate_paths.extend(drive_batches_folder.glob("*.csv"))

# Deduplicate paths
candidate_paths = sorted(set([path for path in candidate_paths if Path(path).exists()]))

raw_frames = []

for path in candidate_paths:
    try:
        temp = pd.read_csv(path)

        if "company" not in temp.columns:
            continue

        raw_cols = [
            "company",
            "date_researched",
            "funding_finding",
            "payer_institutional_finding",
            "outcomes_finding",
            "commercial_scale_finding",
            "fit_brief_json",
            "batch_name"
        ]

        available_cols = [col for col in raw_cols if col in temp.columns]

        evidence_cols = {
            "funding_finding",
            "payer_institutional_finding",
            "outcomes_finding",
            "commercial_scale_finding",
            "fit_brief_json"
        }

        # Only keep files with at least one raw evidence column.
        if len(set(available_cols) & evidence_cols) == 0:
            continue

        temp = temp[available_cols].copy()
        temp["source_file"] = str(path)
        raw_frames.append(temp)

    except Exception as e:
        print(f"Skipping {path}: {e}")

if raw_frames:
    raw_inventory = pd.concat(raw_frames, ignore_index=True)
else:
    raw_inventory = pd.DataFrame(columns=[
        "company",
        "date_researched",
        "funding_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "commercial_scale_finding",
        "fit_brief_json",
        "batch_name",
        "source_file"
    ])

# -----------------------------
# Raw inventory completeness flags
# -----------------------------

for col in [
    "date_researched",
    "funding_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "commercial_scale_finding",
    "fit_brief_json",
    "batch_name",
    "source_file"
]:
    if col not in raw_inventory.columns:
        raw_inventory[col] = ""

raw_inventory["has_funding_raw"] = raw_inventory["funding_finding"].apply(nonblank)
raw_inventory["has_payer_raw"] = raw_inventory["payer_institutional_finding"].apply(nonblank)
raw_inventory["has_outcomes_raw"] = raw_inventory["outcomes_finding"].apply(nonblank)
raw_inventory["has_commercial_scale_raw"] = raw_inventory["commercial_scale_finding"].apply(nonblank)
raw_inventory["has_fit_brief_raw"] = raw_inventory["fit_brief_json"].apply(nonblank)
raw_inventory["fit_brief_json_parseable"] = raw_inventory["fit_brief_json"].apply(parseable_json)

raw_inventory["raw_completeness_score"] = (
    raw_inventory["has_funding_raw"].astype(int)
    + raw_inventory["has_payer_raw"].astype(int)
    + raw_inventory["has_outcomes_raw"].astype(int)
    + raw_inventory["has_commercial_scale_raw"].astype(int)
    + raw_inventory["has_fit_brief_raw"].astype(int)
    + raw_inventory["fit_brief_json_parseable"].astype(int)
)

raw_inventory["date_researched_parsed"] = pd.to_datetime(
    raw_inventory["date_researched"],
    errors="coerce"
)

# -----------------------------
# Classify active vs historical/generated source rows
# -----------------------------

STEP17_GENERATED_OR_HISTORICAL_TOKENS = [
    # Generated / rescore / review artifacts should not count as raw evidence.
    "rescore_existing",
    "full_master_rescore",
    "priority_gate_test",
    "parse_retry_test",
    "_retry_test",
    "mae_parse_retry",
    "role_timing_benchmark_1_rescore_existing",
    "human_review_packet",
    "review_packet",
    "_summary",
    "summary_before",
    "before_step22",
    "step26",

    # Explicit test artifacts.
    "test_research_",
    "commercial_scale_test_",

    # Archive/quarantine folders.
    "_archive_generated_test_runs_",

    # Master/dashboard summary files are useful context but should not be counted as raw evidence.
    "health_tech_market_research_summary_master",
    "health_tech_market_map_snapshot",
    "health_tech_dashboard_export",
]

def step17_source_status(row):
    batch_name = safe_text(row.get("batch_name", "")).lower()
    source_file = safe_text(row.get("source_file", "")).lower()
    source_blob = f"{batch_name} {source_file}"

    if "_archive_generated_test_runs_" in source_blob:
        return "historical_archived_generated_or_test_artifact"

    if "health_tech_market_research_summary_master" in source_blob:
        return "summary_metadata_not_raw_evidence"

    if any(token in source_blob for token in STEP17_GENERATED_OR_HISTORICAL_TOKENS):
        return "historical_generated_review_rescore_or_test_artifact"

    return "active_raw_evidence"

raw_inventory["step17_source_status"] = raw_inventory.apply(step17_source_status, axis=1)
raw_inventory["step17_is_active_raw_evidence"] = (
    raw_inventory["step17_source_status"] == "active_raw_evidence"
)

active_raw_inventory = raw_inventory[
    raw_inventory["step17_is_active_raw_evidence"]
].copy()

historical_raw_inventory = raw_inventory[
    ~raw_inventory["step17_is_active_raw_evidence"]
].copy()

# -----------------------------
# Choose best active raw record per company
# -----------------------------

if not active_raw_inventory.empty:
    raw_best = (
        active_raw_inventory
        .sort_values(
            by=[
                "company",
                "raw_completeness_score",
                "date_researched_parsed"
            ],
            ascending=[True, False, False]
        )
        .drop_duplicates(subset=["company"], keep="first")
    )

    active_raw_source_summary = (
        active_raw_inventory
        .groupby("company", dropna=False)
        .agg(
            raw_record_count=("company", "count"),
            raw_source_files=("source_file", lambda x: " | ".join(sorted(set([str(v) for v in x if nonblank(v)])))),
            raw_batch_names=("batch_name", lambda x: " | ".join(sorted(set([str(v) for v in x if nonblank(v)]))))
        )
        .reset_index()
    )

    raw_best = raw_best.merge(active_raw_source_summary, on="company", how="left")

else:
    raw_best = pd.DataFrame(columns=[
        "company",
        "raw_record_count",
        "raw_source_files",
        "raw_batch_names",
        "has_funding_raw",
        "has_payer_raw",
        "has_outcomes_raw",
        "has_commercial_scale_raw",
        "has_fit_brief_raw",
        "fit_brief_json_parseable",
        "raw_completeness_score"
    ])

# -----------------------------
# Add historical/generated metadata summary separately
# -----------------------------

if not historical_raw_inventory.empty:
    historical_source_summary = (
        historical_raw_inventory
        .groupby("company", dropna=False)
        .agg(
            historical_raw_record_count=("company", "count"),
            historical_raw_source_statuses=("step17_source_status", lambda x: " | ".join(sorted(set([str(v) for v in x if nonblank(v)]))))
        )
        .reset_index()
    )

    raw_best = raw_best.merge(historical_source_summary, on="company", how="outer")

else:
    raw_best["historical_raw_record_count"] = 0
    raw_best["historical_raw_source_statuses"] = ""

# Keep raw_best scoped to known companies if possible.
if "audit_base" in globals() and isinstance(audit_base, pd.DataFrame) and "company" in audit_base.columns:
    known_companies = set(audit_base["company"].astype(str))
    raw_best = raw_best[raw_best["company"].astype(str).isin(known_companies)].copy()

# -----------------------------
# Merge audit base with raw evidence inventory
# -----------------------------

raw_best_cols = [
    "company",
    "raw_record_count",
    "raw_source_files",
    "raw_batch_names",
    "historical_raw_record_count",
    "historical_raw_source_statuses",
    "has_funding_raw",
    "has_payer_raw",
    "has_outcomes_raw",
    "has_commercial_scale_raw",
    "has_fit_brief_raw",
    "fit_brief_json_parseable",
    "raw_completeness_score"
]

raw_best_cols = [col for col in raw_best_cols if col in raw_best.columns]

data_depth_audit = audit_base.merge(
    raw_best[raw_best_cols],
    on="company",
    how="left"
)

# Fill missing raw indicators
for col in [
    "raw_record_count",
    "historical_raw_record_count",
    "has_funding_raw",
    "has_payer_raw",
    "has_outcomes_raw",
    "has_commercial_scale_raw",
    "has_fit_brief_raw",
    "fit_brief_json_parseable",
    "raw_completeness_score"
]:
    if col not in data_depth_audit.columns:
        data_depth_audit[col] = 0

    data_depth_audit[col] = data_depth_audit[col].fillna(0)

for col in [
    "raw_source_files",
    "raw_batch_names",
    "historical_raw_source_statuses"
]:
    if col not in data_depth_audit.columns:
        data_depth_audit[col] = ""

    data_depth_audit[col] = data_depth_audit[col].fillna("")

# Make the semantics explicit for dashboard users.
data_depth_audit["active_raw_record_count"] = data_depth_audit["raw_record_count"]
data_depth_audit["excluded_historical_or_generated_record_count"] = data_depth_audit["historical_raw_record_count"]
data_depth_audit["active_raw_batch_names"] = data_depth_audit["raw_batch_names"]
data_depth_audit["active_raw_source_files"] = data_depth_audit["raw_source_files"]

# -----------------------------
# Data depth status logic
# -----------------------------

def depth_status(row):
    final_priority = safe_text(row.get("final_priority_level", ""))
    evidence = row.get("evidence_confidence_score", 0)
    calibration = safe_text(row.get("calibration_flag", ""))

    try:
        evidence = float(evidence)
    except Exception:
        evidence = 0

    if row["raw_record_count"] == 0:
        return "NEEDS RECOVERY: no raw record found"

    if row["raw_completeness_score"] < 5:
        return "NEEDS QA: incomplete raw evidence"

    if not bool(row["fit_brief_json_parseable"]):
        return "NEEDS QA: fit brief JSON not parseable"

    if "CHECK" in calibration:
        return "REVIEW FLAG: calibration check"

    if "REVIEW" in calibration:
        return "REVIEW FLAG: evidence caveat"

    if evidence < 50:
        return "REVIEW FLAG: low evidence confidence"

    if is_priority_or_diligence_target(final_priority) and evidence < 65:
        return "REVIEW FLAG: priority/diligence target with moderate evidence"

    return "OK"

data_depth_audit["data_depth_status"] = data_depth_audit.apply(depth_status, axis=1)

status_rank = {
    "NEEDS RECOVERY: no raw record found": 1,
    "NEEDS QA: incomplete raw evidence": 2,
    "NEEDS QA: fit brief JSON not parseable": 3,
    "REVIEW FLAG: calibration check": 4,
    "REVIEW FLAG: evidence caveat": 5,
    "REVIEW FLAG: low evidence confidence": 6,
    "REVIEW FLAG: priority/diligence target with moderate evidence": 7,
    "OK": 99
}

data_depth_audit["data_depth_status_rank"] = (
    data_depth_audit["data_depth_status"]
    .map(status_rank)
    .fillna(50)
    .astype(int)
)

# -----------------------------
# Sort audit
# -----------------------------

sort_cols = [
    "data_depth_status_rank",
    "final_priority_rank",
    "evidence_confidence_score",
    "company"
]

sort_cols = [col for col in sort_cols if col in data_depth_audit.columns]

data_depth_audit = data_depth_audit.sort_values(
    by=sort_cols,
    ascending=[True, True, True, True][:len(sort_cols)]
).reset_index(drop=True)

# -----------------------------
# Display audit
# -----------------------------

display_cols = [
    "company",
    "market_segment",
    "final_priority_level",
    "priority_source",
    "priority_level",
    "reviewed_priority_level",
    "review_status",
    "evidence_confidence_score",
    "calibration_flag",
    "active_raw_record_count",
    "excluded_historical_or_generated_record_count",
    "raw_completeness_score",
    "has_funding_raw",
    "has_payer_raw",
    "has_outcomes_raw",
    "has_commercial_scale_raw",
    "has_fit_brief_raw",
    "fit_brief_json_parseable",
    "data_depth_status",
    "active_raw_batch_names",
    "historical_raw_source_statuses",
    "active_raw_source_files"
]

display_cols = [col for col in display_cols if col in data_depth_audit.columns]

print("DATA DEPTH AUDIT")
display(data_depth_audit[display_cols])

print("STEP 17 SOURCE CLASSIFICATION SUMMARY")
if "active_raw_record_count" in data_depth_audit.columns:
    print("Active raw evidence records:", int(data_depth_audit["active_raw_record_count"].fillna(0).sum()))
if "excluded_historical_or_generated_record_count" in data_depth_audit.columns:
    print(
        "Excluded historical/generated records:",
        int(data_depth_audit["excluded_historical_or_generated_record_count"].fillna(0).sum())
    )

print("SUMMARY BY DATA DEPTH STATUS")
display(
    data_depth_audit
    .groupby("data_depth_status", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .assign(
        data_depth_status_rank=lambda df: df["data_depth_status"].map(status_rank).fillna(50).astype(int)
    )
    .sort_values("data_depth_status_rank")
    .drop(columns=["data_depth_status_rank"])
)

print("SUMMARY BY FINAL PRIORITY AND DATA DEPTH STATUS")
display(
    data_depth_audit
    .groupby(["final_priority_level", "data_depth_status"], dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
    .assign(
        final_priority_rank=lambda df: df["final_priority_level"].apply(priority_rank),
        data_depth_status_rank=lambda df: df["data_depth_status"].map(status_rank).fillna(50).astype(int)
    )
    .sort_values(["final_priority_rank", "data_depth_status_rank"])
    .drop(columns=["final_priority_rank", "data_depth_status_rank"])
)


# =============================================================================

# STEP 18 - Segment coverage audit

# =============================================================================

# STEP 18 - Segment coverage audit

# =============================================================================

# 18 - Segment coverage audit
# Purpose:
# - Audit whether each market segment has enough companies for a useful directional read
# - Use final_priority_level / final_priority_rank from Step 12B as the priority source of truth
# - Treat P0, P1, and P2 as priority-or-diligence companies
# - Treat P3 and P4 as watch/low-priority companies
# - Add P0-aware priority counts for dashboard export
#
# Run after:
# 12B -> 13 -> 14 -> 15 -> 16 -> 17
#
# Then continue:
# 19 -> 19A

import pandas as pd
import numpy as np
import re
import sys
from pathlib import Path

# -----------------------------
# Import shared priority helper
# -----------------------------

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"

if SRC_DIR.exists():
    src_path = str(SRC_DIR)
    if src_path not in sys.path:
        sys.path.insert(0, src_path)

try:
    from health_tech_research_agent.priority import (
        apply_priority_fields,
        extract_priority_code,
        priority_rank,
        safe_text,
    )
except Exception as e:
    raise ImportError(
        "STOP: Could not import shared priority helpers. "
        "Run the GitHub pull cell and Step 12B first."
    ) from e

# -----------------------------
# Validate inputs
# -----------------------------

if "market_map_df" not in globals() or not isinstance(market_map_df, pd.DataFrame) or market_map_df.empty:
    raise NameError("STOP: market_map_df not found or empty. Run Step 14 first.")

coverage_source_df = market_map_df.copy()
coverage_source_df = apply_priority_fields(coverage_source_df)

# -----------------------------
# Helpers
# -----------------------------

from health_tech_research_agent.dashboard import (
    existing_cols,
    join_unique,
    coverage_status_from_counts,
    coverage_status_rank,
    companies_needed_for_directional_read,
    companies_needed_for_stronger_read,
)

# -----------------------------
# Required columns
# -----------------------------

required_defaults = {
    "company": "",
    "market_segment": "Unmapped",
    "strategic_bucket": "",
    "final_priority_level": "",
    "priority_source": "",
    "final_priority_code": "",
    "final_priority_rank": 99,
    "thesis_fit_score": np.nan,
    "pmf_scale_score": np.nan,
    "evidence_confidence_score": np.nan,
    "katelynd_role_fit_score": np.nan,
    "operator_timing_score": np.nan
}

for col_name, default_value in required_defaults.items():
    if col_name not in coverage_source_df.columns:
        coverage_source_df[col_name] = default_value

for score_col in [
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]:
    coverage_source_df[score_col] = pd.to_numeric(
        coverage_source_df[score_col],
        errors="coerce"
    )

coverage_source_df["final_priority_code"] = coverage_source_df["final_priority_level"].apply(extract_priority_code)
coverage_source_df["final_priority_rank"] = coverage_source_df["final_priority_level"].apply(priority_rank)

# -----------------------------
# Build segment coverage audit
# -----------------------------

segment_rows = []

for segment, segment_df in coverage_source_df.groupby("market_segment", dropna=False):
    working_df = segment_df.copy()

    company_count = working_df["company"].nunique()

    p0_count = int((working_df["final_priority_code"] == "P0").sum())
    p1_count = int((working_df["final_priority_code"] == "P1").sum())
    p2_count = int((working_df["final_priority_code"] == "P2").sum())
    p3_count = int((working_df["final_priority_code"] == "P3").sum())
    p4_count = int((working_df["final_priority_code"] == "P4").sum())

    priority_or_diligence_count = p0_count + p1_count + p2_count
    watch_or_low_count = p3_count + p4_count

    human_reviewed_count = int(
        working_df["priority_source"]
        .astype(str)
        .str.lower()
        .eq("human reviewed")
        .sum()
    )

    best_rank = int(working_df["final_priority_rank"].min()) if not working_df.empty else 99

    best_priority_levels = (
        working_df.loc[working_df["final_priority_rank"] == best_rank, "final_priority_level"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    best_final_priority_level = best_priority_levels[0] if best_priority_levels else ""

    top_companies_df = working_df.sort_values(
        by=[
            "final_priority_rank",
            "thesis_fit_score",
            "pmf_scale_score",
            "katelynd_role_fit_score",
            "operator_timing_score",
            "evidence_confidence_score",
            "company"
        ],
        ascending=[True, False, False, False, False, False, True]
    ).copy()

    best_companies = join_unique(top_companies_df["company"].head(5).tolist(), max_items=5)

    status = coverage_status_from_counts(company_count, priority_or_diligence_count)

    segment_rows.append({
        "market_segment": safe_text(segment) if safe_text(segment) else "Unmapped",
        "coverage_status": status,
        "coverage_status_rank": coverage_status_rank(status),
        "company_count": int(company_count),
        "priority_or_diligence_count": int(priority_or_diligence_count),
        "watch_or_low_count": int(watch_or_low_count),
        "p0_count": p0_count,
        "p1_count": p1_count,
        "p2_count": p2_count,
        "p3_count": p3_count,
        "p4_count": p4_count,
        "human_reviewed_count": human_reviewed_count,
        "avg_thesis_fit": round(working_df["thesis_fit_score"].mean(), 1),
        "avg_pmf_scale": round(working_df["pmf_scale_score"].mean(), 1),
        "avg_evidence_confidence": round(working_df["evidence_confidence_score"].mean(), 1),
        "avg_katelynd_role_fit": round(working_df["katelynd_role_fit_score"].mean(), 1),
        "avg_operator_timing": round(working_df["operator_timing_score"].mean(), 1),
        "best_final_priority_level": best_final_priority_level,
        "best_final_priority_rank": best_rank,
        "current_best_companies": best_companies,
        "companies_needed_for_directional_read": companies_needed_for_directional_read(
            company_count,
            priority_or_diligence_count
        ),
        "companies_needed_for_stronger_read": companies_needed_for_stronger_read(
            company_count,
            priority_or_diligence_count
        )
    })

segment_coverage_audit = pd.DataFrame(segment_rows)

segment_coverage_audit = segment_coverage_audit.sort_values(
    by=[
        "coverage_status_rank",
        "best_final_priority_rank",
        "priority_or_diligence_count",
        "company_count",
        "avg_thesis_fit",
        "market_segment"
    ],
    ascending=[True, True, False, False, False, True]
).reset_index(drop=True)

# -----------------------------
# Output
# -----------------------------

print("Segment coverage audit built.")
print("Segment count:", segment_coverage_audit["market_segment"].nunique())
print("Shape:", segment_coverage_audit.shape)

print("\nCoverage status summary:")
display(
    segment_coverage_audit
    .groupby("coverage_status", dropna=False)
    .agg(segment_count=("market_segment", "nunique"))
    .reset_index()
    .sort_values("segment_count", ascending=False)
)

print("\nPriority / diligence coverage summary:")
display(
    segment_coverage_audit[[
        "market_segment",
        "coverage_status",
        "company_count",
        "priority_or_diligence_count",
        "p0_count",
        "p1_count",
        "p2_count",
        "p3_count",
        "p4_count",
        "best_final_priority_level",
        "current_best_companies"
    ]]
)

print("\nP0 segment check:")
display(
    segment_coverage_audit[
        segment_coverage_audit["p0_count"] > 0
    ][[
        "market_segment",
        "company_count",
        "priority_or_diligence_count",
        "p0_count",
        "p1_count",
        "p2_count",
        "best_final_priority_level",
        "current_best_companies"
    ]]
)


# =============================================================================

# STEP 19 - Export dashboard workbook

# =============================================================================

# STEP 19 - Export dashboard workbook

# =============================================================================

# Purpose:

# - Export focused dashboard workbook

# - Keep Master Dashboard clean

# - Move priority traceability to Priority Logic Audit

# 19 - Export dashboard workbook
# Purpose:
# - Create a clean Excel workbook from the final-priority dashboard
# - Use final_priority_level as the dashboard priority source of truth
# - Support the P0/P1/P2/P3/P4 priority model
# - Keep Master Dashboard decision-ready
# - Move priority traceability fields to Priority Logic Audit
# - Save to Google Drive
# - Define dashboard_workbook_path for Step 19A formatting
#
# After this, run Step 19A to format and download the workbook.

import pandas as pd
import numpy as np
import sys
from pathlib import Path
from datetime import datetime
from google.colab import drive
import shutil

drive.mount("/content/drive")

# -----------------------------
# Choose dashboard dataframe
# -----------------------------

if "market_map_df" in globals() and isinstance(market_map_df, pd.DataFrame) and not market_map_df.empty:
    dashboard_df = market_map_df.copy()
elif "master_df" in globals() and isinstance(master_df, pd.DataFrame) and not master_df.empty:
    dashboard_df = master_df.copy()
elif "master_summary_df" in globals() and isinstance(master_summary_df, pd.DataFrame) and not master_summary_df.empty:
    dashboard_df = master_summary_df.copy()
else:
    raise NameError("STOP: No dashboard dataframe found. Run Steps 13–14 first.")

# -----------------------------
# Paths
# -----------------------------

drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
drive_folder.mkdir(parents=True, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

local_export_path = Path(f"health_tech_dashboard_export_{timestamp}.xlsx")
drive_export_path = drive_folder / f"health_tech_dashboard_export_{timestamp}.xlsx"

# Step 19A uses these variables.
dashboard_workbook_path = local_export_path
output_workbook_path = local_export_path

# -----------------------------
# Import shared priority helper
# -----------------------------

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"

if SRC_DIR.exists():
    src_path = str(SRC_DIR)
    if src_path not in sys.path:
        sys.path.insert(0, src_path)

try:
    from health_tech_research_agent.priority import (
        apply_priority_fields,
        extract_priority_code,
        priority_rank,
        safe_text,
    )
except Exception as e:
    raise ImportError(
        "STOP: Could not import shared priority helpers. "
        "Run the GitHub pull cell and Step 12B first."
    ) from e

# -----------------------------
# Helpers
# -----------------------------

from health_tech_research_agent.dashboard import (
    existing_cols,
    safe_sort,
    contains_priority,
    join_unique,
)

# -----------------------------
# Ensure final priority fields exist
# -----------------------------

dashboard_df = apply_priority_fields(dashboard_df)

# Backward-compatible aliases. New dashboard logic should use final_priority_level / final_priority_rank.
dashboard_df["decision_priority"] = dashboard_df["final_priority_level"]
dashboard_df["decision_priority_sort"] = dashboard_df["final_priority_rank"]

# -----------------------------
# Ensure required support fields exist
# -----------------------------

default_columns = {
    "company": "",
    "market_segment": "Unmapped",
    "primary_market_segment_code": "",
    "primary_market_segment": "",
    "subsegment_tags": "",
    "product_model_tags": "",
    "distribution_model_tags": "",
    "data_input_tags": "",
    "taxonomy_assignment_method": "",
    "taxonomy_assignment_basis": "",
    "strategic_bucket": "Unmapped",
    "calibration_flag": "",
    "review_status": "",
    "review_notes": "",
    "priority_review_note": "",
    "priority_level": "",
    "reviewed_priority_level": "",
    "priority_source": "Auto Adjudicated",
    "final_priority_level": "",
    "final_priority_code": "",
    "final_priority_rank": 99,
    "final_recommendation": "",
    "business_model_classification": "",
    "commercial_scale_assessment": "",
    "pmf_scale_assessment": "",
    "commercial_scale_finding": "",
    "payer_institutional_finding": "",
    "outcomes_finding": "",
    "funding_finding": "",
    "final_takeaway": "",
    "thesis_fit_score": np.nan,
    "pmf_scale_score": np.nan,
    "evidence_confidence_score": np.nan,
    "katelynd_role_fit_score": np.nan,
    "operator_timing_score": np.nan
}

for col_name, default_value in default_columns.items():
    if col_name not in dashboard_df.columns:
        dashboard_df[col_name] = default_value

dashboard_df["final_priority_code"] = dashboard_df["final_priority_level"].apply(extract_priority_code)

dashboard_df["final_priority_rank"] = pd.to_numeric(
    dashboard_df["final_priority_rank"],
    errors="coerce"
).fillna(
    dashboard_df["final_priority_level"].apply(priority_rank)
).fillna(99).astype(int)

for score_col in [
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score"
]:
    dashboard_df[score_col] = pd.to_numeric(
        dashboard_df[score_col],
        errors="coerce"
    )

# -----------------------------
# Force clean P0-P4 final priority ranking from shared helper
# -----------------------------

dashboard_df["final_priority_code"] = dashboard_df["final_priority_level"].apply(extract_priority_code)
dashboard_df["final_priority_rank"] = dashboard_df["final_priority_level"].apply(priority_rank)

dashboard_df["decision_priority"] = dashboard_df["final_priority_level"]
dashboard_df["decision_priority_sort"] = dashboard_df["final_priority_rank"]

# -----------------------------
# Master Dashboard
# -----------------------------
# Clean decision view. Priority plumbing lives in Priority Logic Audit.

master_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "market_segment",
    "primary_market_segment_code",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_assignment_method",
    "strategic_bucket",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "business_model_classification",
    "commercial_scale_assessment",
    "pmf_scale_assessment",
    "calibration_flag",
    "final_takeaway"
])

master_view = safe_sort(
    dashboard_df,
    [
        "final_priority_rank",
        "thesis_fit_score",
        "pmf_scale_score",
        "katelynd_role_fit_score",
        "operator_timing_score",
        "evidence_confidence_score"
    ],
    [True, False, False, False, False, False]
)[master_cols]

# -----------------------------
# Priority Focus
# -----------------------------
# P0/P1/P2 companies plus companies with calibration flags.

priority_focus_source = dashboard_df[
    dashboard_df["final_priority_level"].apply(lambda value: contains_priority(value, ["P0", "P1", "P2"]))
    | dashboard_df["calibration_flag"].astype(str).str.strip().ne("")
].copy()

priority_focus_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "market_segment",
    "primary_market_segment_code",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_assignment_method",
    "strategic_bucket",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "business_model_classification",
    "commercial_scale_assessment",
    "calibration_flag",
    "final_takeaway"
])

priority_focus = safe_sort(
    priority_focus_source,
    [
        "final_priority_rank",
        "thesis_fit_score",
        "pmf_scale_score",
        "katelynd_role_fit_score",
        "operator_timing_score"
    ],
    [True, False, False, False, False]
)[priority_focus_cols]

# -----------------------------
# Segment Summary
# -----------------------------
# Build directly from dashboard_df so P0 is always included even if older Step 15 outputs exist.

segment_summary_export = (
    dashboard_df
    .groupby("market_segment", dropna=False)
    .agg(
        company_count=("company", "nunique"),
        p0_count=("final_priority_code", lambda x: (x.astype(str).str.upper() == "P0").sum()),
        p1_count=("final_priority_code", lambda x: (x.astype(str).str.upper() == "P1").sum()),
        p2_count=("final_priority_code", lambda x: (x.astype(str).str.upper() == "P2").sum()),
        p3_count=("final_priority_code", lambda x: (x.astype(str).str.upper() == "P3").sum()),
        p4_count=("final_priority_code", lambda x: (x.astype(str).str.upper() == "P4").sum()),
        human_reviewed_count=("priority_source", lambda x: x.astype(str).str.contains("Human Reviewed", case=False, na=False).sum()),
        avg_thesis_fit=("thesis_fit_score", "mean"),
        avg_pmf_scale=("pmf_scale_score", "mean"),
        avg_evidence_confidence=("evidence_confidence_score", "mean"),
        avg_katelynd_role_fit=("katelynd_role_fit_score", "mean"),
        avg_operator_timing=("operator_timing_score", "mean"),
        best_final_priority_rank=("final_priority_rank", "min")
    )
    .reset_index()
)

for avg_col in [
    "avg_thesis_fit",
    "avg_pmf_scale",
    "avg_evidence_confidence",
    "avg_katelynd_role_fit",
    "avg_operator_timing"
]:
    if avg_col in segment_summary_export.columns:
        segment_summary_export[avg_col] = segment_summary_export[avg_col].round(1)

best_priority_lookup = (
    dashboard_df
    .sort_values(["market_segment", "final_priority_rank", "thesis_fit_score"], ascending=[True, True, False])
    .groupby("market_segment", dropna=False)
    .agg(
        best_final_priority_level=("final_priority_level", "first"),
        best_companies=("company", lambda x: join_unique(x, max_items=5))
    )
    .reset_index()
)

segment_summary_export = segment_summary_export.merge(
    best_priority_lookup,
    on="market_segment",
    how="left"
)

segment_summary_export = safe_sort(
    segment_summary_export,
    ["best_final_priority_rank", "p0_count", "p1_count", "p2_count", "avg_thesis_fit"],
    [True, False, False, False, False]
)

# -----------------------------
# Companies by Segment
# -----------------------------

company_by_segment_cols = existing_cols(dashboard_df, [
    "market_segment",
    "company",
    "strategic_bucket",
    "final_priority_level",
    "priority_source",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_takeaway"
])

company_by_segment_export = safe_sort(
    dashboard_df[company_by_segment_cols],
    [
        "market_segment",
        "final_priority_rank",
        "thesis_fit_score",
        "operator_timing_score",
        "pmf_scale_score"
    ],
    [True, True, False, False, False]
)

# -----------------------------
# Commercial Scale Review
# -----------------------------

commercial_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "market_segment",
    "strategic_bucket",
    "pmf_scale_score",
    "evidence_confidence_score",
    "business_model_classification",
    "commercial_scale_assessment",
    "commercial_scale_finding",
    "payer_institutional_finding",
    "outcomes_finding",
    "calibration_flag",
    "final_takeaway"
])

commercial_review = safe_sort(
    dashboard_df[commercial_cols],
    ["final_priority_rank", "pmf_scale_score", "company"],
    [True, False, True]
)

# -----------------------------
# Data Depth Audit
# -----------------------------

if "data_depth_audit" in globals() and isinstance(data_depth_audit, pd.DataFrame) and not data_depth_audit.empty:
    data_depth_audit_export = data_depth_audit.copy()
else:
    audit_rows = []

    for _, row in dashboard_df.iterrows():
        audit_rows.append({
            "company": row.get("company", ""),
            "final_priority_level": row.get("final_priority_level", ""),
            "priority_source": row.get("priority_source", ""),
            "market_segment": row.get("market_segment", ""),
            "strategic_bucket": row.get("strategic_bucket", ""),
            "has_commercial_scale_assessment": safe_text(row.get("commercial_scale_assessment", "")) != "",
            "has_commercial_scale_finding": safe_text(row.get("commercial_scale_finding", "")) != "",
            "has_payer_institutional_finding": safe_text(row.get("payer_institutional_finding", "")) != "",
            "has_outcomes_finding": safe_text(row.get("outcomes_finding", "")) != "",
            "has_funding_finding": safe_text(row.get("funding_finding", "")) != "",
            "has_business_model_classification": safe_text(row.get("business_model_classification", "")) != "",
            "has_calibration_flag": safe_text(row.get("calibration_flag", "")) != ""
        })

    data_depth_audit_export = pd.DataFrame(audit_rows)

# -----------------------------
# Segment Coverage Audit
# -----------------------------

if "segment_coverage_audit" in globals() and isinstance(segment_coverage_audit, pd.DataFrame) and not segment_coverage_audit.empty:
    segment_coverage_audit_export = segment_coverage_audit.copy()
else:
    segment_coverage_audit_export = pd.DataFrame()

# -----------------------------
# Priority Logic Audit
# -----------------------------

priority_logic_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "priority_review_note",
    "priority_level",
    "reviewed_priority_level",
    "final_priority_code",
    "final_priority_rank",
    "decision_priority",
    "decision_priority_sort",
    "review_status",
    "review_notes",
    "calibration_flag",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "market_segment",
    "strategic_bucket",
    "final_recommendation",
    "final_takeaway"
])

priority_logic_audit = safe_sort(
    dashboard_df[priority_logic_cols],
    ["final_priority_rank", "company"],
    [True, True]
)

# -----------------------------
# Read Me
# -----------------------------

read_me = pd.DataFrame([
    {
        "sheet": "Master Dashboard",
        "description": "Main decision dashboard. Uses Final Priority Level as the source of truth. Internal priority plumbing is intentionally excluded."
    },
    {
        "sheet": "Priority Focus",
        "description": "P0/P1/P2 companies plus any company with calibration flags."
    },
    {
        "sheet": "Segment Summary",
        "description": "P0-P4 segment-level scoring and priority roll-up."
    },
    {
        "sheet": "Companies by Segment",
        "description": "Company-level segment view sorted by final priority and fit scores."
    },
    {
        "sheet": "Commercial Scale Review",
        "description": "Commercial-scale and monetization evidence for revenue-quality review."
    },
    {
        "sheet": "Data Depth Audit",
        "description": "QA view showing whether key research evidence and fit-brief fields are populated."
    },
    {
        "sheet": "Segment Coverage Audit",
        "description": "Segment mapping sufficiency audit, included when Step 18 output is available."
    },
    {
        "sheet": "Priority Logic Audit",
        "description": "Traceability view showing automated priority, reviewed priority, final priority, source, and review notes."
    }
])


# -----------------------------
# Candidate Priority V4.2 views
# -----------------------------
# CANDIDATE_PRIORITY_DASHBOARD_V42_PATCH_START
# Purpose:
# - Preserve final_priority_level as the human-reviewed dashboard source of truth.
# - Add candidate_priority_level as the V4.2 recommendation layer.
# - Surface P0/P1 candidate-priority targets in dedicated dashboard sheets.
# - Do not overwrite final priority.

candidate_default_columns = {
    "candidate_priority_level": "",
    "candidate_priority_reason": "",
    "candidate_priority_framework_version": "",
    "candidate_priority_source": "",
    "candidate_priority_audit_file": "",
    "candidate_priority_updated_at": "",
    "target_archetype": "",
    "scale_path_quality": "",
    "katelynd_capability_fit_score": np.nan,
    "operator_agency_entry_score": np.nan,
    "reset_or_restructure_signal": "",
    "reset_or_restructure_basis": "",
    "commercial_scale_signal_inferred": np.nan,
    "institutional_distribution_signal_inferred": np.nan,
    "outcomes_signal_inferred": np.nan,
    "candidate_priority_rank": 99,
    "candidate_priority_code": ""
}

for col_name, default_value in candidate_default_columns.items():
    if col_name not in dashboard_df.columns:
        dashboard_df[col_name] = default_value

dashboard_df["candidate_priority_code"] = dashboard_df["candidate_priority_level"].apply(extract_priority_code)
dashboard_df["candidate_priority_rank"] = dashboard_df["candidate_priority_level"].apply(priority_rank)

for score_col in [
    "katelynd_capability_fit_score",
    "operator_agency_entry_score",
    "commercial_scale_signal_inferred",
    "institutional_distribution_signal_inferred",
    "outcomes_signal_inferred",
    "candidate_priority_rank"
]:
    dashboard_df[score_col] = pd.to_numeric(
        dashboard_df[score_col],
        errors="coerce"
    )

dashboard_df["candidate_priority_rank"] = dashboard_df["candidate_priority_rank"].fillna(99).astype(int)

def candidate_priority_movement(row):
    candidate_rank = row.get("candidate_priority_rank", 99)
    final_rank = row.get("final_priority_rank", 99)

    try:
        candidate_rank = int(candidate_rank)
    except Exception:
        candidate_rank = 99

    try:
        final_rank = int(final_rank)
    except Exception:
        final_rank = 99

    if candidate_rank < final_rank:
        return "Candidate promotes vs final"
    if candidate_rank > final_rank:
        return "Candidate demotes vs final"
    return "No change"

dashboard_df["candidate_vs_final_priority_movement"] = dashboard_df.apply(candidate_priority_movement, axis=1)

candidate_master_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "candidate_priority_level",
    "candidate_vs_final_priority_movement",
    "target_archetype",
    "scale_path_quality",
    "katelynd_capability_fit_score",
    "operator_agency_entry_score",
    "reset_or_restructure_signal",
    "market_segment",
    "strategic_bucket",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "katelynd_role_fit_score",
    "operator_timing_score",
    "final_recommendation",
    "business_model_classification",
    "commercial_scale_assessment",
    "pmf_scale_assessment",
    "calibration_flag",
    "final_takeaway",
    "candidate_priority_reason",
    "reset_or_restructure_basis"
])

master_view = safe_sort(
    dashboard_df,
    [
        "candidate_priority_rank",
        "operator_agency_entry_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "final_priority_rank",
        "company"
    ],
    [True, False, False, False, True, True]
)[candidate_master_cols]

candidate_focus_source = dashboard_df[
    dashboard_df["candidate_priority_level"].apply(lambda value: contains_priority(value, ["P0", "P1"]))
].copy()

candidate_focus_cols = existing_cols(dashboard_df, [
    "company",
    "candidate_priority_level",
    "final_priority_level",
    "candidate_vs_final_priority_movement",
    "target_archetype",
    "scale_path_quality",
    "katelynd_capability_fit_score",
    "operator_agency_entry_score",
    "reset_or_restructure_signal",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "market_segment",
    "strategic_bucket",
    "business_model_classification",
    "commercial_scale_assessment",
    "payer_institutional_finding",
    "outcomes_finding",
    "candidate_priority_reason",
    "reset_or_restructure_basis",
    "final_takeaway"
])

candidate_p0_p1_focus = safe_sort(
    candidate_focus_source,
    [
        "candidate_priority_rank",
        "operator_agency_entry_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "company"
    ],
    [True, False, False, False, True]
)[candidate_focus_cols]

candidate_p0 = candidate_p0_p1_focus[
    candidate_p0_p1_focus["candidate_priority_level"].apply(lambda value: contains_priority(value, ["P0"]))
].copy()

candidate_p1 = candidate_p0_p1_focus[
    candidate_p0_p1_focus["candidate_priority_level"].apply(lambda value: contains_priority(value, ["P1"]))
].copy()

priority_comparison_cols = existing_cols(dashboard_df, [
    "company",
    "final_priority_level",
    "priority_source",
    "candidate_priority_level",
    "candidate_vs_final_priority_movement",
    "target_archetype",
    "scale_path_quality",
    "katelynd_capability_fit_score",
    "operator_agency_entry_score",
    "reset_or_restructure_signal",
    "company_maturity_read",
    "stage_timing_fit",
    "likely_agency_level",
    "thesis_fit_score",
    "pmf_scale_score",
    "evidence_confidence_score",
    "market_segment",
    "strategic_bucket",
    "candidate_priority_reason",
    "reset_or_restructure_basis"
])

priority_comparison = safe_sort(
    dashboard_df[priority_comparison_cols],
    [
        "candidate_priority_rank",
        "operator_agency_entry_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "company"
    ],
    [True, False, False, False, True]
)

candidate_priority_summary = (
    dashboard_df
    .groupby("candidate_priority_level", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
)

candidate_priority_summary["candidate_priority_rank"] = candidate_priority_summary["candidate_priority_level"].apply(priority_rank)
candidate_priority_summary = candidate_priority_summary.sort_values("candidate_priority_rank").drop(columns=["candidate_priority_rank"])

final_priority_summary = (
    dashboard_df
    .groupby("final_priority_level", dropna=False)
    .agg(company_count=("company", "nunique"))
    .reset_index()
)

final_priority_summary["final_priority_rank"] = final_priority_summary["final_priority_level"].apply(priority_rank)
final_priority_summary = final_priority_summary.sort_values("final_priority_rank").drop(columns=["final_priority_rank"])

read_me = pd.concat([
    read_me,
    pd.DataFrame([
        {
            "sheet": "Candidate P0 P1 Focus",
            "description": "Candidate P0/P1 companies from V4.2 priority logic. This is a recommendation layer, not the human-reviewed final priority."
        },
        {
            "sheet": "Candidate P0",
            "description": "Potential active pursuit targets. Requires human review before promotion into final priority."
        },
        {
            "sheet": "Candidate P1",
            "description": "High-priority diligence candidates. Requires human review before promotion into final priority."
        },
        {
            "sheet": "Priority Comparison",
            "description": "Compares human-reviewed final priority with V4.2 candidate priority and flags promotions/demotions."
        },
        {
            "sheet": "Candidate Summary",
            "description": "Counts by candidate_priority_level."
        },
        {
            "sheet": "Final Summary",
            "description": "Counts by human-reviewed final_priority_level."
        }
    ])
], ignore_index=True)

# CANDIDATE_PRIORITY_DASHBOARD_V42_PATCH_END


# -----------------------------
# Export workbook
# -----------------------------

with pd.ExcelWriter(local_export_path, engine="openpyxl") as writer:
    read_me.to_excel(writer, sheet_name="Read Me", index=False)
    master_view.to_excel(writer, sheet_name="Master Dashboard", index=False)
    priority_focus.to_excel(writer, sheet_name="Priority Focus", index=False)
    candidate_p0_p1_focus.to_excel(writer, sheet_name="Candidate P0 P1 Focus", index=False)
    candidate_p0.to_excel(writer, sheet_name="Candidate P0", index=False)
    candidate_p1.to_excel(writer, sheet_name="Candidate P1", index=False)
    priority_comparison.to_excel(writer, sheet_name="Priority Comparison", index=False)
    candidate_priority_summary.to_excel(writer, sheet_name="Candidate Summary", index=False)
    final_priority_summary.to_excel(writer, sheet_name="Final Summary", index=False)
    segment_summary_export.to_excel(writer, sheet_name="Segment Summary", index=False)
    company_by_segment_export.to_excel(writer, sheet_name="Companies by Segment", index=False)
    commercial_review.to_excel(writer, sheet_name="Commercial Scale Review", index=False)
    data_depth_audit_export.to_excel(writer, sheet_name="Data Depth Audit", index=False)

    if not segment_coverage_audit_export.empty:
        segment_coverage_audit_export.to_excel(writer, sheet_name="Segment Coverage Audit", index=False)

    priority_logic_audit.to_excel(writer, sheet_name="Priority Logic Audit", index=False)

shutil.copy(local_export_path, drive_export_path)

print("Dashboard export complete.")
print("Local file:", local_export_path)
print("Drive file:", drive_export_path)

print("\nWorkbook variable for Step 19A:")
print("dashboard_workbook_path =", dashboard_workbook_path)

print("\nExported sheets:")
exported_sheets = [
    "Read Me",
    "Master Dashboard",
    "Priority Focus",
    "Candidate P0 P1 Focus",
    "Candidate P0",
    "Candidate P1",
    "Priority Comparison",
    "Candidate Summary",
    "Final Summary",
    "Segment Summary",
    "Companies by Segment",
    "Commercial Scale Review",
    "Data Depth Audit"
]

if not segment_coverage_audit_export.empty:
    exported_sheets.append("Segment Coverage Audit")

exported_sheets.append("Priority Logic Audit")

for sheet in exported_sheets:
    print("-", sheet)

print("\nPriority Focus includes:")
print("- P0 companies")
print("- P1 companies")
print("- P2 companies")
print("- Any company with calibration flags")


# =============================================================================

# STEP 19A - Format exported dashboard workbook

# =============================================================================

# Purpose:

# - Convert headers to readable labels

# - Apply filters

# - Freeze header row

# - Set column widths

# - Wrap text

# - Save/copy/download formatted workbook

# 19A - Format exported dashboard workbook
# Purpose:
# - Convert snake_case headers to readable labels in the workbook only
# - Apply filters
# - Freeze header row
# - Set readable column widths
# - Wrap text for long cells
# - Save formatted workbook locally
# - Copy formatted workbook to Google Drive
# - Download formatted workbook

from pathlib import Path
import re
import shutil
from google.colab import files
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# Locate workbook
# -----------------------------

if "dashboard_workbook_path" in globals():
    workbook_path = Path(dashboard_workbook_path)
elif "market_map_workbook_path" in globals():
    workbook_path = Path(market_map_workbook_path)
elif "output_workbook_path" in globals():
    workbook_path = Path(output_workbook_path)
elif "local_export_path" in globals():
    workbook_path = Path(local_export_path)
else:
    raise NameError(
        "STOP: Could not find dashboard_workbook_path, market_map_workbook_path, "
        "output_workbook_path, or local_export_path."
    )

if not workbook_path.exists():
    raise FileNotFoundError(f"STOP: Workbook not found: {workbook_path}")

if "drive_export_path" in globals():
    formatted_drive_path = Path(drive_export_path)
else:
    drive_folder = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
    drive_folder.mkdir(parents=True, exist_ok=True)
    formatted_drive_path = drive_folder / workbook_path.name

# -----------------------------
# Header formatting helpers
# -----------------------------

def friendly_header(header):
    if header is None:
        return ""

    text = str(header).strip()

    explicit_map = {
        "company": "Company",
        "priority_level": "Priority Level",
        "reviewed_priority_level": "Reviewed Priority Level",
        "final_priority_level": "Final Priority Level",
        "priority_source": "Priority Source",
        "priority_review_note": "Priority Review Note",
        "final_priority_code": "Final Priority Code",
        "final_priority_rank": "Final Priority Rank",
        "decision_priority": "Decision Priority",
        "decision_priority_sort": "Decision Priority Sort",
        "review_status": "Review Status",
        "review_notes": "Review Notes",
        "thesis_fit_score": "Thesis Fit Score",
        "pmf_scale_score": "PMF / Scale Score",
        "evidence_confidence_score": "Evidence Confidence Score",
        "katelynd_role_fit_score": "Katelynd Role Fit Score",
        "operator_timing_score": "Operator Timing Score",
        "final_recommendation": "Final Recommendation",
        "business_model_classification": "Business Model Classification",
        "commercial_scale_assessment": "Commercial Scale Assessment",
        "pmf_scale_assessment": "PMF / Scale Assessment",
        "scale_signal_assessment": "Scale Signal Assessment",
        "calibration_flag": "Calibration Flag",
        "final_takeaway": "Final Takeaway",
        "date_researched": "Date Researched",
        "funding_finding": "Funding Finding",
        "payer_institutional_finding": "Payer / Institutional Finding",
        "outcomes_finding": "Outcomes Finding",
        "commercial_scale_finding": "Commercial Scale Finding",
        "fit_brief_json": "Fit Brief JSON",
        "segment": "Segment",
        "market_segment": "Market Segment",
        "strategic_bucket": "Strategic Bucket",
        "company_stage": "Company Stage",
        "funding_stage": "Funding Stage",
        "revenue_estimate": "Revenue Estimate",
        "employee_count": "Employee Count",
        "source_urls": "Source URLs",
        "notes": "Notes",
        "company_count": "Company Count",
        "p1_count": "P1 Count",
        "p2_count": "P2 Count",
        "p3_count": "P3 Count",
        "p4_count": "P4 Count",
        "human_reviewed_count": "Human Reviewed Count",
        "avg_thesis_fit": "Avg. Thesis Fit",
        "avg_pmf_scale": "Avg. PMF / Scale",
        "avg_evidence_confidence": "Avg. Evidence Confidence",
        "avg_katelynd_role_fit": "Avg. Katelynd Role Fit",
        "avg_operator_timing": "Avg. Operator Timing",
        "best_final_priority_level": "Best Final Priority Level",
        "best_final_priority_rank": "Best Final Priority Rank",
        "best_companies": "Best Companies",
        "current_best_companies": "Current Best Companies",
        "coverage_status": "Coverage Status",
        "companies_needed_for_directional_read": "Companies Needed for Directional Read",
        "companies_needed_for_stronger_read": "Companies Needed for Stronger Read",
        "data_depth_status": "Data Depth Status",
        "raw_record_count": "Raw Record Count",
        "raw_completeness_score": "Raw Completeness Score",
        "has_funding_raw": "Has Funding Raw",
        "has_payer_raw": "Has Payer Raw",
        "has_outcomes_raw": "Has Outcomes Raw",
        "has_commercial_scale_raw": "Has Commercial Scale Raw",
        "has_fit_brief_raw": "Has Fit Brief Raw",
        "fit_brief_json_parseable": "Fit Brief JSON Parseable",
        "raw_source_files": "Raw Source Files",
        "raw_batch_names": "Raw Batch Names"
    }

    if text in explicit_map:
        return explicit_map[text]

    text = text.replace("_", " ").strip()
    text = re.sub(r"\s+", " ", text)

    small_words = {"and", "or", "to", "of", "in", "for", "with", "by"}

    words = []
    for i, word in enumerate(text.split(" ")):
        lower = word.lower()

        if i > 0 and lower in small_words:
            words.append(lower)
        elif lower == "pmf":
            words.append("PMF")
        elif lower == "d2c":
            words.append("D2C")
        elif lower == "b2b2c":
            words.append("B2B2C")
        elif lower == "arr":
            words.append("ARR")
        elif lower == "cac":
            words.append("CAC")
        elif lower == "api":
            words.append("API")
        elif lower == "json":
            words.append("JSON")
        elif lower == "qa":
            words.append("QA")
        else:
            words.append(word.capitalize())

    return " ".join(words)

def width_for_header(header):
    header_lower = str(header).lower()

    very_long_text_terms = [
        "assessment",
        "finding",
        "takeaway",
        "rationale",
        "notes",
        "claim",
        "source",
        "json",
        "description",
        "summary",
        "commercial",
        "outcomes",
        "institutional",
        "companies needed",
        "current best companies",
        "best companies",
        "raw source files",
        "raw batch names"
    ]

    medium_text_terms = [
        "classification",
        "recommendation",
        "priority",
        "business model",
        "segment",
        "flag",
        "status",
        "bucket",
        "source"
    ]

    if any(term in header_lower for term in very_long_text_terms):
        return 44

    if any(term in header_lower for term in medium_text_terms):
        return 26

    if "company" in header_lower:
        return 24

    if "score" in header_lower or "rank" in header_lower or "count" in header_lower:
        return 14

    if "date" in header_lower:
        return 16

    return 18

# -----------------------------
# Format workbook
# -----------------------------

wb = load_workbook(workbook_path)

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True, color="000000")
thin_border = Border(
    bottom=Side(style="thin", color="B7B7B7")
)

for ws in wb.worksheets:
    if ws.max_row < 1 or ws.max_column < 1:
        continue

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        original_header = cell.value
        display_header = friendly_header(original_header)

        cell.value = display_header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[1].height = 36

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_value = ws.cell(row=1, column=col_idx).value

        ws.column_dimensions[col_letter].width = width_for_header(header_value)

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    for col_idx in range(1, ws.max_column + 1):
        header_value = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()

        if header_value in [
            "final priority rank",
            "final priority code",
            "decision priority sort",
            "data depth status rank",
            "coverage status rank"
        ]:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

# -----------------------------
# Save, sync, download
# -----------------------------

wb.save(workbook_path)
shutil.copy(workbook_path, formatted_drive_path)

print("PASS: Dashboard workbook formatted.")
print("Formatted local workbook:", workbook_path)
print("Formatted Drive workbook:", formatted_drive_path)

files.download(str(workbook_path))

# =============================================================================

# STEP 20 - Dashboard refresh runner

# =============================================================================

# =============================================================================
# STEP 20 - Dashboard refresh runner
# =============================================================================

# =============================================================================
# STEP 20 - Dashboard refresh runner
# =============================================================================

# =============================================================================
# STEP 20 - Dashboard refresh runner
# =============================================================================
# Purpose:
# - Run the dashboard refresh sequence from the checked-out GitHub workflow file.
# - This is orchestration only. It does not change scoring, priority, or export logic.

from pathlib import Path
import re
import sys
import time
import traceback

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. "
        "Run the GitHub pull/setup cell first."
    )

if not SRC_DIR.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {SRC_DIR}. "
        "Run the GitHub pull/setup cell first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

from health_tech_research_agent.priority import (
    apply_priority_fields,
    extract_priority_code,
    normalize_priority_level,
    priority_code,
    priority_rank,
    safe_text,
    is_blank_value,
)

print("PASS: Shared priority helper imported directly by Step 20 runner.")
print("apply_priority_fields available:", "apply_priority_fields" in globals())

workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")

DASHBOARD_REFRESH_STEPS = [
    "13",
    "14",
    "15",
    "16",
    "17",
    "18",
    "19",
    "19A",
]

def _step_marker_pattern(step_id):
    return re.compile(
        rf"(?m)^#\s*STEP\s+{re.escape(step_id)}\b.*$"
    )

def extract_step_code(step_id, source_text):
    """
    Extract a step block from colab_workflow.py.

    Handles duplicate markers for the same step by treating same-step markers
    as part of the same section and stopping only at the next different STEP.
    """
    marker_pattern = re.compile(
        r"(?m)^#\s*STEP\s+([0-9]+[A-Z]?)\b.*$"
    )

    markers = list(marker_pattern.finditer(source_text))

    if not markers:
        raise ValueError(f"STOP: No STEP markers found in {WORKFLOW_PATH}.")

    candidate_blocks = []

    for idx, marker in enumerate(markers):
        marker_step_id = marker.group(1)

        if marker_step_id != step_id:
            continue

        start = marker.start()

        # Stop at the next marker for a different step.
        end = len(source_text)
        for next_marker in markers[idx + 1:]:
            next_step_id = next_marker.group(1)
            if next_step_id != step_id:
                end = next_marker.start()
                break

        block = source_text[start:end].strip()

        candidate_blocks.append(block)

    if not candidate_blocks:
        raise ValueError(
            f"STOP: Could not find Step {step_id} marker in {WORKFLOW_PATH}."
        )

    # Pick the longest candidate so a duplicate header stub does not win.
    step_code = max(candidate_blocks, key=len)

    if not step_code:
        raise ValueError(f"STOP: Step {step_id} block is empty.")

    print(f"Extracted Step {step_id} block length: {len(step_code):,} chars")

    return step_code

runner_globals = globals()

def run_workflow_step(step_id):
    print("\n" + "=" * 80)
    print(f"RUNNING STEP {step_id}")
    print("=" * 80)

    step_code = extract_step_code(step_id, workflow_text)

    start_time = time.time()

    try:
        exec(
            compile(step_code, f"colab_workflow.py::STEP_{step_id}", "exec"),
            runner_globals,
            runner_globals,
        )
    except Exception as exc:
        print("\n" + "!" * 80)
        print(f"FAILED STEP {step_id}")
        print("!" * 80)
        traceback.print_exc()
        raise RuntimeError(f"Dashboard refresh runner stopped at Step {step_id}.") from exc

    elapsed = time.time() - start_time
    print(f"\nPASSED STEP {step_id} in {elapsed:,.1f} seconds")

print("Dashboard refresh runner ready.")
print("Workflow file:", WORKFLOW_PATH)
print("Steps:", " → ".join(DASHBOARD_REFRESH_STEPS))

for step_id in DASHBOARD_REFRESH_STEPS:
    run_workflow_step(step_id)

print("\n" + "=" * 80)
print("DASHBOARD REFRESH RUNNER COMPLETE")
print("=" * 80)

if "dashboard_workbook_path" in globals():
    print("dashboard_workbook_path =", dashboard_workbook_path)
else:
    print("WARNING: dashboard_workbook_path was not created.")

print("\nRun completed from workflow file:")
print(WORKFLOW_PATH)

# =============================================================================
# STEP 21 - Supervised LLM research batch runner
# =============================================================================
# Purpose:
# - Take a user-provided batch_name and batch_companies list.
# - Run LLM research through the supervised review gate.
# - Execute:
#   Step 7 -> Step 8 -> Step 8A -> Step 10 -> Step 10B -> Step 11 -> Step 11A
# - Build a human-review packet.
# - Stop before Step 12 so master is not updated without review.
#
# Required before running:
# - batch_name = "some_batch_name"
# - batch_companies = [{"company": "...", "research_query": "..."}, ...]
#
# Optional before running:
# - batch_wait_between_web_searches = 30
# - STEP_21_DRY_RUN = True

from pathlib import Path
import re
import sys
import time
import traceback
import pandas as pd

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"

STEP_21_DRY_RUN = bool(globals().get("STEP_21_DRY_RUN", False))

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. "
        "Run the GitHub setup / pull cell first."
    )

if not SRC_DIR.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {SRC_DIR}. "
        "Run the GitHub setup / pull cell first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

def step21_plain_display(obj):
    if isinstance(obj, pd.DataFrame):
        print(obj.to_string(index=False))
    else:
        print(obj)

display = step21_plain_display

if "batch_name" not in globals():
    raise NameError(
        "STOP: batch_name is not defined. "
        "Set batch_name before running Step 21."
    )

if "batch_companies" not in globals():
    raise NameError(
        "STOP: batch_companies is not defined. "
        "Set batch_companies before running Step 21."
    )

if not isinstance(batch_companies, list) or len(batch_companies) == 0:
    raise ValueError("STOP: batch_companies must be a non-empty list.")

for idx, item in enumerate(batch_companies, start=1):
    if not isinstance(item, dict):
        raise ValueError(f"STOP: batch_companies item {idx} is not a dict.")
    if not str(item.get("company", "")).strip():
        raise ValueError(f"STOP: batch_companies item {idx} is missing company.")
    if not str(item.get("research_query", "")).strip():
        raise ValueError(f"STOP: batch_companies item {idx} is missing research_query.")

required_runtime_items = [
    "client",
    "MODEL",
    "call_openai",
    "search_funding",
    "search_payer_signal",
    "search_outcomes",
    "search_commercial_scale",
    "run_company_fit_brief",
]

missing_runtime_items = [
    item for item in required_runtime_items
    if item not in globals()
]

if missing_runtime_items and not STEP_21_DRY_RUN:
    print("MISSING RUNTIME ITEMS")
    print("=" * 80)
    for item in missing_runtime_items:
        print("-", item)
    raise RuntimeError("Run the research setup loader before Step 21.")

BATCH_NAME = str(batch_name).strip()
companies = batch_companies

research_batches_folder = Path("research_batches")
research_batches_folder.mkdir(parents=True, exist_ok=True)

batch_checkpoint_path = research_batches_folder / f"{BATCH_NAME}_checkpoint.csv"
batch_raw_export_path = research_batches_folder / f"{BATCH_NAME}_raw.csv"
batch_summary_export_path = research_batches_folder / f"{BATCH_NAME}_summary.csv"
batch_review_packet_path = research_batches_folder / f"{BATCH_NAME}_human_review_packet.csv"

WAIT_BETWEEN_WEB_SEARCHES = int(
    globals().get("batch_wait_between_web_searches", 30)
)

print("STEP 21 - SUPERVISED LLM RESEARCH BATCH RUNNER")
print("=" * 80)
print("BATCH_NAME:", BATCH_NAME)
print("DRY RUN:", STEP_21_DRY_RUN)

if "MODEL" in globals():
    print("MODEL:", MODEL)
else:
    print("MODEL: not loaded")

print("WAIT_BETWEEN_WEB_SEARCHES:", WAIT_BETWEEN_WEB_SEARCHES)
print("Companies:")
for item in companies:
    print("-", item["company"])

print("batch_checkpoint_path:", batch_checkpoint_path)
print("batch_raw_export_path:", batch_raw_export_path)
print("batch_summary_export_path:", batch_summary_export_path)
print("batch_review_packet_path:", batch_review_packet_path)

workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")

marker_pattern = re.compile(
    r"(?m)^#\s*STEP\s+([0-9]+[A-Z]?)\b.*$"
)

markers = list(marker_pattern.finditer(workflow_text))

def step21_extract_step_code(step_id):
    candidate_blocks = []

    for idx, marker in enumerate(markers):
        marker_step_id = marker.group(1)

        if marker_step_id != step_id:
            continue

        start = marker.start()
        end = len(workflow_text)

        for next_marker in markers[idx + 1:]:
            next_step_id = next_marker.group(1)
            if next_step_id != step_id:
                end = next_marker.start()
                break

        candidate_blocks.append(workflow_text[start:end].strip())

    if not candidate_blocks:
        raise ValueError(f"STOP: Could not find Step {step_id} in {WORKFLOW_PATH}.")

    return max(candidate_blocks, key=len)

def step21_run_workflow_step(step_id):
    print("")
    print("=" * 80)
    print(f"RUNNING STEP {step_id}")
    print("=" * 80)

    step_code = step21_extract_step_code(step_id)
    print(f"Step {step_id} block length:", len(step_code))

    start_time = time.time()

    try:
        exec(
            compile(step_code, f"colab_workflow.py::STEP_{step_id}_FROM_STEP_21", "exec"),
            globals(),
            globals(),
        )
    except Exception:
        print(f"FAILED: Step {step_id} stopped.")
        traceback.print_exc()
        raise

    elapsed = time.time() - start_time
    print(f"PASS: Step {step_id} completed in {elapsed:.1f} seconds.")

def step21_validate_checkpoint():
    if not batch_checkpoint_path.exists():
        raise FileNotFoundError("STOP: Batch checkpoint was not created.")

    checkpoint_df = pd.read_csv(batch_checkpoint_path)

    required_cols = [
        "company",
        "date_researched",
        "funding_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "commercial_scale_finding",
        "fit_brief_json",
    ]

    print("")
    print("CHECKPOINT VALIDATION")
    print("=" * 80)
    print("checkpoint shape:", checkpoint_df.shape)

    all_complete = True

    for _, row in checkpoint_df.iterrows():
        company = row.get("company", "")
        missing = []

        for col in required_cols:
            value = row.get(col, "")
            if pd.isna(value) or str(value).strip() == "":
                missing.append(col)

        if missing:
            all_complete = False
            print(f"{company}: MISSING {', '.join(missing)}")
        else:
            print(f"{company}: COMPLETE")

    if not all_complete:
        raise RuntimeError("STOP: Checkpoint has incomplete rows. Stop before parsing/export.")

    return checkpoint_df

def step21_build_review_packet():
    if "summary_df" not in globals():
        raise NameError("STOP: summary_df was not created.")

    review_cols = [
        "company",
        "thesis_fit_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "katelynd_role_fit_score",
        "operator_timing_score",
        "operator_timing_score_raw",
        "operator_timing_score_cap",
        "public_company_hard_signal_present",
        "public_company_signal_basis",
        "public_company_language_without_hard_signal",
        "public_false_positive_language_present",
        "company_maturity_read",
        "likely_agency_level",
        "stage_timing_fit",
        "why_now_or_why_not",
        "timing_penalty_applied",
        "priority_level",
        "final_recommendation",
        "business_model_classification",
        "primary_market_segment_code",
        "primary_market_segment",
        "market_segment",
        "subsegment_tags",
        "product_model_tags",
        "distribution_model_tags",
        "data_input_tags",
        "taxonomy_assignment_method",
        "taxonomy_assignment_basis",
        "commercial_scale_signal",
        "institutional_distribution_signal",
        "outcomes_signal",
        "plausible_near_term_scale_path",
        "scale_engine_type",
        "strong_scale_engine_present",
        "calibration_flag",
        "final_takeaway",
    ]

    existing_review_cols = [
        col for col in review_cols
        if col in summary_df.columns
    ]

    review_packet_df = summary_df[existing_review_cols].copy()

    if "qa_df" in globals() and isinstance(qa_df, pd.DataFrame) and not qa_df.empty:
        qa_rollup = (
            qa_df
            .groupby("company", dropna=False)["issue"]
            .apply(lambda values: " | ".join([str(value) for value in values]))
            .reset_index()
            .rename(columns={"issue": "qa_flags"})
        )

        review_packet_df = review_packet_df.merge(
            qa_rollup,
            on="company",
            how="left",
        )
    else:
        review_packet_df["qa_flags"] = ""

    review_packet_df.to_csv(batch_review_packet_path, index=False)

    print("")
    print("BATCH REVIEW PACKET")
    print("=" * 80)
    print("review packet path:", batch_review_packet_path)
    print("review packet shape:", review_packet_df.shape)

    for _, row in review_packet_df.iterrows():
        print("")
        print("-" * 80)
        print("company:", row.get("company", ""))
        print("priority_level:", row.get("priority_level", ""))
        print("thesis_fit_score:", row.get("thesis_fit_score", ""))
        print("pmf_scale_score:", row.get("pmf_scale_score", ""))
        print("evidence_confidence_score:", row.get("evidence_confidence_score", ""))
        print("katelynd_role_fit_score:", row.get("katelynd_role_fit_score", ""))
        print("operator_timing_score:", row.get("operator_timing_score", ""))
        print("commercial_scale_signal:", row.get("commercial_scale_signal", ""))
        print("institutional_distribution_signal:", row.get("institutional_distribution_signal", ""))
        print("outcomes_signal:", row.get("outcomes_signal", ""))
        print("plausible_near_term_scale_path:", row.get("plausible_near_term_scale_path", ""))
        print("scale_engine_type:", row.get("scale_engine_type", ""))
        print("strong_scale_engine_present:", row.get("strong_scale_engine_present", ""))
        print("calibration_flag:", row.get("calibration_flag", ""))
        print("qa_flags:", row.get("qa_flags", ""))
        print("final_takeaway:", row.get("final_takeaway", ""))

    return review_packet_df

step21_steps = ["7", "8", "8A", "10", "10B", "11", "11A"]

print("")
print("Step 21 will run:")
print(" -> ".join(step21_steps))

print("")
print("Step extraction check:")
for step_id in step21_steps:
    step_code = step21_extract_step_code(step_id)
    print(f"- Step {step_id}: {len(step_code):,} chars")

if STEP_21_DRY_RUN:
    print("")
    print("=" * 80)
    print("STEP 21 DRY RUN COMPLETE")
    print("=" * 80)
    print("No research was run. No files were written except folder creation if needed.")
else:
    step21_run_workflow_step("7")
    checkpoint_df = step21_validate_checkpoint()

    for step_id in ["8", "8A", "10", "10B", "11", "11A"]:
        step21_run_workflow_step(step_id)

    review_packet_df = step21_build_review_packet()

    print("")
    print("=" * 80)
    print("STEP 21 SUPERVISED BATCH COMPLETE")
    print("=" * 80)
    print("Stopped before Step 12. Master was not updated.")
    print("Next decision: approve / downgrade / hold each company before master update.")

# =============================================================================
# STEP 22 - Approved batch to master and dashboard runner
# =============================================================================
# Purpose:
# - Take human review decisions for a completed Step 21 batch.
# - Apply those decisions to the batch summary export.
# - Exclude held/skipped companies from master update.
# - Run Step 12 to update master.
# - Run Step 20 to refresh the dashboard.
#
# Required before running:
# - batch_name = "completed_batch_name"
# - batch_review_decisions = {
#       "Company Name": {
#           "decision": "approve",  # approve/include OR hold/exclude/skip
#           "reviewed_priority_level": "P1: High-priority diligence",
#           "review_status": "Human reviewed",
#           "review_notes": "Why this decision was made.",
#           "priority_review_note": "Short dashboard-facing note."
#       }
#   }
#
# Optional before running:
# - STEP_22_DRY_RUN = True

from pathlib import Path
from datetime import datetime
import re
import sys
import time
import traceback
import pandas as pd

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"

STEP_22_DRY_RUN = bool(globals().get("STEP_22_DRY_RUN", False))

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. "
        "Run the GitHub setup / pull cell first."
    )

if not SRC_DIR.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {SRC_DIR}. "
        "Run the GitHub setup / pull cell first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

def step22_plain_display(obj):
    if isinstance(obj, pd.DataFrame):
        print(obj.to_string(index=False))
    else:
        print(obj)

display = step22_plain_display

if "batch_name" not in globals():
    raise NameError(
        "STOP: batch_name is not defined. "
        "Set batch_name before running Step 22."
    )

if "batch_review_decisions" not in globals():
    raise NameError(
        "STOP: batch_review_decisions is not defined. "
        "Set batch_review_decisions before running Step 22."
    )

if not isinstance(batch_review_decisions, dict) or not batch_review_decisions:
    raise ValueError("STOP: batch_review_decisions must be a non-empty dict.")

BATCH_NAME = str(batch_name).strip()

research_batches_folder = Path("research_batches")
research_batches_folder.mkdir(parents=True, exist_ok=True)

batch_checkpoint_path = research_batches_folder / f"{BATCH_NAME}_checkpoint.csv"
batch_raw_export_path = research_batches_folder / f"{BATCH_NAME}_raw.csv"
batch_summary_export_path = research_batches_folder / f"{BATCH_NAME}_summary.csv"
batch_review_packet_path = research_batches_folder / f"{BATCH_NAME}_human_review_packet.csv"

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

batch_summary_before_step22_path = (
    research_batches_folder / f"{BATCH_NAME}_summary_before_step22_{timestamp}.csv"
)

batch_summary_excluded_path = (
    research_batches_folder / f"{BATCH_NAME}_excluded_by_step22_{timestamp}.csv"
)

print("STEP 22 - APPROVED BATCH TO MASTER + DASHBOARD RUNNER")
print("=" * 80)
print("BATCH_NAME:", BATCH_NAME)
print("DRY RUN:", STEP_22_DRY_RUN)
print("batch_summary_export_path:", batch_summary_export_path)
print("batch_summary_before_step22_path:", batch_summary_before_step22_path)
print("batch_summary_excluded_path:", batch_summary_excluded_path)

if not batch_summary_export_path.exists():
    raise FileNotFoundError(
        f"STOP: Could not find batch summary export: {batch_summary_export_path}. "
        "Run Step 21 first, or make sure you are in the same Colab runtime."
    )

original_summary_df = pd.read_csv(batch_summary_export_path)
summary_df = original_summary_df.copy()

if "company" not in summary_df.columns:
    raise ValueError("STOP: batch summary is missing company column.")

print("")
print("Loaded batch summary.")
print("summary_df shape:", summary_df.shape)
print("companies:")
for company in summary_df["company"].tolist():
    print("-", company)

# -----------------------------
# Decision matching helpers
# -----------------------------

COMPANY_ALIASES = {
    "fay nutrition": "fay",
}

def step22_safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def step22_normalize_company_key(value):
    text = step22_safe_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return COMPANY_ALIASES.get(text, text)

decision_by_key = {}

for company_name, decision in batch_review_decisions.items():
    if not isinstance(decision, dict):
        raise ValueError(f"STOP: decision for {company_name} must be a dict.")

    key = step22_normalize_company_key(company_name)

    if key in decision_by_key:
        raise ValueError(f"STOP: duplicate decision key after normalization: {company_name}")

    decision_by_key[key] = decision

summary_df["_step22_company_key"] = summary_df["company"].apply(step22_normalize_company_key)

summary_keys = set(summary_df["_step22_company_key"].tolist())
decision_keys = set(decision_by_key.keys())

missing_decisions = sorted(summary_keys - decision_keys)
extra_decisions = sorted(decision_keys - summary_keys)

if missing_decisions:
    missing_companies = summary_df[
        summary_df["_step22_company_key"].isin(missing_decisions)
    ]["company"].tolist()

    raise ValueError(
        "STOP: Missing review decisions for batch companies: "
        + ", ".join(missing_companies)
    )

if extra_decisions:
    print("")
    print("WARNING: Review decisions were provided for companies not in the batch:")
    for key in extra_decisions:
        print("-", key)

# -----------------------------
# Apply decisions
# -----------------------------

for col in [
    "reviewed_priority_level",
    "review_status",
    "review_notes",
    "priority_review_note",
    "reviewed_priority_code",
]:
    if col not in summary_df.columns:
        summary_df[col] = ""

approved_rows = []
excluded_rows = []

for _, row in summary_df.iterrows():
    company = row["company"]
    key = row["_step22_company_key"]
    decision = decision_by_key[key]

    action = step22_safe_text(decision.get("decision", "approve")).lower()

    if action in ["hold", "exclude", "skip", "do not add", "do_not_add"]:
        excluded_row = row.copy()
        excluded_row["step22_decision"] = action
        excluded_row["step22_exclusion_reason"] = step22_safe_text(
            decision.get("reason", decision.get("review_notes", "Held from master update."))
        )
        excluded_rows.append(excluded_row)
        continue

    if action not in ["approve", "include", "update", "add"]:
        raise ValueError(
            f"STOP: unsupported decision '{action}' for {company}. "
            "Use approve/include/update/add or hold/exclude/skip."
        )

    reviewed_priority_level = step22_safe_text(decision.get("reviewed_priority_level", ""))

    if reviewed_priority_level == "":
        raise ValueError(f"STOP: approved company {company} is missing reviewed_priority_level.")

    review_status = step22_safe_text(decision.get("review_status", "Human reviewed"))

    review_notes = step22_safe_text(decision.get("review_notes", ""))
    priority_review_note = step22_safe_text(decision.get("priority_review_note", ""))

    if review_notes == "":
        raise ValueError(f"STOP: approved company {company} is missing review_notes.")

    if priority_review_note == "":
        priority_review_note = review_notes

    priority_code_match = re.search(r"\bP[0-4]\b", reviewed_priority_level)
    reviewed_priority_code = priority_code_match.group(0) if priority_code_match else ""

    updated_row = row.copy()
    updated_row["reviewed_priority_level"] = reviewed_priority_level
    updated_row["review_status"] = review_status
    updated_row["review_notes"] = review_notes
    updated_row["priority_review_note"] = priority_review_note
    updated_row["reviewed_priority_code"] = reviewed_priority_code
    updated_row["step22_decision"] = action

    approved_rows.append(updated_row)

approved_df = pd.DataFrame(approved_rows)
excluded_df = pd.DataFrame(excluded_rows)

if not approved_df.empty and "_step22_company_key" in approved_df.columns:
    approved_df = approved_df.drop(columns=["_step22_company_key"])

if not excluded_df.empty and "_step22_company_key" in excluded_df.columns:
    excluded_df = excluded_df.drop(columns=["_step22_company_key"])

print("")
print("STEP 22 DECISION SUMMARY")
print("=" * 80)
print("Approved/include rows:", len(approved_df))
print("Held/excluded rows:", len(excluded_df))

if not approved_df.empty:
    print("")
    print("Approved companies:")
    print(
        approved_df[
            [
                col for col in [
                    "company",
                    "priority_level",
                    "reviewed_priority_level",
                    "review_status",
                    "priority_review_note",
                ]
                if col in approved_df.columns
            ]
        ].to_string(index=False)
    )

if not excluded_df.empty:
    print("")
    print("Excluded companies:")
    print(
        excluded_df[
            [
                col for col in [
                    "company",
                    "priority_level",
                    "step22_decision",
                    "step22_exclusion_reason",
                ]
                if col in excluded_df.columns
            ]
        ].to_string(index=False)
    )

if approved_df.empty:
    raise ValueError("STOP: No approved companies remain for master update.")

# -----------------------------
# Step extraction runner
# -----------------------------

workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")

marker_pattern = re.compile(
    r"(?m)^#\s*STEP\s+([0-9]+[A-Z]?)\b.*$"
)

markers = list(marker_pattern.finditer(workflow_text))

def step22_extract_step_code(step_id):
    candidate_blocks = []

    for idx, marker in enumerate(markers):
        marker_step_id = marker.group(1)

        if marker_step_id != step_id:
            continue

        start = marker.start()
        end = len(workflow_text)

        for next_marker in markers[idx + 1:]:
            next_step_id = next_marker.group(1)
            if next_step_id != step_id:
                end = next_marker.start()
                break

        candidate_blocks.append(workflow_text[start:end].strip())

    if not candidate_blocks:
        raise ValueError(f"STOP: Could not find Step {step_id} in {WORKFLOW_PATH}.")

    return max(candidate_blocks, key=len)

def step22_run_workflow_step(step_id):
    print("")
    print("=" * 80)
    print(f"RUNNING STEP {step_id}")
    print("=" * 80)

    step_code = step22_extract_step_code(step_id)
    print(f"Step {step_id} block length:", len(step_code))

    start_time = time.time()

    try:
        exec(
            compile(step_code, f"colab_workflow.py::STEP_{step_id}_FROM_STEP_22", "exec"),
            globals(),
            globals(),
        )
    except Exception:
        print(f"FAILED: Step {step_id} stopped.")
        traceback.print_exc()
        raise

    elapsed = time.time() - start_time
    print(f"PASS: Step {step_id} completed in {elapsed:.1f} seconds.")

print("")
print("Step extraction check:")
for step_id in ["12", "20"]:
    step_code = step22_extract_step_code(step_id)
    print(f"- Step {step_id}: {len(step_code):,} chars")

if STEP_22_DRY_RUN:
    print("")
    print("=" * 80)
    print("STEP 22 DRY RUN COMPLETE")
    print("=" * 80)
    print("No files were overwritten. Master was not updated. Dashboard was not refreshed.")
else:
    original_summary_df.to_csv(batch_summary_before_step22_path, index=False)

    if not excluded_df.empty:
        excluded_df.to_csv(batch_summary_excluded_path, index=False)

    summary_df = approved_df.copy()
    summary_df.to_csv(batch_summary_export_path, index=False)

    print("")
    print("Saved approved batch summary for Step 12:")
    print(batch_summary_export_path)

    if not excluded_df.empty:
        print("Saved excluded rows:")
        print(batch_summary_excluded_path)

    step22_run_workflow_step("12")

    # # STEP 22 PATCH - reapply human review decisions after Step 12

    # Step 12 can overwrite review_status/review_notes for newly added companies.

    # Reapply the Step 22/24 human decisions to the active master before dashboard refresh.

    from pathlib import Path as _Step22Path


    _step22_active_master_path = _Step22Path(

        "/content/drive/MyDrive/Job Search/Health Tech Research/health_tech_market_research_summary_MASTER.csv"

    )


    if not _step22_active_master_path.exists():

        raise FileNotFoundError(f"STOP: active master not found: {_step22_active_master_path}")


    _step22_master_df = pd.read_csv(_step22_active_master_path)


    for _col in [

        "reviewed_priority_level",

        "review_status",

        "review_notes",

        "priority_review_note",

    ]:

        if _col not in _step22_master_df.columns:

            _step22_master_df[_col] = ""


    _step22_reapplied_count = 0


    for _, _approved_row in approved_df.iterrows():

        _company_raw = str(_approved_row.get("company", "")).strip()

        if "canonical_company_name" in globals():

            _company = canonical_company_name(_company_raw)

        else:

            _company = _company_raw


        _mask = _step22_master_df["company"].astype(str).str.strip().eq(str(_company).strip())


        if _mask.sum() != 1:

            raise RuntimeError(

                f"STOP: expected exactly one master row for {_company} after Step 12, found {_mask.sum()}."

            )


        _reviewed_priority = str(_approved_row.get("reviewed_priority_level", "")).strip()

        _review_status = str(_approved_row.get("review_status", "Human reviewed")).strip() or "Human reviewed"

        _review_notes = str(_approved_row.get("review_notes", "")).strip()

        _priority_review_note = str(_approved_row.get("priority_review_note", _review_notes)).strip() or _review_notes


        if _reviewed_priority:

            _step22_master_df.loc[_mask, "reviewed_priority_level"] = _reviewed_priority


        _step22_master_df.loc[_mask, "review_status"] = _review_status


        if _review_notes:

            _step22_master_df.loc[_mask, "review_notes"] = _review_notes


        if _priority_review_note:

            _step22_master_df.loc[_mask, "priority_review_note"] = _priority_review_note


        _step22_reapplied_count += 1


    _step22_master_df.to_csv(_step22_active_master_path, index=False)


    print("")

    print("STEP 22 PATCH: re-applied human review decisions to active master.")

    print("Rows updated:", _step22_reapplied_count)

    print("Active master re-saved to:", _step22_active_master_path)
    step22_run_workflow_step("20")

    print("")
    print("=" * 80)
    print("STEP 22 APPROVED BATCH COMPLETE")
    print("=" * 80)
    print("Master was updated and dashboard was refreshed.")

    if "dashboard_workbook_path" in globals():
        print("dashboard_workbook_path =", dashboard_workbook_path)

# =============================================================================
# STEP 23 - Google Sheet research queue launcher
# =============================================================================
# Purpose:
# - Read READY companies from the Google Sheet Research Queue tab.
# - Build batch_name and batch_companies automatically.
# - Run Step 21.
# - Write readable outputs to Review Packet and Evidence Detail.
# - Stop before master update.
#
# Required Google Sheet tabs:
# - Research Queue
# - Review Packet
# - Evidence Detail
#
# Optional before running:
# - BATCH_CONTROL_SHEET_URL = "..."
# - STEP23_SELECTED_BATCH_NAME = "..."  # optional intentional override
# - STEP_23_DRY_RUN = True

from pathlib import Path
from datetime import datetime
import json
import os
import re
import subprocess
import sys
import time
import traceback
import pandas as pd

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"

DEFAULT_BATCH_CONTROL_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1psVpW5SddVlgbnwkToXjbMvL0NWuyM3feyY5fwKCjjY/edit"
)

BATCH_CONTROL_SHEET_URL = globals().get(
    "BATCH_CONTROL_SHEET_URL",
    DEFAULT_BATCH_CONTROL_SHEET_URL,
)

STEP_23_DRY_RUN = bool(globals().get("STEP_23_DRY_RUN", False))

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. "
        "Run the GitHub setup / pull cell first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

def step23_plain_display(obj):
    if isinstance(obj, pd.DataFrame):
        print(obj.to_string(index=False))
    else:
        print(obj)

display = step23_plain_display

def step23_install_or_import_gspread():
    try:
        import gspread
        return gspread
    except ImportError:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "-q", "gspread", "google-auth"],
            check=True,
        )
        import gspread
        return gspread

def step23_authenticate_sheet():
    from google.colab import auth
    import google.auth

    gspread = step23_install_or_import_gspread()

    auth.authenticate_user()
    creds, _ = google.auth.default()

    return gspread.authorize(creds)

def step23_safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def step23_normalize_key(value):
    text = step23_safe_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def step23_dataframe_from_sheet(worksheet):
    values = worksheet.get_all_values()

    if not values:
        return pd.DataFrame()

    headers = [step23_safe_text(header) for header in values[0]]

    rows = []
    for raw_row in values[1:]:
        padded = raw_row + [""] * (len(headers) - len(raw_row))
        padded = raw_row + [""] * (len(headers) - len(raw_row))
        rows.append(padded[:len(headers)])

    return pd.DataFrame(rows, columns=headers)

def step23_autogenerated_query(company, notes="", priority_context="", custom_query=""):
    if step23_safe_text(custom_query):
        return step23_safe_text(custom_query)

    company = step23_safe_text(company)
    notes = step23_safe_text(notes)
    priority_context = step23_safe_text(priority_context)

    query_parts = [
        company,
        "health tech funding revenue growth payer employer provider health plan partnerships",
        "outcomes retention engagement commercial scale business model implementation",
    ]

    if notes:
        query_parts.append(notes)

    if priority_context:
        query_parts.append(priority_context)

    return " ".join(query_parts)


def step23_load_research_runtime_if_needed():
    required_runtime_items = [
        "client",
        "MODEL",
        "call_openai",
        "search_funding",
        "search_payer_signal",
        "search_outcomes",
        "search_commercial_scale",
        "run_company_fit_brief",
    ]

    missing = [name for name in required_runtime_items if name not in globals()]

    # Default to reloading the research runtime from the current workflow file.
    # Colab keeps old globals alive across cell edits, which can otherwise run stale prompts/parsers.
    force_reload = bool(globals().get("STEP23_FORCE_RESEARCH_RUNTIME_RELOAD", True))

    if not missing and not force_reload:
        print("Research runtime already loaded and forced reload disabled.")
        return

    if STEP_23_DRY_RUN:
        print("Dry run: research runtime not loaded, but dry run does not need it.")
        print("Missing runtime items:", missing)
        return

    if force_reload:
        print("Forcing research runtime reload from current colab_workflow.py.")
    else:
        print("Research runtime missing. Loading setup/function block from colab_workflow.py.")

    if missing:
        print("Missing runtime items:", missing)

    workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")
    step_6_match = re.search(r"(?m)^#\s*STEP\s+6\b", workflow_text)

    if not step_6_match:
        raise RuntimeError("Could not find Step 6 marker for setup loader.")

    setup_code = workflow_text[:step_6_match.start()]

    # colab_workflow.py contains notebook shell lines like !pip install, which cannot be compiled
    # through exec(). Install dependencies with subprocess, then strip shell/magic lines.
    import sys as _sys
    import subprocess as _subprocess

    _subprocess.run(
        [_sys.executable, "-m", "pip", "install", "-q", "openai", "pandas"],
        check=False,
    )

    setup_code = "\n".join(
        line for line in setup_code.splitlines()
        if not line.lstrip().startswith("!") and not line.lstrip().startswith("%")
    )

    exec(
        compile(setup_code, "colab_workflow.py::SETUP_LOADER_FROM_STEP_23", "exec"),
        globals(),
        globals(),
    )

    missing_after = [name for name in required_runtime_items if name not in globals()]

    if missing_after:
        raise RuntimeError(
            "Research runtime is still incomplete after setup load: "
            + ", ".join(missing_after)
        )

    print("PASS: research runtime loaded from current workflow.")
    print("MODEL:", MODEL)

def step23_extract_step_code(step_id):
    workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")

    marker_pattern = re.compile(
        r"(?m)^#\s*STEP\s+([0-9]+[A-Z]?)\b.*$"
    )

    markers = list(marker_pattern.finditer(workflow_text))
    candidate_blocks = []

    for idx, marker in enumerate(markers):
        marker_step_id = marker.group(1)

        if marker_step_id != step_id:
            continue

        start = marker.start()
        end = len(workflow_text)

        for next_marker in markers[idx + 1:]:
            next_step_id = next_marker.group(1)
            if next_step_id != step_id:
                end = next_marker.start()
                break

        candidate_blocks.append(workflow_text[start:end].strip())

    if not candidate_blocks:
        raise ValueError(f"STOP: Could not find Step {step_id} in {WORKFLOW_PATH}.")

    return max(candidate_blocks, key=len)

def step23_run_workflow_step(step_id):
    print("")
    print("=" * 80)
    print(f"RUNNING STEP {step_id}")
    print("=" * 80)

    step_code = step23_extract_step_code(step_id)
    print(f"Step {step_id} block length:", len(step_code))

    start_time = time.time()

    try:
        exec(
            compile(step_code, f"colab_workflow.py::STEP_{step_id}_FROM_STEP_23", "exec"),
            globals(),
            globals(),
        )
    except Exception:
        print(f"FAILED: Step {step_id} stopped.")
        traceback.print_exc()
        raise

    elapsed = time.time() - start_time
    print(f"PASS: Step {step_id} completed in {elapsed:.1f} seconds.")

def step23_parse_json(value):
    text = step23_safe_text(value)

    if text == "":
        return {}

    try:
        return json.loads(text)
    except Exception:
        return {}

def step23_text_from_any(value):
    if value is None:
        return ""

    if isinstance(value, float) and pd.isna(value):
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, list):
        pieces = []
        for item in value:
            item_text = step23_text_from_any(item)
            if item_text:
                pieces.append(item_text)
        return "\n".join(f"- {piece}" for piece in pieces)

    if isinstance(value, dict):
        pieces = []
        for key, item in value.items():
            item_text = step23_text_from_any(item)
            if item_text:
                pieces.append(f"{key}: {item_text}")
        return "\n".join(pieces)

    return str(value).strip()

def step23_first_available(source, keys):
    if not isinstance(source, dict):
        return ""

    for key in keys:
        if key in source:
            text = step23_text_from_any(source.get(key))
            if text:
                return text

    return ""

def step23_get_row_value(row, candidates):
    for col in candidates:
        if col in row.index:
            value = step23_safe_text(row.get(col, ""))
            if value:
                return value
    return ""

def step23_build_qa_rollup():
    if "qa_df" not in globals():
        return pd.DataFrame(columns=["company", "qa_flags"])

    if not isinstance(qa_df, pd.DataFrame) or qa_df.empty:
        return pd.DataFrame(columns=["company", "qa_flags"])

    issue_col = "issue" if "issue" in qa_df.columns else None

    if issue_col is None:
        return pd.DataFrame(columns=["company", "qa_flags"])

    return (
        qa_df
        .groupby("company", dropna=False)[issue_col]
        .apply(lambda values: " | ".join([str(value) for value in values if str(value).strip()]))
        .reset_index()
        .rename(columns={issue_col: "qa_flags"})
    )

def step23_build_readable_outputs(batch_name):
    if "summary_df" not in globals():
        raise NameError("STOP: summary_df was not created by Step 21.")

    local_summary_df = summary_df.copy()

    checkpoint_path = Path("research_batches") / f"{batch_name}_checkpoint.csv"

    if checkpoint_path.exists():
        checkpoint_df = pd.read_csv(checkpoint_path)
    else:
        checkpoint_df = pd.DataFrame(columns=["company"])

    if "company" in checkpoint_df.columns:
        checkpoint_lookup = {
            step23_normalize_key(row.get("company", "")): row
            for _, row in checkpoint_df.iterrows()
        }
    else:
        checkpoint_lookup = {}

    qa_rollup = step23_build_qa_rollup()

    if not qa_rollup.empty:
        local_summary_df = local_summary_df.merge(
            qa_rollup,
            on="company",
            how="left",
        )
    else:
        local_summary_df["qa_flags"] = ""

    required_taxonomy_review_cols = [
        "primary_market_segment_code",
        "subsegment_tags",
        "product_model_tags",
        "distribution_model_tags",
        "data_input_tags",
        "taxonomy_assignment_method",
    ]

    missing_taxonomy_review_cols = [
        col for col in required_taxonomy_review_cols
        if col not in local_summary_df.columns
    ]

    if missing_taxonomy_review_cols:
        raise RuntimeError(
            "STOP: Step 23 review output is missing taxonomy columns: "
            + ", ".join(missing_taxonomy_review_cols)
        )

    blank_taxonomy_rows = local_summary_df[
        local_summary_df["primary_market_segment_code"].astype(str).str.strip().eq("")
    ]

    if not blank_taxonomy_rows.empty:
        raise RuntimeError(
            "STOP: Step 23 review output has blank primary_market_segment_code for: "
            + ", ".join(blank_taxonomy_rows["company"].astype(str).tolist())
            + ". This usually means the research runtime was stale or the fit brief omitted taxonomy_classification."
        )


    review_packet_rows = []
    evidence_rows = []

    for _, row in local_summary_df.iterrows():
        company = step23_safe_text(row.get("company", ""))
        company_key = step23_normalize_key(company)

        checkpoint_row = checkpoint_lookup.get(company_key, pd.Series(dtype=object))
        brief_json = step23_parse_json(checkpoint_row.get("fit_brief_json", ""))

        model_priority = step23_get_row_value(row, ["priority_level", "final_priority_level"])
        recommended_decision = "review"

        if model_priority.startswith("P0") or model_priority.startswith("P1"):
            recommended_decision = "consider_approve"
        elif model_priority.startswith("P2"):
            recommended_decision = "deeper_diligence"
        elif model_priority.startswith("P3"):
            recommended_decision = "watchlist"
        elif model_priority.startswith("P4"):
            recommended_decision = "likely_reject"

        final_takeaway = step23_get_row_value(row, ["final_takeaway"])
        if not final_takeaway:
            final_takeaway = step23_first_available(
                brief_json,
                ["final_takeaway", "summary", "recommendation_summary"]
            )

        why_fits = step23_first_available(
            brief_json,
            [
                "why_this_fits_katelynd",
                "katelynd_role_fit",
                "role_fit_assessment",
                "operator_fit",
            ],
        )

        why_not = step23_first_available(
            brief_json,
            [
                "why_this_might_not_fit",
                "risks",
                "risk_flags",
                "weaknesses",
                "concerns",
            ],
        )

        verified_facts = step23_first_available(
            brief_json,
            ["verified_facts", "key_verified_facts", "verified_claims"]
        )

        weak_claims = step23_first_available(
            brief_json,
            [
                "key_weak_or_unverified_claims",
                "unverified_or_weak_claims",
                "weak_claims",
                "unknowns",
                "evidence_gaps",
            ],
        )

        sources_summary = step23_first_available(
            brief_json,
            ["sources", "source_summary", "source_urls", "source_notes"]
        )

        suggested_next_diligence = step23_first_available(
            brief_json,
            ["suggested_next_diligence", "next_diligence", "diligence_questions"]
        )

        review_packet_rows.append({
            "batch_name": batch_name,
            "company": company,
            "model_priority": model_priority,
            "recommended_decision": recommended_decision,
            "final_takeaway": final_takeaway,
            "why_this_fits_katelynd": why_fits,
            "why_this_might_not_fit": why_not,
            "commercial_scale_signal": step23_get_row_value(row, ["commercial_scale_signal"]),
            "institutional_distribution_signal": step23_get_row_value(row, ["institutional_distribution_signal"]),
            "outcomes_signal": step23_get_row_value(row, ["outcomes_signal"]),
            "evidence_confidence_score": step23_get_row_value(row, ["evidence_confidence_score"]),
            "thesis_fit_score": step23_get_row_value(row, ["thesis_fit_score"]),
            "pmf_scale_score": step23_get_row_value(row, ["pmf_scale_score"]),
            "role_fit_score": step23_get_row_value(row, ["katelynd_role_fit_score", "role_fit_score"]),
            "operator_timing_score": step23_get_row_value(row, ["operator_timing_score"]),
            "operator_timing_score_raw": step23_get_row_value(row, ["operator_timing_score_raw"]),
            "operator_timing_score_cap": step23_get_row_value(row, ["operator_timing_score_cap"]),
            "company_maturity_read": step23_get_row_value(row, ["company_maturity_read"]),
            "likely_agency_level": step23_get_row_value(row, ["likely_agency_level"]),
            "stage_timing_fit": step23_get_row_value(row, ["stage_timing_fit"]),
            "why_now_or_why_not": step23_get_row_value(row, ["why_now_or_why_not"]),
            "timing_penalty_applied": step23_get_row_value(row, ["timing_penalty_applied"]),
            "business_model_classification": step23_get_row_value(row, ["business_model_classification"]),
            "primary_market_segment_code": step23_get_row_value(row, ["primary_market_segment_code"]),
            "primary_market_segment": step23_get_row_value(row, ["primary_market_segment"]),
            "market_segment": step23_get_row_value(row, ["market_segment"]),
            "subsegment_tags": step23_get_row_value(row, ["subsegment_tags"]),
            "product_model_tags": step23_get_row_value(row, ["product_model_tags"]),
            "distribution_model_tags": step23_get_row_value(row, ["distribution_model_tags"]),
            "data_input_tags": step23_get_row_value(row, ["data_input_tags"]),
            "taxonomy_assignment_method": step23_get_row_value(row, ["taxonomy_assignment_method"]),
            "taxonomy_assignment_basis": step23_get_row_value(row, ["taxonomy_assignment_basis"]),
            "key_verified_facts": verified_facts,
            "key_weak_or_unverified_claims": weak_claims,
            "sources_summary": sources_summary,
            "suggested_next_diligence": suggested_next_diligence,
            "qa_flags": step23_safe_text(row.get("qa_flags", "")),
            "review_decision": "",
            "reviewed_priority": model_priority,
            "review_notes": "",
            "ready_for_master_update": "NO",
            "reviewed_at": "",
            "step_22_status": "",
        })

        for evidence_type, claim_text in [
            ("funding_finding", checkpoint_row.get("funding_finding", "")),
            ("payer_institutional_finding", checkpoint_row.get("payer_institutional_finding", "")),
            ("outcomes_finding", checkpoint_row.get("outcomes_finding", "")),
            ("commercial_scale_finding", checkpoint_row.get("commercial_scale_finding", "")),
            ("verified_facts", verified_facts),
            ("weak_or_unverified_claims", weak_claims),
            ("sources_summary", sources_summary),
        ]:
            claim_text = step23_text_from_any(claim_text)

            if claim_text:
                evidence_rows.append({
                    "batch_name": batch_name,
                    "company": company,
                    "evidence_type": evidence_type,
                    "claim": claim_text,
                    "source": "",
                    "source_quality": "",
                    "confidence": "",
                    "notes": "",
                })

    return pd.DataFrame(review_packet_rows), pd.DataFrame(evidence_rows)

def step23_write_dataframe_to_worksheet(worksheet, headers, df):
    worksheet.clear()

    # Make sure the sheet is wide enough for newly added review columns.
    # This prevents future schema additions from silently failing or truncating.
    try:
        target_rows = max(len(df) + 1, 2)
        target_cols = max(len(headers), getattr(worksheet, "col_count", len(headers)))
        worksheet.resize(rows=target_rows, cols=target_cols)
    except Exception as resize_error:
        print(f"WARNING: Could not resize worksheet before write: {resize_error}")

    worksheet.update("A1", [headers], value_input_option="USER_ENTERED")

    if df.empty:
        return

    rows = []

    for _, row in df.iterrows():
        rows.append([step23_safe_text(row.get(header, "")) for header in headers])

    worksheet.update("A2", rows, value_input_option="USER_ENTERED")

def step23_write_outputs_to_sheet(spreadsheet, batch_name):
    review_headers = [
        "batch_name",
        "company",
        "model_priority",
        "recommended_decision",
        "final_takeaway",
        "why_this_fits_katelynd",
        "why_this_might_not_fit",
        "commercial_scale_signal",
        "institutional_distribution_signal",
        "outcomes_signal",
        "evidence_confidence_score",
        "thesis_fit_score",
        "pmf_scale_score",
        "role_fit_score",
        "operator_timing_score",
        "operator_timing_score_raw",
        "operator_timing_score_cap",
        "company_maturity_read",
        "likely_agency_level",
        "stage_timing_fit",
        "why_now_or_why_not",
        "timing_penalty_applied",
        "business_model_classification",
        "primary_market_segment_code",
        "primary_market_segment",
        "market_segment",
        "subsegment_tags",
        "product_model_tags",
        "distribution_model_tags",
        "data_input_tags",
        "taxonomy_assignment_method",
        "taxonomy_assignment_basis",
        "key_verified_facts",
        "key_weak_or_unverified_claims",
        "sources_summary",
        "suggested_next_diligence",
        "qa_flags",
        "review_decision",
        "reviewed_priority",
        "review_notes",
        "ready_for_master_update",
        "reviewed_at",
        "step_22_status",
    ]

    evidence_headers = [
        "batch_name",
        "company",
        "evidence_type",
        "claim",
        "source",
        "source_quality",
        "confidence",
        "notes",
    ]

    review_packet_df, evidence_detail_df = step23_build_readable_outputs(batch_name)

    review_ws = spreadsheet.worksheet("Review Packet")
    evidence_ws = spreadsheet.worksheet("Evidence Detail")

    step23_write_dataframe_to_worksheet(review_ws, review_headers, review_packet_df)
    step23_write_dataframe_to_worksheet(evidence_ws, evidence_headers, evidence_detail_df)

    print("")
    print("WROTE READABLE REVIEW OUTPUTS TO GOOGLE SHEET")
    print("=" * 80)
    print("Review Packet rows:", len(review_packet_df))
    print("Evidence Detail rows:", len(evidence_detail_df))

    return review_packet_df, evidence_detail_df


def step23_column_letter(zero_based_index):
    """Convert zero-based column index to Google Sheets column letters."""
    number = int(zero_based_index) + 1
    letters = ""

    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters

    return letters

def step23_mark_queue_done(queue_ws, queue_df, selected_batch_name_value):
    values = queue_ws.get_all_values()

    if not values:
        return

    headers = values[0]
    status_idx = headers.index("status") if "status" in headers else None
    batch_idx = headers.index("batch_name") if "batch_name" in headers else None

    if status_idx is None or batch_idx is None:
        return

    updates = []

    for row_number, row_values in enumerate(values[1:], start=2):
        padded = row_values + [""] * (len(headers) - len(row_values))
        row_batch = step23_safe_text(padded[batch_idx])

        if row_batch == selected_batch_name_value:
            cell_a1 = f"{step23_column_letter(status_idx)}{row_number}"
            updates.append({
                "range": cell_a1,
                "values": [["DONE"]],
            })

    if updates:
        queue_ws.batch_update(updates)

print("STEP 23 - GOOGLE SHEET RESEARCH QUEUE LAUNCHER")
print("=" * 80)
print("BATCH_CONTROL_SHEET_URL:", BATCH_CONTROL_SHEET_URL)
print("DRY RUN:", STEP_23_DRY_RUN)

client_for_sheet = step23_authenticate_sheet()
spreadsheet = client_for_sheet.open_by_url(BATCH_CONTROL_SHEET_URL)

queue_ws = spreadsheet.worksheet("Research Queue")
queue_df = step23_dataframe_from_sheet(queue_ws)

if queue_df.empty:
    raise ValueError("STOP: Research Queue is empty.")

required_queue_cols = ["batch_name", "company", "status"]
missing_queue_cols = [col for col in required_queue_cols if col not in queue_df.columns]

if missing_queue_cols:
    raise ValueError("STOP: Research Queue missing columns: " + ", ".join(missing_queue_cols))

ready_df = queue_df[
    queue_df["status"].astype(str).str.strip().str.upper().eq("READY")
].copy()

if ready_df.empty:
    raise ValueError("STOP: No READY rows found in Research Queue.")

step23_selected_batch_override = step23_safe_text(globals().get("STEP23_SELECTED_BATCH_NAME", ""))

if step23_selected_batch_override:
    selected_batch_name_value = step23_selected_batch_override
    ready_df = ready_df[
        ready_df["batch_name"].astype(str).str.strip().eq(selected_batch_name_value)
    ].copy()

    if ready_df.empty:
        raise ValueError(
            f"STOP: No READY rows found for STEP23_SELECTED_BATCH_NAME={selected_batch_name_value}"
        )
else:
    unique_ready_batches = sorted(
        [batch for batch in ready_df["batch_name"].astype(str).str.strip().unique() if batch]
    )

    if len(unique_ready_batches) != 1:
        raise ValueError(
            "STOP: Multiple READY batch_name values found. "
            "Set STEP23_SELECTED_BATCH_NAME before running Step 23 only if multiple READY batches exist. "
            f"READY batches: {unique_ready_batches}"
        )

    selected_batch_name_value = unique_ready_batches[0]

batch_name = selected_batch_name_value

batch_companies = []

for _, row in ready_df.iterrows():
    company = step23_safe_text(row.get("company", ""))
    notes = step23_safe_text(row.get("notes", row.get("notes_optional", "")))
    priority_context = step23_safe_text(row.get("priority_context", ""))
    custom_query = step23_safe_text(
        row.get("custom_research_query", row.get("custom_research_query_optional", ""))
    )

    if company == "":
        continue

    batch_companies.append({
        "company": company,
        "research_query": step23_autogenerated_query(
            company=company,
            notes=notes,
            priority_context=priority_context,
            custom_query=custom_query,
        ),
    })

if not batch_companies:
    raise ValueError("STOP: READY rows did not contain any company names.")

batch_wait_between_web_searches = int(globals().get("batch_wait_between_web_searches", 30))
STEP_21_DRY_RUN = STEP_23_DRY_RUN

print("")
print("BATCH SELECTED")
print("=" * 80)
print("batch_name:", batch_name)
print("batch_wait_between_web_searches:", batch_wait_between_web_searches)
print("companies:")
for item in batch_companies:
    print("-", item["company"])

print("")
print("Step 21 extraction check:")
step21_code = step23_extract_step_code("21")
print("Step 21 block length:", len(step21_code))

if STEP_23_DRY_RUN:
    print("")
    print("=" * 80)
    print("STEP 23 DRY RUN COMPLETE")
    print("=" * 80)
    print("No research was run. No Google Sheet outputs were overwritten.")
else:
    step23_load_research_runtime_if_needed()
    step23_run_workflow_step("21")

    review_packet_df, evidence_detail_df = step23_write_outputs_to_sheet(
        spreadsheet=spreadsheet,
        batch_name=batch_name,
    )

    step23_mark_queue_done(
        queue_ws=queue_ws,
        queue_df=queue_df,
        selected_batch_name_value=selected_batch_name_value,
    )

    print("")
    print("=" * 80)
    print("STEP 23 COMPLETE")
    print("=" * 80)
    print("Research is complete.")
    print("Review the results in Google Sheets:")
    print(BATCH_CONTROL_SHEET_URL)
    print("")
    print("Next: fill review_decision, reviewed_priority, review_notes, and ready_for_master_update in Review Packet.")

# =============================================================================
# STEP 24 - Google Sheet review approval launcher
# =============================================================================
# Purpose:
# - Read human review decisions from the Google Sheet Review Packet tab.
# - Convert those decisions into batch_review_decisions for Step 22.
# - Run Step 22 to update master and refresh the dashboard.
# - Update Review Packet step_22_status after successful completion.
#
# Required Google Sheet tab:
# - Review Packet
#
# Required before running, usually:
# - selected_batch_name = "completed_batch_name"
#
# Optional before running:
# - BATCH_CONTROL_SHEET_URL = "..."
# - STEP_24_DRY_RUN = True

from pathlib import Path
from datetime import datetime
import os
import re
import subprocess
import sys
import time
import traceback
import pandas as pd

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"

DEFAULT_BATCH_CONTROL_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1psVpW5SddVlgbnwkToXjbMvL0NWuyM3feyY5fwKCjjY/edit"
)

BATCH_CONTROL_SHEET_URL = globals().get(
    "BATCH_CONTROL_SHEET_URL",
    DEFAULT_BATCH_CONTROL_SHEET_URL,
)

STEP_24_DRY_RUN = bool(globals().get("STEP_24_DRY_RUN", False))

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. "
        "Run the GitHub setup / pull cell first."
    )

src_path = str(SRC_DIR)
if src_path not in sys.path:
    sys.path.insert(0, src_path)

def step24_plain_display(obj):
    if isinstance(obj, pd.DataFrame):
        print(obj.to_string(index=False))
    else:
        print(obj)

display = step24_plain_display

def step24_install_or_import_gspread():
    try:
        import gspread
        return gspread
    except ImportError:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "-q", "gspread", "google-auth"],
            check=True,
        )
        import gspread
        return gspread

def step24_authenticate_sheet():
    from google.colab import auth
    import google.auth

    gspread = step24_install_or_import_gspread()

    auth.authenticate_user()
    creds, _ = google.auth.default()

    return gspread.authorize(creds)

def step24_safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def step24_dataframe_from_sheet(worksheet):
    values = worksheet.get_all_values()

    if not values:
        return pd.DataFrame()

    headers = [step24_safe_text(header) for header in values[0]]

    rows = []
    for raw_row in values[1:]:
        padded = raw_row + [""] * (len(headers) - len(raw_row))
        rows.append(padded[:len(headers)])

    return pd.DataFrame(rows, columns=headers)

def step24_extract_step_code(step_id):
    workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")

    marker_pattern = re.compile(
        r"(?m)^#\s*STEP\s+([0-9]+[A-Z]?)\b.*$"
    )

    markers = list(marker_pattern.finditer(workflow_text))
    candidate_blocks = []

    for idx, marker in enumerate(markers):
        marker_step_id = marker.group(1)

        if marker_step_id != step_id:
            continue

        start = marker.start()
        end = len(workflow_text)

        for next_marker in markers[idx + 1:]:
            next_step_id = next_marker.group(1)
            if next_step_id != step_id:
                end = next_marker.start()
                break

        candidate_blocks.append(workflow_text[start:end].strip())

    if not candidate_blocks:
        raise ValueError(f"STOP: Could not find Step {step_id} in {WORKFLOW_PATH}.")

    return max(candidate_blocks, key=len)

def step24_run_workflow_step(step_id):
    print("")
    print("=" * 80)
    print(f"RUNNING STEP {step_id}")
    print("=" * 80)

    step_code = step24_extract_step_code(step_id)
    print(f"Step {step_id} block length:", len(step_code))

    start_time = time.time()

    try:
        exec(
            compile(step_code, f"colab_workflow.py::STEP_{step_id}_FROM_STEP_24", "exec"),
            globals(),
            globals(),
        )
    except Exception:
        print(f"FAILED: Step {step_id} stopped.")
        traceback.print_exc()
        raise

    elapsed = time.time() - start_time
    print(f"PASS: Step {step_id} completed in {elapsed:.1f} seconds.")

def step24_normalize_decision(value):
    text = step24_safe_text(value).lower()

    mapping = {
        "approve": "approve",
        "approved": "approve",
        "include": "approve",
        "update": "approve",
        "add": "approve",
        "downgrade": "approve",
        "hold": "hold",
        "exclude": "hold",
        "skip": "hold",
        "reject": "hold",
        "needs_more_research": "hold",
        "needs more research": "hold",
    }

    return mapping.get(text, text)

def step24_status_for_row(decision, ready):
    if ready == "YES" and decision == "approve":
        return "READY_FOR_STEP_22"
    return "HELD_FROM_STEP_22"

def step24_build_decisions(review_df, selected_batch_name_value):
    required_cols = [
        "batch_name",
        "company",
        "review_decision",
        "reviewed_priority",
        "review_notes",
        "ready_for_master_update",
    ]

    missing_cols = [col for col in required_cols if col not in review_df.columns]

    if missing_cols:
        raise ValueError("STOP: Review Packet missing columns: " + ", ".join(missing_cols))

    batch_df = review_df[
        review_df["batch_name"].astype(str).str.strip().eq(selected_batch_name_value)
    ].copy()

    if batch_df.empty:
        raise ValueError(
            f"STOP: No Review Packet rows found for batch_name={selected_batch_name_value}"
        )

    decisions = {}
    decision_rows = []

    for _, row in batch_df.iterrows():
        company = step24_safe_text(row.get("company", ""))
        raw_review_decision = step24_safe_text(row.get("review_decision", ""))
        ready_for_master_update = step24_safe_text(row.get("ready_for_master_update", "")).upper()
        reviewed_priority = step24_safe_text(row.get("reviewed_priority", ""))
        review_notes = step24_safe_text(row.get("review_notes", ""))

        if company == "":
            continue

        normalized_decision = step24_normalize_decision(raw_review_decision)

        # Human review consistency checks. These prevent accidental partial approvals
        # or silent holds caused by a blank/ambiguous review_decision cell.
        if raw_review_decision == "":
            raise ValueError(f"STOP: {company} has blank review_decision. Use approve or hold.")

        if normalized_decision not in ["approve", "hold"]:
            raise ValueError(
                f"STOP: {company} has unsupported review_decision={raw_review_decision}. "
                "Use approve or hold."
            )

        if ready_for_master_update not in ["YES", "NO", ""]:
            raise ValueError(
                f"STOP: {company} has unsupported ready_for_master_update={ready_for_master_update}. "
                "Use YES or NO."
            )

        if ready_for_master_update == "YES" and normalized_decision != "approve":
            raise ValueError(
                f"STOP: {company} is marked ready_for_master_update=YES but review_decision is not approve."
            )

        if normalized_decision == "approve" and ready_for_master_update != "YES":
            raise ValueError(
                f"STOP: {company} has review_decision=approve but ready_for_master_update is not YES."
            )

        if ready_for_master_update != "YES":
            normalized_decision = "hold"

        if normalized_decision == "approve":
            if reviewed_priority == "":
                raise ValueError(f"STOP: {company} is approved but reviewed_priority is blank.")

            if not re.search(r"\bP[0-4]\b", reviewed_priority):
                raise ValueError(
                    f"STOP: {company} reviewed_priority must include P0, P1, P2, P3, or P4. "
                    f"Got: {reviewed_priority}"
                )

            if review_notes == "":
                raise ValueError(f"STOP: {company} is approved but review_notes is blank.")

            decisions[company] = {
                "decision": "approve",
                "reviewed_priority_level": reviewed_priority,
                "review_status": "Human reviewed",
                "review_notes": review_notes,
                "priority_review_note": step24_safe_text(
                    row.get("priority_review_note", review_notes)
                ) or review_notes,
            }
        else:
            decisions[company] = {
                "decision": "hold",
                "reason": review_notes or f"Not approved for master update. Review decision: {raw_review_decision}",
                "review_notes": review_notes or f"Held from master update. Review decision: {raw_review_decision}",
            }

        decision_rows.append({
            "company": company,
            "raw_review_decision": raw_review_decision,
            "normalized_step22_decision": decisions[company]["decision"],
            "reviewed_priority": reviewed_priority,
            "ready_for_master_update": ready_for_master_update,
            "step24_preflight_status": step24_status_for_row(
                decisions[company]["decision"],
                ready_for_master_update,
            ),
        })

    if not decisions:
        raise ValueError("STOP: No company decisions were built from Review Packet.")

    return decisions, pd.DataFrame(decision_rows), batch_df


def step24_column_letter(zero_based_index):
    """Convert zero-based column index to Google Sheets column letters."""
    number = int(zero_based_index) + 1
    letters = ""

    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters

    return letters

def step24_update_review_packet_status(review_ws, selected_batch_name_value, status_text):
    values = review_ws.get_all_values()

    if not values:
        return

    headers = values[0]

    required = ["batch_name", "step_22_status"]

    for col in required:
        if col not in headers:
            print(f"WARNING: Cannot update {col}; column not found.")
            return

    batch_idx = headers.index("batch_name")
    status_idx = headers.index("step_22_status")

    updates = []

    for row_number, row_values in enumerate(values[1:], start=2):
        padded = row_values + [""] * (len(headers) - len(row_values))
        row_batch = step24_safe_text(padded[batch_idx])

        if row_batch == selected_batch_name_value:
            cell_a1 = f"{step24_column_letter(status_idx)}{row_number}"
            updates.append({
                "range": cell_a1,
                "values": [[status_text]],
            })

    if updates:
        review_ws.batch_update(updates)

print("STEP 24 - GOOGLE SHEET REVIEW APPROVAL LAUNCHER")
print("=" * 80)
print("BATCH_CONTROL_SHEET_URL:", BATCH_CONTROL_SHEET_URL)
print("DRY RUN:", STEP_24_DRY_RUN)

client_for_sheet = step24_authenticate_sheet()
spreadsheet = client_for_sheet.open_by_url(BATCH_CONTROL_SHEET_URL)

review_ws = spreadsheet.worksheet("Review Packet")
review_df = step24_dataframe_from_sheet(review_ws)

if review_df.empty:
    raise ValueError("STOP: Review Packet is empty.")

step24_selected_batch_override = step24_safe_text(globals().get("STEP24_SELECTED_BATCH_NAME", ""))

if step24_selected_batch_override:
    selected_batch_name_value = step24_selected_batch_override
else:
    candidate_df = review_df[
        review_df["ready_for_master_update"].astype(str).str.strip().str.upper().eq("YES")
    ].copy()

    unique_batches = sorted(
        [batch for batch in candidate_df["batch_name"].astype(str).str.strip().unique() if batch]
    )

    if len(unique_batches) != 1:
        raise ValueError(
            "STOP: Set STEP24_SELECTED_BATCH_NAME before running Step 24 only if multiple batches are ready. "
            f"Ready batches found: {unique_batches}"
        )

    selected_batch_name_value = unique_batches[0]

batch_name = selected_batch_name_value

batch_review_decisions, decision_preview_df, selected_review_df = step24_build_decisions(
    review_df=review_df,
    selected_batch_name_value=selected_batch_name_value,
)

print("")
print("BATCH SELECTED")
print("=" * 80)
print("batch_name:", batch_name)

print("")
print("STEP 24 DECISION PREFLIGHT")
print("=" * 80)
print(decision_preview_df.to_string(index=False))

approved_count = sum(
    1 for decision in batch_review_decisions.values()
    if decision.get("decision") == "approve"
)

held_count = sum(
    1 for decision in batch_review_decisions.values()
    if decision.get("decision") == "hold"
)

print("")
print("Approved for master update:", approved_count)
print("Held from master update:", held_count)

print("")
print("Step 22 extraction check:")
step22_code = step24_extract_step_code("22")
print("Step 22 block length:", len(step22_code))

STEP_22_DRY_RUN = STEP_24_DRY_RUN

if STEP_24_DRY_RUN:
    print("")
    print("Running Step 22 in dry-run mode for validation.")
    step24_run_workflow_step("22")

    print("")
    print("=" * 80)
    print("STEP 24 DRY RUN COMPLETE")
    print("=" * 80)
    print("No master update was performed. No dashboard refresh was performed.")
else:
    print("")
    print("Running Step 22 for real master/dashboard update.")
    step24_run_workflow_step("22")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    step24_update_review_packet_status(
        review_ws=review_ws,
        selected_batch_name_value=selected_batch_name_value,
        status_text=f"UPDATED_BY_STEP_24 {timestamp}",
    )

    print("")
    print("=" * 80)
    print("STEP 24 COMPLETE")
    print("=" * 80)
    print("Approved batch was sent through Step 22.")
    print("Master should be updated and dashboard should be refreshed.")

# =============================================================================
# STEP 26 - Rescore existing companies without fresh web research
# =============================================================================
# Purpose:
# - Reuse previously archived evidence for selected companies.
# - Re-run only the company fit synthesis / scoring prompt.
# - Do NOT run fresh web research.
# - Stop before master update.
#
# Optional before running:
# - STEP_26_DRY_RUN = True / False
# - STEP_26_BATCH_NAME = "..."
# - STEP_26_COMPANIES = [...]

from pathlib import Path
from datetime import datetime
import json
import os
import re
import shutil
import subprocess
import sys
import pandas as pd

REPO_DIR = Path("/content/health-tech-research-agent")
WORKFLOW_PATH = REPO_DIR / "colab_workflow.py"
DRIVE_FOLDER = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
DRIVE_BATCHES_FOLDER = DRIVE_FOLDER / "research_batches"
LOCAL_BATCHES_FOLDER = Path("research_batches")
LOCAL_BATCHES_FOLDER.mkdir(parents=True, exist_ok=True)

STEP_26_DRY_RUN = bool(globals().get("STEP_26_DRY_RUN", True))
STEP_26_BATCH_NAME = globals().get(
    "STEP_26_BATCH_NAME",
    "role_timing_benchmark_1_rescore_existing",
)
STEP_26_COMPANIES = globals().get(
    "STEP_26_COMPANIES",
    [
        "Omada Health",
        "Function Health",
        "Midi Health",
        "Mae Health",
        "Oshi Health",
        "Levels Health",
    ],
)

if not WORKFLOW_PATH.exists():
    raise FileNotFoundError(
        f"STOP: Could not find {WORKFLOW_PATH}. Run the GitHub setup / pull cell first."
    )

def step26_safe_text(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()

def step26_normalize_key(value):
    text = step26_safe_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def step26_text_from_any(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        pieces = [step26_text_from_any(item) for item in value]
        pieces = [piece for piece in pieces if piece]
        return "\n".join(f"- {piece}" for piece in pieces)
    if isinstance(value, dict):
        pieces = []
        for key, item in value.items():
            item_text = step26_text_from_any(item)
            if item_text:
                pieces.append(f"{key}: {item_text}")
        return "\n".join(pieces)
    return str(value).strip()

def step26_load_latest_scoring_runtime():
    """
    Reloads the setup/prompt section from the current colab_workflow.py so Step 26 uses
    the latest Step 5 scoring rubric. Strips notebook magic/shell lines before exec.
    """
    workflow_text = WORKFLOW_PATH.read_text(encoding="utf-8")
    step_6_match = re.search(r"(?m)^#\s*STEP\s+6\b", workflow_text)

    if not step_6_match:
        raise RuntimeError("Could not find Step 6 marker for setup loader.")

    setup_code = workflow_text[:step_6_match.start()]

    clean_lines = []
    for line in setup_code.splitlines():
        stripped = line.lstrip()
        if stripped.startswith("!") or stripped.startswith("%"):
            continue
        clean_lines.append(line)

    setup_code = "\n".join(clean_lines)

    exec(
        compile(setup_code, "colab_workflow.py::SETUP_LOADER_FROM_STEP_26", "exec"),
        globals(),
        globals(),
    )

    required = ["client", "MODEL", "call_openai", "run_company_fit_brief"]
    missing = [name for name in required if name not in globals()]

    if missing:
        raise RuntimeError("Research runtime missing after load: " + ", ".join(missing))

    print("PASS: loaded latest Step 5 scoring rubric.")
    print("MODEL:", globals().get("MODEL"))

def step26_load_csv_safely(path, source_name):
    try:
        if not Path(path).exists():
            return pd.DataFrame()

        temp = pd.read_csv(path)

        if "company" not in temp.columns:
            return pd.DataFrame()

        temp["source_name"] = source_name
        temp["source_file"] = str(path)
        return temp

    except Exception as e:
        print(f"Skipping unreadable CSV: {path} | {e}")
        return pd.DataFrame()


def step26_candidate_source_paths():
    """
    Returns previously archived research files that are safe to use as input evidence.

    Important:
    - Do NOT use Step 26 output files as input evidence.
    - Do NOT use rescore files as input evidence.
    - Do NOT use the accidental Step 25 benchmark files if they were partial.
    """
    paths = []

    def should_skip(path):
        name = path.name.lower()
        full = str(path).lower()

        skip_tokens = [
            # Never use generated rescore / review / test outputs as evidence.
            "rescore_existing",
            "rescore",
            "full_master_rescore",
            "priority_gate_test",
            "parse_retry_test",
            "_retry_test",
            "mae_parse_retry",
            "human_review_packet",
            "review_packet",
            "_summary",
            "step26",

            # Historical quarantine folders / partial batches.
            "_ignored_partial_step25",
        ]

        if any(token in name for token in skip_tokens):
            return True

        if any(token in full for token in skip_tokens):
            return True

        # Never let the current Step 26 batch feed back into itself.
        try:
            current_batch = str(STEP_26_BATCH_NAME).lower()
            if current_batch and current_batch in name:
                return True
        except Exception:
            pass

        return False

    archive_candidates = [
        (Path("health_tech_raw_research_ARCHIVE.csv"), "local_raw_archive"),
        (DRIVE_FOLDER / "health_tech_raw_research_ARCHIVE.csv", "drive_raw_archive"),
    ]

    for path, label in archive_candidates:
        if not should_skip(path):
            paths.append((path, label))

    for folder, label in [
        (Path("research_batches"), "local_research_batches"),
        (DRIVE_BATCHES_FOLDER, "drive_research_batches"),
    ]:
        if folder.exists():
            for pattern in ["*_checkpoint.csv", "*_raw.csv", "*_raw_*.csv"]:
                for path in folder.glob(pattern):
                    if should_skip(path):
                        continue
                    paths.append((path, label))

    seen = set()
    deduped = []

    for path, label in paths:
        key = str(path)
        if key not in seen:
            seen.add(key)
            deduped.append((path, label))

    return deduped

def step26_load_existing_evidence(companies):
    try:
        from google.colab import drive
        drive.mount("/content/drive")
    except Exception as e:
        print("Drive mount warning:", e)

    required_cols = [
        "company",
        "date_researched",
        "funding_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "commercial_scale_finding",
        "org_events_finding",
        "operating_characteristics_finding",
        "fit_brief_json",
    ]

    company_keys = {step26_normalize_key(company): company for company in companies}
    source_frames = []

    for path, label in step26_candidate_source_paths():
        temp = step26_load_csv_safely(path, label)

        if temp.empty:
            continue

        for col in required_cols:
            if col not in temp.columns:
                temp[col] = ""

        temp["company_key"] = temp["company"].apply(step26_normalize_key)
        temp = temp[temp["company_key"].isin(company_keys.keys())].copy()

        if not temp.empty:
            source_frames.append(temp)

    if not source_frames:
        raise RuntimeError("STOP: No existing archived evidence found for requested companies.")

    all_sources = pd.concat(source_frames, ignore_index=True)

    evidence_cols = [
        "funding_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "commercial_scale_finding",
        "org_events_finding",
        "operating_characteristics_finding",
    ]

    for col in evidence_cols:
        all_sources[f"has_{col}"] = all_sources[col].apply(
            lambda value: step26_safe_text(value) != ""
        )

    all_sources["evidence_completeness"] = all_sources[
        [f"has_{col}" for col in evidence_cols]
    ].sum(axis=1)

    all_sources["has_fit_brief_json"] = all_sources["fit_brief_json"].apply(
        lambda value: step26_safe_text(value) != ""
    )

    source_priority = {
        "local_research_batches": 4,
        "drive_research_batches": 3,
        "local_raw_archive": 2,
        "drive_raw_archive": 1,
    }

    all_sources["source_priority"] = all_sources["source_name"].map(source_priority).fillna(0)

    if "archive_saved_at" not in all_sources.columns:
        all_sources["archive_saved_at"] = ""

    all_sources["sort_timestamp"] = (
        all_sources["archive_saved_at"].astype(str)
        + " "
        + all_sources["date_researched"].astype(str)
    )

    picked_rows = []
    missing_companies = []

    for company in companies:
        key = step26_normalize_key(company)
        matches = all_sources[all_sources["company_key"].eq(key)].copy()

        if matches.empty:
            missing_companies.append(company)
            continue

        matches = matches.sort_values(
            by=[
                "evidence_completeness",
                "has_fit_brief_json",
                "source_priority",
                "sort_timestamp",
            ],
            ascending=[False, False, False, False],
        )

        picked = matches.iloc[0].copy()
        picked["requested_company_name"] = company
        picked_rows.append(picked)

    if missing_companies:
        raise RuntimeError("STOP: Missing existing evidence for: " + ", ".join(missing_companies))

    evidence_df = pd.DataFrame(picked_rows).reset_index(drop=True)

    return evidence_df


# STEP26 SOURCE HYGIENE ASSERTION - START
# Hard guardrail: Step 26 must use raw archived evidence, not prior Step 26 outputs.

STEP26_FORBIDDEN_EVIDENCE_SOURCE_TOKENS = [
    "rescore_existing",
    "rescore",
    "full_master_rescore",
    "priority_gate_test",
    "parse_retry_test",
    "_retry_test",
    "mae_parse_retry",
    "human_review_packet",
    "review_packet",
    "_summary",
    "step26",
]


def step26_assert_clean_evidence_sources(evidence_df):
    if evidence_df is None or evidence_df.empty:
        return

    if "source_file" not in evidence_df.columns:
        raise RuntimeError("STOP: evidence_df has no source_file column; cannot validate source hygiene.")

    generated_output_columns = [
        "fit_brief_json",
        "has_fit_brief_json",
        "rescore_batch_name",
        "rescore_saved_at",
        "parse_attempt_count",
        "parse_retry_errors",
    ]

    bad_rows = []

    for _, row in evidence_df.iterrows():
        source_file = step26_safe_text(row.get("source_file", "")).lower()
        source_name = step26_safe_text(row.get("source_name", "")).lower()
        company = step26_safe_text(row.get("requested_company_name", row.get("company", "")))

        matched_tokens = [
            token for token in STEP26_FORBIDDEN_EVIDENCE_SOURCE_TOKENS
            if token in source_file
        ]

        generated_markers = []

        for col in generated_output_columns:
            if col in evidence_df.columns:
                value = step26_safe_text(row.get(col, ""))

                if value:
                    generated_markers.append(col)

        # Generated Step 26/test outputs often come from local research_batches
        # and contain model-output columns. Raw archived evidence should not.
        if matched_tokens or generated_markers:
            bad_rows.append({
                "company": company,
                "source_name": source_name,
                "source_file": source_file,
                "matched_forbidden_tokens": ",".join(matched_tokens),
                "generated_output_markers": ",".join(generated_markers),
            })

    if bad_rows:
        preview = "\\n".join(
            (
                f"- {item['company']} | "
                f"tokens={item['matched_forbidden_tokens'] or 'none'} | "
                f"generated_markers={item['generated_output_markers'] or 'none'} | "
                f"source={item['source_file']}"
            )
            for item in bad_rows[:20]
        )

        raise RuntimeError(
            "STOP: Step 26 selected generated/rescore/test output files as evidence. "
            "This would create a scoring feedback loop.\\n" + preview
        )


# STEP26 SOURCE HYGIENE ASSERTION - END


def step26_build_status_findings(row):
    company = step26_safe_text(row.get("requested_company_name", row.get("company", "")))

    return f"""
RESCORE-ONLY PASS USING PREVIOUSLY ARCHIVED EVIDENCE.
Do not infer new external facts. Do not rely on web search. Use only the evidence below.

Company: {company}

Funding / stage evidence:
{step26_safe_text(row.get("funding_finding", ""))}

Payer / employer / provider / institutional distribution evidence:
{step26_safe_text(row.get("payer_institutional_finding", ""))}

Outcomes / engagement / product-value evidence:
{step26_safe_text(row.get("outcomes_finding", ""))}

Commercial scale / revenue-quality evidence:
{step26_safe_text(row.get("commercial_scale_finding", ""))}

Recent org / leadership events (last ~12-18 months):
{step26_safe_text(row.get("org_events_finding", ""))}

Operating characteristics (product-engagement + operational strain):
{step26_safe_text(row.get("operating_characteristics_finding", ""))}

Prior batch/source metadata, for traceability only:
- prior_batch_name: {step26_safe_text(row.get("batch_name", ""))}
- date_researched: {step26_safe_text(row.get("date_researched", ""))}
- source_file: {step26_safe_text(row.get("source_file", ""))}
""".strip()


def step26_parse_json_object(text):
    """
    Parse a JSON object from model output.

    Durable behavior:
    - Handles fenced JSON.
    - Extracts the first JSON object.
    - Repairs the common model failure where the final closing brace/bracket is missing.
    - Raises a clear error if repair still fails.
    """
    raw = step26_safe_text(text)
    raw = re.sub(r"^```json", "", raw).strip()
    raw = re.sub(r"^```", "", raw).strip()
    raw = re.sub(r"```$", "", raw).strip()

    start = raw.find("{")

    if start == -1:
        raise ValueError("No JSON object found")

    candidate = raw[start:].strip()

    def try_parse(value):
        return json.loads(value)

    try:
        return try_parse(candidate)
    except Exception as first_error:
        first_error_message = f"{type(first_error).__name__}: {first_error}"

    # Repair common EOF truncation / missing closer cases.
    repaired = candidate

    stack = []
    in_string = False
    escape = False

    for char in repaired:
        if escape:
            escape = False
            continue

        if char == "\\" and in_string:
            escape = True
            continue

        if char == '"':
            in_string = not in_string
            continue

        if in_string:
            continue

        if char in "{[":
            stack.append(char)
        elif char in "}]":
            if not stack:
                continue

            opener = stack[-1]
            if opener == "{" and char == "}":
                stack.pop()
            elif opener == "[" and char == "]":
                stack.pop()

    closer_map = {
        "{": "}",
        "[": "]",
    }

    if stack:
        repaired = repaired + "".join(closer_map[item] for item in reversed(stack))

        try:
            parsed = try_parse(repaired)

            if isinstance(parsed, dict):
                return parsed

            raise ValueError("Repaired JSON parsed but did not return an object.")

        except Exception as repair_error:
            raise ValueError(
                "Could not parse model JSON. "
                f"Initial error: {first_error_message}. "
                f"Repair error: {type(repair_error).__name__}: {repair_error}. "
                f"Raw preview: {candidate[:500]}"
            )

    raise ValueError(
        "Could not parse model JSON and no simple missing-closer repair was available. "
        f"Initial error: {first_error_message}. "
        f"Raw preview: {candidate[:500]}"
    )


def step26_extract_score(scores, key):
    value = scores.get(key, None) if isinstance(scores, dict) else None

    if isinstance(value, dict):
        value = value.get("score", value.get("value", None))

    try:
        return float(value)
    except Exception:
        return None



def step26_infer_maturity_and_cap(company, evidence_text, parsed):
    """
    Infer company maturity and timing cap.

    Public-company classification is intentionally conservative:
    - DO classify as public when there is a hard signal: IPO, ticker, exchange,
      SEC/S-1, or explicitly publicly traded.
    - DO NOT classify as public from phrases like "public evidence",
      "public outcomes data", or "public commercial metrics."
    """
    role_timing = parsed.get("role_timing_assessment", {}) if isinstance(parsed, dict) else {}

    combined = " ".join([
        company,
        evidence_text,
        step26_text_from_any(role_timing),
        step26_text_from_any(parsed.get("business_model_classification", "")),
        step26_text_from_any(parsed.get("final_takeaway", "")),
        step26_text_from_any(parsed.get("calibration_flag", "")),
    ])

    combined_lower = combined.lower()

    model_maturity = step26_safe_text(role_timing.get("company_maturity_read", "")).lower()
    maturity = model_maturity

    if maturity not in ["early", "early-growth", "scale-up", "late-stage", "public", "unclear"]:
        maturity = ""

    public_false_positive_patterns = [
        r"\bpublic evidence\b",
        r"\bpublicly available evidence\b",
        r"\bpublic outcomes?\b",
        r"\bpublic outcomes? data\b",
        r"\bpublic commercial\b",
        r"\bpublic commercial metrics?\b",
        r"\bpublic proof\b",
        r"\bpublic metrics?\b",
        r"\bpublic data\b",
        r"\bpublic source",
        r"\bpublicly available\b",
        r"\bpublic evidence does not\b",
        r"\bpublic evidence is\b",
        r"\bpublic evidence remains\b",
    ]

    # Hard public-company signals only.
    # These are intentionally stronger than plain text like "public company."
    public_company_signal_patterns = {
        "exchange_ticker": [
            r"\b(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*[:：]\s*[A-Z]{1,6}\b",
            r"\b[A-Z]{1,6}\s*\(\s*(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*\)",
            r"\b(?:ticker|stock ticker|ticker symbol)\s*(?:is|:|=)\s*[A-Z]{1,6}\b",
            r"\btrades?\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
            r"\blisted\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
        ],
        "ipo": [
            r"\bipo date\b",
            r"\binitial public offering\b",
            r"\bcompleted\s+(?:its|an|the)?\s*ipo\b",
            r"\bwent public\b",
            r"\bpost[- ]ipo\b",
        ],
        "sec_s1": [
            r"\bsec filing\b",
            r"\bfiled\s+(?:an?\s+)?s[- ]1\b",
            r"\bform\s+s[- ]1\b",
            r"\bs[- ]1 filing\b",
            r"\bregistration statement\b",
        ],
        "publicly_traded": [
            r"\bpublicly traded\b",
            r"\bpublicly listed\b",
            r"\blisted company\b",
            r"\bpublicly held company\b",
        ],
    }

    public_signal_basis = []

    for basis, patterns in public_company_signal_patterns.items():
        for pattern in patterns:
            if re.search(pattern, combined, flags=re.IGNORECASE):
                public_signal_basis.append(basis)
                break

    reliable_public_company_signal = len(public_signal_basis) > 0

    public_false_positive_language_present = any(
        re.search(pattern, combined_lower)
        for pattern in public_false_positive_patterns
    )

    public_company_language_present = any(
        re.search(pattern, combined_lower)
        for pattern in [
            r"\bas a public company\b",
            r"\bis a public company\b",
            r"\bpublic[- ]company\b",
        ]
    )

    public_company_language_without_hard_signal = (
        public_company_language_present and not reliable_public_company_signal
    )

    late_terms = [
        "series d",
        "series e",
        "series f",
        "late-stage",
        "late stage",
        "$100m arr",
        "100m arr",
        "100 million arr",
        "$100m revenue",
        "100m revenue",
        "100 million revenue",
        "unicorn",
    ]

    scaleup_terms = [
        "series c",
        "growth-stage",
        "growth stage",
        "$75m",
        "75m arr",
        "national scale",
        "scaled",
        "multi-state",
        "multi state",
    ]

    early_growth_terms = [
        "series a",
        "series b",
        "seed",
        "early-stage",
        "early stage",
        "early growth",
    ]

    # Slice 2: the maturity LABEL is DERIVED deterministically from the researched
    # funding_stage / ipo_status (structured_evidence.derive_maturity), not the
    # keyword/regex text-scan. The public-detection machinery above now only feeds the
    # reporting fields in the return dict; the label no longer comes from it.
    from health_tech_research_agent.structured_evidence import derive_maturity as _derive_maturity
    inferred, _maturity_needs_review = _derive_maturity(
        parsed.get("maturity_evidence", {}) if isinstance(parsed, dict) else {}
    )

    high_agency_terms = [
        "new business line",
        "operating model",
        "operationally immature",
        "rebuild",
        "scaling operations",
        "implementation complexity",
        "complex implementation",
        "founder-led",
        "white space",
        "zero to one",
        "0 to 1",
        "new market",
        "needs professionalization",
        "pre-professionalized",
        "underbuilt",
        "messy",
    ]

    high_agency = any(term in combined_lower for term in high_agency_terms)

    if inferred == "public":
        cap = 65 if high_agency else 50
    elif inferred == "late-stage":
        cap = 70 if high_agency else 60
    elif inferred == "scale-up":
        cap = 90 if high_agency else 80
    else:
        cap = 100

    if inferred in ["early", "early-growth"]:
        timing_fit = "ideal"
        agency = "high"
    elif inferred == "scale-up":
        timing_fit = "good" if high_agency else "borderline"
        agency = "high" if high_agency else "medium"
    elif inferred in ["late-stage", "public"]:
        timing_fit = "borderline" if high_agency else "too late"
        agency = "medium" if high_agency else "low"
    else:
        timing_fit = "unclear"
        agency = "role-dependent"

    why = step26_safe_text(role_timing.get("why_now_or_why_not", ""))

    if not why:
        why = (
            f"Inferred maturity={inferred}; "
            f"hard_public_signal={reliable_public_company_signal}; "
            f"public_signal_basis={','.join(sorted(set(public_signal_basis))) or 'none'}; "
            f"high-agency exception evidence={high_agency}."
        )

    return {
        "company_maturity_read": inferred,
        "likely_agency_level": agency,
        "stage_timing_fit": timing_fit,
        "why_now_or_why_not": why,
        "operator_timing_score_cap": cap,
        "public_company_hard_signal_present": reliable_public_company_signal,
        "public_company_signal_basis": ",".join(sorted(set(public_signal_basis))) if public_signal_basis else "",
        "public_company_language_without_hard_signal": public_company_language_without_hard_signal,
        "public_false_positive_language_present": public_false_positive_language_present,
    }

def step26_apply_timing_cap(raw_score, cap_info):
    if raw_score is None:
        return None, False

    cap = cap_info.get("operator_timing_score_cap", 100)
    capped = min(float(raw_score), float(cap))

    return capped, capped < float(raw_score)


# STEP26 PRIORITY GATE CORRECTION - START
# Deterministic correction layer:
# The LLM can suggest a priority, but final Step 26 output must obey hard gates.

def step26_priority_code(priority):
    priority = step26_safe_text(priority)
    match = re.search(r"\bP([0-4])\b", priority)
    if not match:
        return None
    return int(match.group(1))

def step26_priority_label(code):
    labels = {
        0: "P0: Highest-priority target",
        1: "P1: High-priority diligence",
        2: "P2: Worth deeper diligence",
        3: "P3: Watch list",
        4: "P4: Low priority / likely reject",
    }
    return labels.get(code, "P3: Watch list")

def step26_cap_priority_code(code, max_code):
    if code is None:
        return max_code
    return max(code, max_code)

def step26_apply_priority_gates(parsed_priority, scores, cap_info):
    original_code = step26_priority_code(parsed_priority)

    thesis = step26_extract_score(scores, "thesis_fit_score") or 0
    pmf = step26_extract_score(scores, "pmf_scale_score") or 0
    confidence = (
        step26_extract_score(scores, "evidence_confidence_score")
        or step26_extract_score(scores, "overall_confidence")
        or 0
    )
    role = step26_extract_score(scores, "katelynd_role_fit_score") or 0
    timing = step26_extract_score(scores, "operator_timing_score") or 0

    maturity = step26_safe_text(cap_info.get("company_maturity_read", "")).lower()
    stage_fit = step26_safe_text(cap_info.get("stage_timing_fit", "")).lower()
    agency = step26_safe_text(cap_info.get("likely_agency_level", "")).lower()

    # First assign a score-based priority from gates.
    if (
        thesis >= 85
        and pmf >= 80
        and role >= 80
        and timing >= 75
        and confidence >= 60
        and stage_fit != "too late"
        and maturity != "public"
        and agency != "low"
    ):
        gated_code = 0
    elif (
        thesis >= 80
        and pmf >= 75
        and role >= 75
        and timing >= 65
        and confidence >= 55
        and stage_fit != "too late"
        and agency != "low"
        and maturity != "public"
    ):
        gated_code = 1
    elif (
        thesis >= 70
        and pmf >= 65
        and role >= 65
        and timing >= 55
        and confidence >= 50
    ):
        gated_code = 2
    elif (
        thesis >= 55
        or pmf >= 50
        or role >= 55
    ):
        gated_code = 3
    else:
        gated_code = 4

    # Then apply hard blockers.
    # Lower numeric code = higher priority, so max() demotes.
    if maturity == "public":
        gated_code = step26_cap_priority_code(gated_code, 2)

    if stage_fit == "too late":
        gated_code = step26_cap_priority_code(gated_code, 2)

    if agency == "low":
        gated_code = step26_cap_priority_code(gated_code, 2)

    if confidence < 50:
        gated_code = step26_cap_priority_code(gated_code, 3)

    # If the LLM was more conservative than the gates, keep the conservative result.
    # This prevents the correction layer from upgrading companies.
    if original_code is not None:
        gated_code = max(gated_code, original_code)

    return step26_priority_label(gated_code)

# STEP26 PRIORITY GATE CORRECTION - END


def step26_recommended_decision(priority):
    priority = step26_safe_text(priority)

    if priority.startswith("P0") or priority.startswith("P1"):
        return "consider_approve"
    if priority.startswith("P2"):
        return "deeper_diligence"
    if priority.startswith("P3"):
        return "watchlist"
    if priority.startswith("P4"):
        return "likely_reject"

    return "review"


# HARD PUBLIC SIGNAL OVERRIDE - START
# This override intentionally sits after the earlier maturity function and before step26_main().
# Python will use this latest definition when step26_main() runs.

PUBLIC_COMPANY_OVERRIDES = {
    "omada health": {
        "is_public": True,
        "ticker": "OMDA",
        "exchange": "Nasdaq",
        "ipo_date": "2025-06-06",
        "basis": "manual_override_verified_public_company",
    },
}

def step26_infer_maturity_and_cap(company, evidence_text, parsed):
    """
    Infer company maturity and timing cap.

    Public-company classification is conservative:
    - Public only when there is a manual override or hard signal:
      ticker/exchange, IPO, SEC/S-1, or explicitly publicly traded.
    - Never classify as public from phrases like:
      "public evidence", "public outcomes data", or "public commercial metrics."
    """
    role_timing = parsed.get("role_timing_assessment", {}) if isinstance(parsed, dict) else {}

    combined = " ".join([
        company,
        evidence_text,
        step26_text_from_any(role_timing),
        step26_text_from_any(parsed.get("business_model_classification", "")),
        step26_text_from_any(parsed.get("final_takeaway", "")),
        step26_text_from_any(parsed.get("calibration_flag", "")),
    ])

    combined_lower = combined.lower()
    company_key = step26_normalize_key(company)

    override = PUBLIC_COMPANY_OVERRIDES.get(company_key, {})
    override_public = bool(override.get("is_public", False))

    model_maturity = step26_safe_text(role_timing.get("company_maturity_read", "")).lower()
    maturity = model_maturity

    if maturity not in ["early", "early-growth", "scale-up", "late-stage", "public", "unclear"]:
        maturity = ""

    public_false_positive_patterns = [
        r"\bpublic evidence\b",
        r"\bpublicly available evidence\b",
        r"\bpublic outcomes?\b",
        r"\bpublic outcomes? data\b",
        r"\bpublic commercial\b",
        r"\bpublic commercial metrics?\b",
        r"\bpublic proof\b",
        r"\bpublic metrics?\b",
        r"\bpublic data\b",
        r"\bpublic source",
        r"\bpublicly available\b",
        r"\bpublic evidence does not\b",
        r"\bpublic evidence is\b",
        r"\bpublic evidence remains\b",
    ]

    public_company_signal_patterns = {
        "exchange_ticker": [
            r"\b(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*[:：]\s*[A-Z]{1,6}\b",
            r"\b[A-Z]{1,6}\s*\(\s*(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*\)",
            r"\b(?:ticker|stock ticker|ticker symbol)\s*(?:is|:|=)\s*[A-Z]{1,6}\b",
            r"\btrades?\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
            r"\blisted\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
        ],
        "ipo": [
            r"\bipo date\b",
            r"\binitial public offering\b",
            r"\bcompleted\s+(?:its|an|the)?\s*ipo\b",
            r"\bwent public\b",
            r"\bpost[- ]ipo\b",
        ],
        "sec_s1": [
            r"\bsec filing\b",
            r"\bfiled\s+(?:an?\s+)?s[- ]1\b",
            r"\bform\s+s[- ]1\b",
            r"\bs[- ]1 filing\b",
            r"\bregistration statement\b",
        ],
        "publicly_traded": [
            r"\bpublicly traded\b",
            r"\bpublicly listed\b",
            r"\blisted company\b",
            r"\bpublicly held company\b",
        ],
    }

    public_signal_basis = []

    if override_public:
        override_basis = override.get("basis", "manual_override")
        ticker = override.get("ticker", "")
        if ticker:
            public_signal_basis.append(f"{override_basis}:{ticker}")
        else:
            public_signal_basis.append(override_basis)

    for basis, patterns in public_company_signal_patterns.items():
        for pattern in patterns:
            if re.search(pattern, combined, flags=re.IGNORECASE):
                public_signal_basis.append(basis)
                break

    reliable_public_company_signal = len(public_signal_basis) > 0

    public_false_positive_language_present = any(
        re.search(pattern, combined_lower)
        for pattern in public_false_positive_patterns
    )

    public_company_language_present = any(
        re.search(pattern, combined_lower)
        for pattern in [
            r"\bas a public company\b",
            r"\bis a public company\b",
            r"\bpublic[- ]company\b",
            r"\bpublic company\b",
        ]
    )

    public_company_language_without_hard_signal = (
        public_company_language_present and not reliable_public_company_signal
    )

    late_terms = [
        "series d",
        "series e",
        "series f",
        "late-stage",
        "late stage",
        "$100m arr",
        "100m arr",
        "100 million arr",
        "$100m revenue",
        "100m revenue",
        "100 million revenue",
        "unicorn",
    ]

    scaleup_terms = [
        "series c",
        "growth-stage",
        "growth stage",
        "$75m",
        "75m arr",
        "national scale",
        "scaled",
        "multi-state",
        "multi state",
    ]

    early_growth_terms = [
        "series a",
        "series b",
        "seed",
        "early-stage",
        "early stage",
        "early growth",
    ]

    # Slice 2: the maturity LABEL is DERIVED deterministically from the researched
    # funding_stage / ipo_status (structured_evidence.derive_maturity), not the
    # keyword/regex text-scan. The public-detection machinery above now only feeds the
    # reporting fields in the return dict; the label no longer comes from it.
    from health_tech_research_agent.structured_evidence import derive_maturity as _derive_maturity
    inferred, _maturity_needs_review = _derive_maturity(
        parsed.get("maturity_evidence", {}) if isinstance(parsed, dict) else {}
    )

    high_agency_terms = [
        "new business line",
        "operating model",
        "operationally immature",
        "rebuild",
        "scaling operations",
        "implementation complexity",
        "complex implementation",
        "founder-led",
        "white space",
        "zero to one",
        "0 to 1",
        "new market",
        "needs professionalization",
        "pre-professionalized",
        "underbuilt",
        "messy",
    ]

    high_agency = any(term in combined_lower for term in high_agency_terms)

    if inferred == "public":
        cap = 65 if high_agency else 50
    elif inferred == "late-stage":
        cap = 70 if high_agency else 60
    elif inferred == "scale-up":
        cap = 90 if high_agency else 80
    else:
        cap = 100

    if inferred in ["early", "early-growth"]:
        timing_fit = "ideal"
        agency = "high"
    elif inferred == "scale-up":
        timing_fit = "good" if high_agency else "borderline"
        agency = "high" if high_agency else "medium"
    elif inferred in ["late-stage", "public"]:
        timing_fit = "borderline" if high_agency else "too late"
        agency = "medium" if high_agency else "low"
    else:
        timing_fit = "unclear"
        agency = "role-dependent"

    why = step26_safe_text(role_timing.get("why_now_or_why_not", ""))

    if not why:
        why = (
            f"Inferred maturity={inferred}; "
            f"hard_public_signal={reliable_public_company_signal}; "
            f"public_signal_basis={','.join(sorted(set(public_signal_basis))) or 'none'}; "
            f"public_false_positive_language_present={public_false_positive_language_present}; "
            f"high-agency exception evidence={high_agency}."
        )

    return {
        "company_maturity_read": inferred,
        "likely_agency_level": agency,
        "stage_timing_fit": timing_fit,
        "why_now_or_why_not": why,
        "operator_timing_score_cap": cap,
        "public_company_hard_signal_present": reliable_public_company_signal,
        "public_company_signal_basis": ",".join(sorted(set(public_signal_basis))) if public_signal_basis else "",
        "public_company_language_without_hard_signal": public_company_language_without_hard_signal,
        "public_false_positive_language_present": public_false_positive_language_present,
    }

# HARD PUBLIC SIGNAL OVERRIDE - END



# STEP26 PUBLIC STATUS SAFETY CORRECTION - START
# Safety rule:
# - Public-company status requires a hard signal or manual override.
# - "public evidence", "public outcomes", etc. do not count.

STEP26_PUBLIC_COMPANY_OVERRIDES = {
    "omada health": {
        "is_public": True,
        "ticker": "OMDA",
        "exchange": "Nasdaq",
        "ipo_date": "2025-06-06",
        "basis": "manual_override_verified_public_company",
    },
}

def step26_recompute_timing_fields_for_maturity(cap_info, inferred, combined_lower):
    high_agency_terms = [
        "new business line",
        "operating model",
        "operationally immature",
        "rebuild",
        "scaling operations",
        "implementation complexity",
        "complex implementation",
        "founder-led",
        "white space",
        "zero to one",
        "0 to 1",
        "new market",
        "needs professionalization",
        "pre-professionalized",
        "underbuilt",
        "messy",
    ]

    high_agency = any(term in combined_lower for term in high_agency_terms)

    if inferred == "public":
        cap = 65 if high_agency else 50
    elif inferred == "late-stage":
        cap = 70 if high_agency else 60
    elif inferred == "scale-up":
        cap = 90 if high_agency else 80
    else:
        cap = 100

    if inferred in ["early", "early-growth"]:
        timing_fit = "ideal"
        agency = "high"
    elif inferred == "scale-up":
        timing_fit = "good" if high_agency else "borderline"
        agency = "high" if high_agency else "medium"
    elif inferred in ["late-stage", "public"]:
        timing_fit = "borderline" if high_agency else "too late"
        agency = "medium" if high_agency else "low"
    else:
        timing_fit = "unclear"
        agency = "role-dependent"

    cap_info["company_maturity_read"] = inferred
    cap_info["operator_timing_score_cap"] = cap
    cap_info["stage_timing_fit"] = timing_fit
    cap_info["likely_agency_level"] = agency

    return cap_info


def step26_correct_public_status(company, evidence_text, parsed, cap_info, raw_evidence_text=None):
    """
    Correct public-company status using only raw archived evidence for hard public signals.

    Important:
    - Do not trust model-generated parsed/final_takeaway/calibration text for hard public detection.
    - Raw evidence includes archived funding, payer/institutional, outcomes, and commercial findings.
    """
    company_key = step26_normalize_key(company)

    # Hard public detection scans raw archived evidence only.
    raw_scan_text = " ".join([
        company,
        step26_safe_text(raw_evidence_text if raw_evidence_text is not None else evidence_text),
    ])

    raw_scan_lower = raw_scan_text.lower()

    # Combined text is used only for audit flags and maturity fallback, not hard public detection.
    combined = " ".join([
        company,
        step26_safe_text(raw_evidence_text if raw_evidence_text is not None else evidence_text),
        step26_text_from_any(parsed),
        step26_text_from_any(cap_info),
    ])

    combined_lower = combined.lower()

    override = STEP26_PUBLIC_COMPANY_OVERRIDES.get(company_key, {})
    override_public = bool(override.get("is_public", False))

    public_signal_patterns = {
        "exchange_ticker": [
            r"\b(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*[:：]\s*[A-Z]{1,6}\b",
            r"\b[A-Z]{1,6}\s*\(\s*(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\s*\)",
            r"\b(?:ticker|stock ticker|ticker symbol)\s*(?:is|:|=)\s*[A-Z]{1,6}\b",
            r"\btrades?\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
            r"\blisted\s+(?:on|under)\s+(?:the\s+)?(?:nasdaq|nyse|amex|otc|otcqx|otcqb)\b",
        ],
        "ipo": [
            r"\bipo date\b",
            r"\binitial public offering\b",
            r"\bcompleted\s+(?:its|an|the)?\s*ipo\b",
            r"\bwent public\b",
            r"\bpost[- ]ipo\b",
        ],
        "sec_s1": [
            r"\bfiled\s+(?:an?\s+)?(?:form\s+)?s[- ]1\b",
            r"\bform\s+s[- ]1\b",
            r"\bs[- ]1\s+(?:registration statement|filing)\b",
            r"\bregistration statement\b.{0,80}\b(?:ipo|public offering)\b",
        ],
        "publicly_traded": [
            r"\bpublicly traded\b",
            r"\bpublicly listed\b",
            r"\blisted company\b",
            r"\bpublicly held company\b",
        ],
    }

    negative_context_terms = [
        "no ",
        "not ",
        "without ",
        "none ",
        "no evidence",
        "not public",
        "privately held",
        "private company",
        "has not",
        "have not",
        "does not",
        "did not",
    ]

    def positive_pattern_present(pattern):
        for match in re.finditer(pattern, raw_scan_text, flags=re.IGNORECASE | re.DOTALL):
            left = max(0, match.start() - 80)
            right = min(len(raw_scan_text), match.end() + 80)
            context = raw_scan_text[left:right].lower()

            if any(term in context for term in negative_context_terms):
                continue

            return True

        return False

    public_signal_basis = []

    if override_public:
        ticker = override.get("ticker", "")
        basis = override.get("basis", "manual_override")
        public_signal_basis.append(f"{basis}:{ticker}" if ticker else basis)

    for basis, patterns in public_signal_patterns.items():
        for pattern in patterns:
            if positive_pattern_present(pattern):
                public_signal_basis.append(basis)
                break

    hard_public_signal = len(public_signal_basis) > 0

    false_positive_public_language = any(
        phrase in combined_lower
        for phrase in [
            "public evidence",
            "publicly available evidence",
            "public outcomes",
            "public outcomes data",
            "public commercial",
            "public commercial metrics",
            "public proof",
            "public metrics",
            "public data",
        ]
    )

    public_company_language = any(
        phrase in combined_lower
        for phrase in [
            "as a public company",
            "is a public company",
            "public-company",
            "public company",
        ]
    )

    cap_info["public_company_hard_signal_present"] = hard_public_signal
    cap_info["public_company_signal_basis"] = ",".join(sorted(set(public_signal_basis))) if public_signal_basis else ""
    cap_info["public_company_language_without_hard_signal"] = (
        public_company_language and not hard_public_signal
    )
    cap_info["public_false_positive_language_present"] = false_positive_public_language

    # If there is no hard public signal, do not allow public-company maturity.
    if cap_info.get("company_maturity_read") == "public" and not hard_public_signal:
        if any(term in combined_lower for term in [
            "series d",
            "series e",
            "series f",
            "late-stage",
            "late stage",
            "$100m arr",
            "100m arr",
            "100 million arr",
            "$100m revenue",
            "100m revenue",
            "100 million revenue",
            "unicorn",
        ]):
            corrected = "late-stage"
        elif any(term in combined_lower for term in [
            "series c",
            "growth-stage",
            "growth stage",
            "$75m",
            "75m arr",
            "national scale",
            "scaled",
            "multi-state",
            "multi state",
        ]):
            corrected = "scale-up"
        elif any(term in combined_lower for term in [
            "series a",
            "series b",
            "seed",
            "early-stage",
            "early stage",
            "early growth",
        ]):
            corrected = "early-growth"
        else:
            corrected = "unclear"

        cap_info = step26_recompute_timing_fields_for_maturity(
            cap_info=cap_info,
            inferred=corrected,
            combined_lower=combined_lower,
        )

    # If there is a true hard public signal, enforce public maturity.
    if hard_public_signal:
        cap_info = step26_recompute_timing_fields_for_maturity(
            cap_info=cap_info,
            inferred="public",
            combined_lower=combined_lower,
        )

    return cap_info


# STEP26 PUBLIC STATUS SAFETY CORRECTION - END



# STEP26 PARSE ERROR ASSERTION - START
# Hard guardrail: Step 26 should not silently succeed with unparsed company rows.

def step26_assert_no_parse_errors(summary_df):
    if summary_df is None or summary_df.empty:
        return

    if "error" not in summary_df.columns:
        return

    error_rows = summary_df[
        summary_df["error"].apply(lambda value: step26_safe_text(value) != "")
    ].copy()

    if error_rows.empty:
        return

    preview_rows = []

    for _, row in error_rows.head(20).iterrows():
        preview_rows.append(
            f"- {step26_safe_text(row.get('company', 'UNKNOWN'))}: "
            f"{step26_safe_text(row.get('error', ''))[:300]}"
        )

    preview = "\\n".join(preview_rows)

    raise RuntimeError(
        "STOP: Step 26 produced parse errors for one or more companies. "
        "Do not update the workbook until these are fixed.\\n" + preview
    )

# STEP26 PARSE ERROR ASSERTION - END



# STEP26 COMPANY RESOLUTION - START
# Durable guardrail: STEP_26_COMPANIES=None means "all companies from master."

def step26_find_master_csv_path():
    candidate_paths = [
        Path("/content/drive/MyDrive/Job Search/Health Tech Research/health_tech_market_research_summary_MASTER.csv"),
        Path("health_tech_market_research_summary_MASTER.csv"),
    ]

    for path in candidate_paths:
        if path.exists():
            return path

    search_roots = [
        Path("/content/drive/MyDrive/Job Search/Health Tech Research"),
        Path("."),
    ]

    for root in search_roots:
        if not root.exists():
            continue

        matches = sorted(root.glob("**/health_tech_market_research_summary_MASTER.csv"))

        if matches:
            return matches[0]

    raise FileNotFoundError(
        "Could not find health_tech_market_research_summary_MASTER.csv. "
        "Cannot resolve STEP_26_COMPANIES=None."
    )


def step26_resolve_target_companies(step_26_companies):
    """
    Resolve Step 26 target companies.

    Supported modes:
    - list/tuple/set/pandas Series of company names
    - comma-separated string
    - None, empty list, empty string, or 'ALL' = all companies from master
    """
    use_master = False

    if step_26_companies is None:
        use_master = True

    elif isinstance(step_26_companies, str):
        if step_26_companies.strip().lower() in ["", "all", "*"]:
            use_master = True
        else:
            companies = [
                item.strip()
                for item in step_26_companies.split(",")
                if item.strip()
            ]

            if companies:
                return companies

            use_master = True

    else:
        try:
            companies = [
                step26_safe_text(item).strip()
                for item in list(step_26_companies)
                if step26_safe_text(item).strip()
            ]
        except TypeError:
            raise TypeError(
                "STEP_26_COMPANIES must be None, 'ALL', a comma-separated string, "
                "or an iterable of company names."
            )

        if companies:
            return companies

        use_master = True

    if use_master:
        master_path = step26_find_master_csv_path()
        master_df = pd.read_csv(master_path)

        possible_company_cols = [
            "company",
            "Company",
            "company_name",
            "Company Name",
            "requested_company_name",
        ]

        company_col = None

        for col in possible_company_cols:
            if col in master_df.columns:
                company_col = col
                break

        if company_col is None:
            raise ValueError(
                "Could not find a company-name column in master CSV. "
                f"Columns found: {list(master_df.columns)}"
            )

        companies = (
            master_df[company_col]
            .dropna()
            .astype(str)
            .map(lambda value: value.strip())
        )

        companies = [company for company in companies if company]
        companies = list(dict.fromkeys(companies))

        if not companies:
            raise ValueError(
                f"Master CSV was found at {master_path}, but no companies were resolved."
            )

        print(f"Resolved {len(companies)} companies from master: {master_path}")
        return companies

    raise RuntimeError("Unexpected Step 26 company-resolution state.")


# STEP26 COMPANY RESOLUTION - END



# STEP26 FIT BRIEF PARSE RETRY - START
# Durable guardrail: retry once when model output is malformed/truncated JSON.

def step26_run_fit_brief_with_parse_retry(company, latest_status_findings, max_attempts=2):
    """
    Run Step 5 fit brief generation and parse it.

    Behavior:
    - Attempt normal fit brief generation.
    - Parse with repair-capable parser.
    - If parsing fails, retry once.
    - If retry still fails, return the last raw output and parse errors.
      Step 26 then records the error and hard-stops before writes.
    """
    parse_errors = []
    last_raw = ""

    for attempt in range(1, max_attempts + 1):
        fit_brief_json = run_company_fit_brief(company, latest_status_findings)
        last_raw = step26_safe_text(fit_brief_json)

        try:
            parsed = step26_parse_json_object(fit_brief_json)

            if attempt > 1:
                print(f"PASS: {company} parsed successfully after retry {attempt}.")

            return fit_brief_json, parsed, attempt, parse_errors

        except Exception as error:
            error_message = f"attempt {attempt}: {type(error).__name__}: {error}"
            parse_errors.append(error_message)

            if attempt < max_attempts:
                print(f"WARNING: {company} JSON parse failed on {error_message}. Retrying once...")
                continue

            return last_raw, None, attempt, parse_errors

    return last_raw, None, max_attempts, parse_errors


# STEP26 FIT BRIEF PARSE RETRY - END


def step26_main():
    global df, summary_df

    print("STEP 26 - RESCORE EXISTING COMPANIES WITHOUT FRESH WEB RESEARCH")
    print("=" * 80)
    print("BATCH:", STEP_26_BATCH_NAME)
    print("DRY RUN:", STEP_26_DRY_RUN)
    resolved_companies = step26_resolve_target_companies(STEP_26_COMPANIES)

    print("Companies:")

    for company in resolved_companies:
        print("-", company)

    evidence_df = step26_load_existing_evidence(resolved_companies)

    print()
    print("EXISTING EVIDENCE SOURCE CHECK")
    print("=" * 80)

    source_check_cols = [
        "requested_company_name",
        "company",
        "evidence_completeness",
        "source_name",
        "source_file",
        "batch_name",
        "date_researched",
    ]

    source_check_cols = [col for col in source_check_cols if col in evidence_df.columns]
    display(evidence_df[source_check_cols])

    if STEP_26_DRY_RUN:
        print()
        print("STEP 26 DRY RUN COMPLETE")
        print("No LLM scoring was run. No files were overwritten.")
        return

    step26_load_latest_scoring_runtime()

    scored_rows = []
    summary_rows = []

    for idx, row in evidence_df.iterrows():
        company = step26_safe_text(row.get("requested_company_name", row.get("company", "")))

        print()
        print("=" * 80)
        print(f"RESCORING {idx + 1}/{len(evidence_df)}: {company}")
        print("=" * 80)
        print("Source:", row.get("source_file", ""))

        latest_status_findings = step26_build_status_findings(row)
        fit_brief_json, parsed, parse_attempt_count, parse_retry_errors = step26_run_fit_brief_with_parse_retry(
            company=company,
            latest_status_findings=latest_status_findings,
            max_attempts=2,
        )

        raw_row = row.copy()
        raw_row["company"] = company
        raw_row["fit_brief_json"] = fit_brief_json
        raw_row["parse_attempt_count"] = parse_attempt_count
        raw_row["parse_retry_errors"] = " | ".join(parse_retry_errors)
        raw_row["rescore_batch_name"] = STEP_26_BATCH_NAME
        raw_row["rescore_saved_at"] = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            if parsed is None:
                raise ValueError(
                    "Could not parse model JSON after retry. "
                    + " | ".join(parse_retry_errors)
                )
            scores = parsed.get("scores", {}) if isinstance(parsed, dict) else {}
            scale = parsed.get("scale_signal_assessment", {}) if isinstance(parsed, dict) else {}

            evidence_text = latest_status_findings + " " + step26_text_from_any(parsed)

            cap_info = step26_infer_maturity_and_cap(
                company=company,
                evidence_text=evidence_text,
                parsed=parsed,
            )

            cap_info = step26_correct_public_status(
                company=company,
                evidence_text=evidence_text,
                parsed=parsed,
                cap_info=cap_info,
                raw_evidence_text=latest_status_findings,
            )

            operator_raw = step26_extract_score(scores, "operator_timing_score")
            operator_capped, penalty_applied = step26_apply_timing_cap(operator_raw, cap_info)

            calibration_parts = []

            existing_flag = step26_safe_text(parsed.get("calibration_flag", ""))
            if existing_flag:
                calibration_parts.append(existing_flag)

            if penalty_applied:
                calibration_parts.append(
                    "CHECK: operator timing score capped for maturity / lower high-agency fit"
                )

            if cap_info["company_maturity_read"] in ["late-stage", "public"]:
                calibration_parts.append(
                    "CHECK: mature company may be too late for high-agency operator entry"
                )

            gated_priority_level = step26_apply_priority_gates(
                parsed_priority=parsed.get("priority_level", ""),
                scores=scores,
                cap_info=cap_info,
            )

            summary_rows.append({
                "batch_name": STEP_26_BATCH_NAME,
                "company": company,
                "thesis_fit_score": step26_extract_score(scores, "thesis_fit_score"),
                "pmf_scale_score": step26_extract_score(scores, "pmf_scale_score"),
                "evidence_confidence_score": (
                    step26_extract_score(scores, "evidence_confidence_score")
                    or step26_extract_score(scores, "overall_confidence")
                ),
                "katelynd_role_fit_score": step26_extract_score(scores, "katelynd_role_fit_score"),
                "operator_timing_score_raw": operator_raw,
                "operator_timing_score": operator_capped,
                "operator_timing_score_cap": cap_info["operator_timing_score_cap"],
                "public_company_hard_signal_present": cap_info.get("public_company_hard_signal_present", False),
                "public_company_signal_basis": cap_info.get("public_company_signal_basis", ""),
                "public_company_language_without_hard_signal": cap_info.get("public_company_language_without_hard_signal", False),
                "public_false_positive_language_present": cap_info.get("public_false_positive_language_present", False),
                "company_maturity_read": cap_info["company_maturity_read"],
                "likely_agency_level": cap_info["likely_agency_level"],
                "stage_timing_fit": cap_info["stage_timing_fit"],
                "why_now_or_why_not": cap_info["why_now_or_why_not"],
                "timing_penalty_applied": penalty_applied,
                "final_recommendation": parsed.get("final_recommendation", ""),
                "priority_level": gated_priority_level,
                "recommended_decision": step26_recommended_decision(gated_priority_level),
                "business_model_classification": parsed.get("business_model_classification", ""),
                "commercial_scale_assessment": step26_text_from_any(
                    parsed.get("commercial_scale_assessment", "")
                ),
                "commercial_scale_signal": scale.get("commercial_scale_signal", ""),
                "institutional_distribution_signal": scale.get(
                    "institutional_distribution_signal", ""
                ),
                "outcomes_signal": scale.get("outcomes_signal", ""),
                "plausible_near_term_scale_path": scale.get(
                    "plausible_near_term_scale_path", ""
                ),
                "scale_engine_type": scale.get("scale_engine_type", ""),
                "strong_scale_engine_present": scale.get("strong_scale_engine_present", ""),
                "calibration_flag": " | ".join([part for part in calibration_parts if part]),
                "final_takeaway": step26_text_from_any(parsed.get("final_takeaway", "")),
                "commercial_scale_finding": row.get("commercial_scale_finding", ""),
                "source_file": row.get("source_file", ""),
                "parse_attempt_count": parse_attempt_count,
                "parse_retry_errors": " | ".join(parse_retry_errors),
            })

        except Exception as e:
            summary_rows.append({
                "batch_name": STEP_26_BATCH_NAME,
                "company": company,
                "error": f"Could not parse JSON: {e}",
                "raw_preview": step26_safe_text(fit_brief_json)[:500],
                "source_file": row.get("source_file", ""),
            })

        scored_rows.append(raw_row)

    df = pd.DataFrame(scored_rows)
    summary_df = pd.DataFrame(summary_rows)
    step26_assert_no_parse_errors(summary_df)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    LOCAL_BATCHES_FOLDER.mkdir(parents=True, exist_ok=True)
    DRIVE_BATCHES_FOLDER.mkdir(parents=True, exist_ok=True)

    checkpoint_path = LOCAL_BATCHES_FOLDER / f"{STEP_26_BATCH_NAME}_checkpoint.csv"
    summary_path = LOCAL_BATCHES_FOLDER / f"{STEP_26_BATCH_NAME}_summary.csv"
    review_path = LOCAL_BATCHES_FOLDER / f"{STEP_26_BATCH_NAME}_human_review_packet.csv"

    df.to_csv(checkpoint_path, index=False)
    summary_df.to_csv(summary_path, index=False)
    summary_df.to_csv(review_path, index=False)

    shutil.copy(checkpoint_path, DRIVE_BATCHES_FOLDER / checkpoint_path.name)
    shutil.copy(summary_path, DRIVE_BATCHES_FOLDER / summary_path.name)
    shutil.copy(review_path, DRIVE_BATCHES_FOLDER / review_path.name)

    print()
    print("STEP 26 COMPLETE")
    print("=" * 80)
    print("No fresh web research was run.")
    print("Master was not updated.")
    print("Local checkpoint:", checkpoint_path)
    print("Local summary:", summary_path)
    print("Local review packet:", review_path)

    print()
    print("RESCORE SUMMARY")
    print("=" * 80)

    display_cols = [
        "company",
        "error",
        "raw_preview",
        "priority_level",
        "recommended_decision",
        "thesis_fit_score",
        "pmf_scale_score",
        "evidence_confidence_score",
        "katelynd_role_fit_score",
        "operator_timing_score_raw",
        "operator_timing_score",
        "operator_timing_score_cap",
        "public_company_hard_signal_present",
        "public_company_signal_basis",
        "public_company_language_without_hard_signal",
        "public_false_positive_language_present",
        "company_maturity_read",
        "likely_agency_level",
        "stage_timing_fit",
        "timing_penalty_applied",
        "calibration_flag",
        "final_takeaway",
    ]

    display_cols = [col for col in display_cols if col in summary_df.columns]

    display(summary_df[display_cols])

step26_main()



# STEP 27 - LLM taxonomy backfill for existing master
# Purpose:
# - Classify existing master companies using the controlled taxonomy.
# - Make LLM taxonomy assignment the source of truth.
# - Validate all outputs against allowed taxonomy codes.
# - Persist taxonomy fields to the active master without rerunning research.

from pathlib import Path
from datetime import datetime
import os
import re
import json
import shutil
import subprocess
import sys
import pandas as pd

from google.colab import drive

drive.mount("/content/drive", force_remount=False)

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
TAXONOMY_DIR = REPO_DIR / "taxonomy"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from health_tech_research_agent.taxonomy import (
    safe_text,
    split_tags,
    join_tags,
    normalize_code,
    allowed_codes,
    load_taxonomy_tables,
    code_label_maps,
)

MASTER_PATH = Path("/content/drive/MyDrive/Job Search/Health Tech Research/health_tech_market_research_summary_MASTER.csv")
RESEARCH_DIR = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
OUTPUT_DIR = RESEARCH_DIR / "taxonomy_backfills"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TAXONOMY_BACKFILL_MODEL = globals().get("TAXONOMY_BACKFILL_MODEL", "gpt-5.4-mini")
TAXONOMY_BACKFILL_CONFIDENCE_REVIEW_THRESHOLD = int(globals().get("TAXONOMY_BACKFILL_CONFIDENCE_REVIEW_THRESHOLD", 65))

print("Taxonomy backfill model:", TAXONOMY_BACKFILL_MODEL)
print("Confidence review threshold:", TAXONOMY_BACKFILL_CONFIDENCE_REVIEW_THRESHOLD)

if not MASTER_PATH.exists():
    raise FileNotFoundError(f"Master file not found: {MASTER_PATH}")

# ---------------------------------------------------------------------
# Load taxonomy tables.
# ---------------------------------------------------------------------

tables = load_taxonomy_tables(TAXONOMY_DIR)
code_to_label, label_to_code = code_label_maps(tables)

market_segments = tables["market_segments"]
subsegments = tables["subsegment_tags"]
product_models = tables["product_models"]
distribution_models = tables["distribution_models"]
data_layers = tables["data_input_layers"]

allowed_primary = allowed_codes(market_segments, "segment_code")
allowed_subsegments = allowed_codes(subsegments, "tag_code")
allowed_product = allowed_codes(product_models, "product_model_code")
allowed_distribution = allowed_codes(distribution_models, "distribution_model_code")
allowed_data = allowed_codes(data_layers, "data_input_code")

if not allowed_primary:
    raise RuntimeError("No allowed primary market segments loaded.")

print("Allowed primary segments:", len(allowed_primary))
print("Allowed subsegment tags:", len(allowed_subsegments))
print("Allowed product model tags:", len(allowed_product))
print("Allowed distribution model tags:", len(allowed_distribution))
print("Allowed data input tags:", len(allowed_data))

# ---------------------------------------------------------------------
# Build taxonomy prompt block directly from current CSVs.
# ---------------------------------------------------------------------

def taxonomy_options_block():
    lines = []
    lines.append("CONTROLLED HEALTH-TECH TAXONOMY")
    lines.append("")
    lines.append("Primary market segment: choose exactly ONE segment_code.")
    lines.append("Do not invent new primary market segments.")
    lines.append("If multiple categories seem plausible, choose the broader mutually exclusive umbrella segment and put nuance in tags.")
    lines.append("")

    lines.append("PRIMARY MARKET SEGMENTS")
    for _, row in market_segments.iterrows():
        code = safe_text(row.get("segment_code"))
        label = safe_text(row.get("segment_label"))
        definition = safe_text(row.get("definition"))
        assignment_rule = safe_text(row.get("assignment_rule"))
        if code:
            lines.append(f"- {code}: {label}. Definition: {definition}. Assignment rule: {assignment_rule}")

    lines.append("")
    lines.append("SUBSEGMENT TAGS")
    for _, row in subsegments.iterrows():
        code = safe_text(row.get("tag_code"))
        label = safe_text(row.get("tag_label"))
        parent = safe_text(row.get("parent_market_segment"))
        definition = safe_text(row.get("definition"))
        if code:
            lines.append(f"- {code}: {label}. Parent: {parent}. Definition: {definition}")

    lines.append("")
    lines.append("PRODUCT MODEL TAGS")
    for _, row in product_models.iterrows():
        code = safe_text(row.get("product_model_code"))
        label = safe_text(row.get("product_model_label"))
        definition = safe_text(row.get("definition"))
        if code:
            lines.append(f"- {code}: {label}. Definition: {definition}")

    lines.append("")
    lines.append("DISTRIBUTION MODEL TAGS")
    for _, row in distribution_models.iterrows():
        code = safe_text(row.get("distribution_model_code"))
        label = safe_text(row.get("distribution_model_label"))
        definition = safe_text(row.get("definition"))
        if code:
            lines.append(f"- {code}: {label}. Definition: {definition}")

    lines.append("")
    lines.append("DATA INPUT TAGS")
    for _, row in data_layers.iterrows():
        code = safe_text(row.get("data_input_code"))
        label = safe_text(row.get("data_input_label"))
        definition = safe_text(row.get("definition"))
        if code:
            lines.append(f"- {code}: {label}. Definition: {definition}")

    return "\n".join(lines)

TAXONOMY_PROMPT_BLOCK = taxonomy_options_block()

# ---------------------------------------------------------------------
# OpenAI client.
# ---------------------------------------------------------------------

try:
    from openai import OpenAI
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openai"], check=True)
    from openai import OpenAI

api_key = os.environ.get("OPENAI_API_KEY", "")

try:
    from google.colab import userdata
    api_key = userdata.get("OPENAI_API_KEY") or api_key
except Exception:
    pass

if api_key:
    client = OpenAI(api_key=api_key)
else:
    client = OpenAI()

# ---------------------------------------------------------------------
# Output schema.
# ---------------------------------------------------------------------

taxonomy_schema = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "primary_market_segment_code": {
            "type": "string",
            "enum": sorted(list(allowed_primary)),
        },
        "subsegment_tags": {
            "type": "array",
            "items": {
                "type": "string",
                "enum": sorted(list(allowed_subsegments)),
            },
        },
        "product_model_tags": {
            "type": "array",
            "items": {
                "type": "string",
                "enum": sorted(list(allowed_product)),
            },
        },
        "distribution_model_tags": {
            "type": "array",
            "items": {
                "type": "string",
                "enum": sorted(list(allowed_distribution)),
            },
        },
        "data_input_tags": {
            "type": "array",
            "items": {
                "type": "string",
                "enum": sorted(list(allowed_data)),
            },
        },
        "taxonomy_confidence": {
            "type": "integer",
        },
        "taxonomy_rationale": {
            "type": "string",
        },
        "taxonomy_needs_review": {
            "type": "boolean",
        },
        "taxonomy_review_reason": {
            "type": "string",
        },
    },
    "required": [
        "primary_market_segment_code",
        "subsegment_tags",
        "product_model_tags",
        "distribution_model_tags",
        "data_input_tags",
        "taxonomy_confidence",
        "taxonomy_rationale",
        "taxonomy_needs_review",
        "taxonomy_review_reason",
    ],
}

SYSTEM_PROMPT = """
You are a strict health-tech taxonomy classifier.

Your job:
- Assign exactly one primary_market_segment_code from the controlled taxonomy.
- Put nuance into subsegment_tags, product_model_tags, distribution_model_tags, and data_input_tags.
- Do not invent codes.
- Do not use distribution, modality, data source, or care model as the primary segment unless the taxonomy explicitly defines that as the market problem.

Critical rules:
- D2C, employer, payer, Medicaid, and provider network are distribution_model_tags, not primary market segments.
- Wearables, CGM, labs, imaging, biomarkers, and EHR are data/product inputs, not primary market segments.
- Virtual care, marketplace, coaching, diagnostics, and navigation are product models unless the company’s actual market is infrastructure or access/navigation.
- Consumer labs/biomarkers/screening/longevity companies such as Function Health, InsideTracker, and Neko Health map to CONSUMER_HEALTH_OPTIMIZATION, not DIAGNOSTICS_LIFE_SCIENCES.
- Clinical diagnostics, pharma, trials, RWE, drug discovery, and medicine/treatment advancement map to DIAGNOSTICS_LIFE_SCIENCES.
- Equip-style eating disorder care maps to MENTAL_BEHAVIORAL_HEALTH.
- Pomelo/Mae-style maternity or maternal care maps to WOMENS_FAMILY_HEALTH.
- Oncology patient support/navigation maps to SPECIALTY_CONDITION_CARE with oncology_cancer, not diagnostics/life sciences merely because it uses data or patient support.
"""

def extract_response_text(response):
    text = getattr(response, "output_text", None)
    if text:
        return text

    try:
        chunks = []
        for item in response.output:
            for content in getattr(item, "content", []):
                value = getattr(content, "text", None)
                if value:
                    chunks.append(value)
        if chunks:
            return "\n".join(chunks)
    except Exception:
        pass

    return str(response)

def parse_json_object(raw_text):
    raw_text = safe_text(raw_text)

    try:
        return json.loads(raw_text)
    except Exception:
        pass

    match = re.search(r"\{.*\}", raw_text, flags=re.DOTALL)
    if match:
        return json.loads(match.group(0))

    raise ValueError(f"Could not parse JSON object from response: {raw_text[:500]}")

def call_taxonomy_llm(company_context):
    user_prompt = f"""
{TAXONOMY_PROMPT_BLOCK}

Classify this company.

COMPANY CONTEXT:
{company_context}

Return only the requested JSON object.
"""

    # Prefer Responses API structured output.
    try:
        response = client.responses.create(
            model=TAXONOMY_BACKFILL_MODEL,
            input=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            text={
                "format": {
                    "type": "json_schema",
                    "name": "taxonomy_classification",
                    "strict": True,
                    "schema": taxonomy_schema,
                }
            },
        )

        return parse_json_object(extract_response_text(response)), "responses_structured_output"

    except Exception as structured_error:
        print("Structured output call failed; falling back to JSON mode.")
        print("Structured error:", structured_error)

        response = client.chat.completions.create(
            model=TAXONOMY_BACKFILL_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
        )

        return parse_json_object(response.choices[0].message.content), "chat_json_mode_fallback"

def build_company_context(row):
    fields = [
        "company",
        "business_model_classification",
        "final_takeaway",
        "commercial_scale_assessment",
        "pmf_scale_assessment",
        "market_segment",
        "strategic_bucket",
        "priority_level",
        "reviewed_priority_level",
        "final_priority_level",
        "priority_review_note",
        "review_status",
        "review_notes",
        "calibration_flag",
        "funding_finding",
        "commercial_scale_finding",
        "payer_institutional_finding",
        "outcomes_finding",
        "katelynd_fit_reasoning",
        "why_now_or_why_not",
    ]

    parts = []
    for field in fields:
        if field in row.index:
            value = safe_text(row.get(field, ""))
            if value:
                parts.append(f"{field}: {value}")

    return "\n".join(parts)

def validate_llm_result(result):
    result = dict(result)

    primary = normalize_code(
        result.get("primary_market_segment_code", ""),
        allowed_primary,
        label_to_code,
    )

    if not primary:
        primary = "OTHER_REVIEW"

    def clean_tags(raw_value, allowed):
        cleaned = []
        if isinstance(raw_value, list):
            values = raw_value
        else:
            values = split_tags(raw_value)

        for value in values:
            value = safe_text(value)
            if value in allowed and value not in cleaned:
                cleaned.append(value)

        return cleaned

    sub_tags = clean_tags(result.get("subsegment_tags", []), allowed_subsegments)
    product_tags = clean_tags(result.get("product_model_tags", []), allowed_product)
    distribution_tags = clean_tags(result.get("distribution_model_tags", []), allowed_distribution)
    data_tags = clean_tags(result.get("data_input_tags", []), allowed_data)

    try:
        confidence = int(result.get("taxonomy_confidence", 0))
    except Exception:
        confidence = 0

    confidence = max(0, min(100, confidence))

    needs_review = bool(result.get("taxonomy_needs_review", False))
    review_reason = safe_text(result.get("taxonomy_review_reason", ""))
    rationale = safe_text(result.get("taxonomy_rationale", ""))

    if primary == "OTHER_REVIEW":
        needs_review = True
        review_reason = review_reason or "LLM could not assign an approved primary market segment."

    if confidence < TAXONOMY_BACKFILL_CONFIDENCE_REVIEW_THRESHOLD:
        needs_review = True
        review_reason = review_reason or f"Taxonomy confidence below threshold: {confidence}"

    return {
        "primary_market_segment_code": primary,
        "primary_market_segment": code_to_label.get(primary, primary),
        "market_segment": code_to_label.get(primary, primary),
        "subsegment_tags": join_tags(sub_tags),
        "product_model_tags": join_tags(product_tags),
        "distribution_model_tags": join_tags(distribution_tags),
        "data_input_tags": join_tags(data_tags),
        "taxonomy_confidence": confidence,
        "taxonomy_rationale": rationale,
        "taxonomy_needs_review": "YES" if needs_review else "NO",
        "taxonomy_review_reason": review_reason,
    }

# ---------------------------------------------------------------------
# Run backfill.
# ---------------------------------------------------------------------

master_df = pd.read_csv(MASTER_PATH).fillna("")

if "company" not in master_df.columns:
    raise RuntimeError("Master file must contain a company column.")

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_path = MASTER_PATH.with_name(f"{MASTER_PATH.stem}_backup_before_taxonomy_backfill_{timestamp}.csv")
shutil.copy2(MASTER_PATH, backup_path)

print("Master backup created:", backup_path)
print("Companies to classify:", master_df["company"].nunique())

results = []

for idx, row in master_df.iterrows():
    company = safe_text(row.get("company"))

    print("")
    print("=" * 80)
    print(f"Classifying taxonomy for {idx + 1}/{len(master_df)}: {company}")
    print("=" * 80)

    context = build_company_context(row)
    raw_result, api_mode = call_taxonomy_llm(context)
    validated = validate_llm_result(raw_result)

    validated["company"] = company
    validated["api_mode"] = api_mode
    validated["taxonomy_assignment_method"] = "llm_taxonomy_backfill"
    validated["taxonomy_assignment_basis"] = validated["taxonomy_rationale"]

    results.append(validated)

    print(
        company,
        "→",
        validated["primary_market_segment_code"],
        "| confidence:",
        validated["taxonomy_confidence"],
        "| review:",
        validated["taxonomy_needs_review"],
    )

results_df = pd.DataFrame(results)

# Persist fields to master.
fields_to_write = [
    "primary_market_segment_code",
    "primary_market_segment",
    "market_segment",
    "subsegment_tags",
    "product_model_tags",
    "distribution_model_tags",
    "data_input_tags",
    "taxonomy_confidence",
    "taxonomy_rationale",
    "taxonomy_needs_review",
    "taxonomy_review_reason",
    "taxonomy_assignment_method",
    "taxonomy_assignment_basis",
]

for col in fields_to_write:
    if col not in master_df.columns:
        master_df[col] = ""

for _, result_row in results_df.iterrows():
    company = result_row["company"]
    mask = master_df["company"].astype(str).str.strip().eq(company)

    if not mask.any():
        continue

    idx = master_df[mask].index[0]

    for col in fields_to_write:
        master_df.at[idx, col] = result_row.get(col, "")

master_df.to_csv(MASTER_PATH, index=False)

backfill_path = OUTPUT_DIR / f"taxonomy_backfill_results_{timestamp}.csv"
results_df.to_csv(backfill_path, index=False)

review_packet = results_df[
    (results_df["taxonomy_needs_review"].astype(str).str.upper() == "YES")
    | (results_df["primary_market_segment_code"].astype(str) == "OTHER_REVIEW")
].copy()

review_packet_path = OUTPUT_DIR / f"taxonomy_backfill_review_packet_{timestamp}.csv"
review_packet.to_csv(review_packet_path, index=False)

print("")
print("=" * 80)
print("TAXONOMY BACKFILL COMPLETE")
print("=" * 80)
print("Master updated:", MASTER_PATH)
print("Backfill results:", backfill_path)
print("Review packet:", review_packet_path)
print("Companies needing review:", len(review_packet))

print("")
print("Primary segment summary:")
display(
    results_df
    .groupby(["primary_market_segment", "primary_market_segment_code"], dropna=False)["company"]
    .nunique()
    .reset_index(name="company_count")
    .sort_values(["company_count", "primary_market_segment"], ascending=[False, True])
)

print("")
print("Review packet preview:")
if len(review_packet):
    display(review_packet[[
        "company",
        "primary_market_segment_code",
        "taxonomy_confidence",
        "taxonomy_review_reason",
        "taxonomy_rationale",
    ]])
else:
    print("No taxonomy review items.")



# STEP 28_DEPRECATED - Taxonomy QA and adjudication pass
# Purpose:
# - Audit LLM-created taxonomy assignments for suspicious mappings.
# - Use deterministic QA checks to find likely mistakes.
# - Use a second-pass LLM adjudicator only on flagged rows.
# - Auto-apply only high-confidence corrections.
# - Preserve human review only for genuinely ambiguous cases.

from pathlib import Path
from datetime import datetime
import os
import re
import json
import shutil
import subprocess
import sys
import pandas as pd

from google.colab import drive

drive.mount("/content/drive", force_remount=False)

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
TAXONOMY_DIR = REPO_DIR / "taxonomy"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from health_tech_research_agent.taxonomy import (
    safe_text,
    split_tags,
    join_tags,
    normalize_code,
    allowed_codes,
    load_taxonomy_tables,
    code_label_maps,
)

MASTER_PATH = Path("/content/drive/MyDrive/Job Search/Health Tech Research/health_tech_market_research_summary_MASTER.csv")
RESEARCH_DIR = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
OUTPUT_DIR = RESEARCH_DIR / "taxonomy_backfills"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TAXONOMY_QA_MODEL = globals().get("TAXONOMY_QA_MODEL", "gpt-5.4-mini")
TAXONOMY_QA_CONFIDENCE_AUTO_APPLY_THRESHOLD = int(globals().get("TAXONOMY_QA_CONFIDENCE_AUTO_APPLY_THRESHOLD", 80))

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

if not MASTER_PATH.exists():
    raise FileNotFoundError(f"Master not found: {MASTER_PATH}")

master_df = pd.read_csv(MASTER_PATH).fillna("")

tables = load_taxonomy_tables(TAXONOMY_DIR)
code_to_label, label_to_code = code_label_maps(tables)

market_segments = tables["market_segments"]
subsegments = tables["subsegment_tags"]
product_models = tables["product_models"]
distribution_models = tables["distribution_models"]
data_layers = tables["data_input_layers"]

allowed_primary = allowed_codes(market_segments, "segment_code")
allowed_subsegments = allowed_codes(subsegments, "tag_code")
allowed_product = allowed_codes(product_models, "product_model_code")
allowed_distribution = allowed_codes(distribution_models, "distribution_model_code")
allowed_data = allowed_codes(data_layers, "data_input_code")

subsegment_parent = {}
for _, row in subsegments.iterrows():
    tag = safe_text(row.get("tag_code"))
    parent = safe_text(row.get("parent_market_segment"))
    if tag and parent:
        subsegment_parent[tag] = parent

def filter_subsegments_for_primary(tags, primary_code):
    clean = []
    for tag in tags:
        tag = safe_text(tag)
        if not tag:
            continue
        parent = subsegment_parent.get(tag)
        if parent and parent != primary_code:
            continue
        if tag not in clean:
            clean.append(tag)
    return clean

def row_text(row):
    fields = [
        "company",
        "business_model_classification",
        "final_takeaway",
        "taxonomy_rationale",
        "taxonomy_assignment_basis",
        "commercial_scale_assessment",
        "pmf_scale_assessment",
        "priority_review_note",
        "review_notes",
        "funding_finding",
        "commercial_scale_finding",
        "payer_institutional_finding",
        "outcomes_finding",
    ]
    return " | ".join(safe_text(row.get(field, "")) for field in fields).lower()

def tags_from_row(row, col):
    return split_tags(row.get(col, ""))

def deterministic_taxonomy_flags(row):
    flags = []
    primary = safe_text(row.get("primary_market_segment_code"))
    company = safe_text(row.get("company"))
    text = row_text(row)

    sub_tags = tags_from_row(row, "subsegment_tags")

    mismatched_tags = []
    for tag in sub_tags:
        expected_parent = subsegment_parent.get(tag)
        if expected_parent and expected_parent != primary:
            mismatched_tags.append(f"{tag} expects {expected_parent}")

    if mismatched_tags:
        flags.append("SUBSEGMENT_PARENT_MISMATCH: " + "; ".join(mismatched_tags))

    metabolic_tags = {
        "obesity_glp1",
        "diabetes_cardiometabolic",
        "nutrition_care",
        "food_as_medicine",
        "cgm_glucose_behavior",
    }

    if primary == "CONSUMER_HEALTH_OPTIMIZATION":
        if any(tag in metabolic_tags for tag in sub_tags):
            flags.append("CONSUMER_PRIMARY_WITH_METABOLIC_SUBSEGMENT")

    if primary == "DIAGNOSTICS_LIFE_SCIENCES":
        consumer_longevity_signals = [
            "consumer",
            "longevity",
            "health optimization",
            "preventive wellness",
            "membership",
            "cash-pay",
            "cash pay",
        ]
        if any(signal in text for signal in consumer_longevity_signals):
            flags.append("DIAGNOSTICS_PRIMARY_WITH_CONSUMER_LONGEVITY_TEXT")

    if primary == "CARE_NAVIGATION_ACCESS":
        # These are regex-based and intentionally avoid bare substring checks.
        # Example: "gi" should only match GI as a standalone acronym, not inside unrelated words.
        condition_signal_patterns = [
            (r"\boncology\b|\bcancer\b", "SPECIALTY_CONDITION_CARE", "oncology/cancer"),
            (r"\bgi\b|\bgastroenterology\b|\bgastrointestinal\b|\bdigestive\b|\bibs\b|\bibd\b|\bcrohn\b|\bcolitis\b", "SPECIALTY_CONDITION_CARE", "GI/digestive"),
            (r"\bmsk\b|\bmusculoskeletal\b|\bphysical therapy\b", "SPECIALTY_CONDITION_CARE", "MSK"),
            (r"\bkidney\b|\brenal\b", "SPECIALTY_CONDITION_CARE", "kidney/renal"),
        ]

        for pattern, expected_parent, label in condition_signal_patterns:
            if re.search(pattern, text, flags=re.IGNORECASE):
                flags.append(f"CARE_NAV_PRIMARY_WITH_CONDITION_SIGNAL:{label}->{expected_parent}")
                break

    if primary == "PAYER_BENEFITS_INFRASTRUCTURE":
        clinical_terms = [
            "maternity",
            "maternal",
            "pregnancy",
            "postpartum",
            "eating disorder",
            "therapy",
            "oncology",
            "cancer",
            "diabetes",
            "nutrition",
            "gi",
            "msk",
        ]
        if any(term in text for term in clinical_terms):
            flags.append("PAYER_PRIMARY_WITH_CLINICAL_CONDITION_TEXT")

    known_high_risk = {
        "levels health": "Known CGM/metabolic behavior company; verify against METABOLIC_NUTRITION_HEALTH.",
        "zoe": "Known nutrition/metabolic/microbiome company; verify against METABOLIC_NUTRITION_HEALTH.",
        "jasper health": "Known oncology support/navigation company; verify against SPECIALTY_CONDITION_CARE.",
    }

    key = company.lower().strip()
    if key in known_high_risk:
        flags.append("KNOWN_HIGH_RISK_TAXONOMY_CHECK: " + known_high_risk[key])

    return flags

audit_rows = []

for idx, row in master_df.iterrows():
    flags = deterministic_taxonomy_flags(row)
    if flags:
        audit_rows.append({
            "row_index": idx,
            "company": safe_text(row.get("company")),
            "current_primary_market_segment_code": safe_text(row.get("primary_market_segment_code")),
            "current_primary_market_segment": safe_text(row.get("primary_market_segment")),
            "subsegment_tags": safe_text(row.get("subsegment_tags")),
            "product_model_tags": safe_text(row.get("product_model_tags")),
            "distribution_model_tags": safe_text(row.get("distribution_model_tags")),
            "data_input_tags": safe_text(row.get("data_input_tags")),
            "taxonomy_confidence": safe_text(row.get("taxonomy_confidence")),
            "taxonomy_rationale": safe_text(row.get("taxonomy_rationale")),
            "qa_flags": " | ".join(flags),
        })

audit_df = pd.DataFrame(audit_rows)

audit_path = OUTPUT_DIR / f"taxonomy_qa_audit_{timestamp}.csv"
audit_df.to_csv(audit_path, index=False)

print("Deterministic taxonomy QA audit complete.")
print("Flagged companies:", len(audit_df))
print("Audit path:", audit_path)

if len(audit_df):
    display(audit_df)
else:
    print("No deterministic flags found.")
    print("No second-pass LLM adjudication needed.")

if not audit_df.empty:
    try:
        from openai import OpenAI
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openai"], check=True)
        from openai import OpenAI

    api_key = os.environ.get("OPENAI_API_KEY", "")
    try:
        from google.colab import userdata
        api_key = userdata.get("OPENAI_API_KEY") or api_key
    except Exception:
        pass

    if api_key:
        client = OpenAI(api_key=api_key)
    else:
        client = OpenAI()

    taxonomy_options = {
        "primary_market_segment_codes": sorted(list(allowed_primary)),
        "subsegment_tag_codes": sorted(list(allowed_subsegments)),
        "product_model_codes": sorted(list(allowed_product)),
        "distribution_model_codes": sorted(list(allowed_distribution)),
        "data_input_codes": sorted(list(allowed_data)),
    }

    adjudication_schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "current_assignment_valid": {"type": "boolean"},
            "corrected_primary_market_segment_code": {
                "type": "string",
                "enum": sorted(list(allowed_primary)),
            },
            "corrected_subsegment_tags": {
                "type": "array",
                "items": {"type": "string", "enum": sorted(list(allowed_subsegments))},
            },
            "corrected_product_model_tags": {
                "type": "array",
                "items": {"type": "string", "enum": sorted(list(allowed_product))},
            },
            "corrected_distribution_model_tags": {
                "type": "array",
                "items": {"type": "string", "enum": sorted(list(allowed_distribution))},
            },
            "corrected_data_input_tags": {
                "type": "array",
                "items": {"type": "string", "enum": sorted(list(allowed_data))},
            },
            "qa_confidence": {"type": "integer"},
            "qa_decision_rationale": {"type": "string"},
            "needs_human_review": {"type": "boolean"},
            "human_review_reason": {"type": "string"},
        },
        "required": [
            "current_assignment_valid",
            "corrected_primary_market_segment_code",
            "corrected_subsegment_tags",
            "corrected_product_model_tags",
            "corrected_distribution_model_tags",
            "corrected_data_input_tags",
            "qa_confidence",
            "qa_decision_rationale",
            "needs_human_review",
            "human_review_reason",
        ],
    }

    system_prompt = """
You are a skeptical taxonomy QA reviewer.

Your job:
- Review a prior taxonomy assignment.
- Decide whether it follows the controlled taxonomy.
- Correct it when the primary market segment is wrong.
- Be strict about not confusing distribution, data input, or product model with market segment.
- Treat deterministic QA flags as hypotheses to investigate, not facts.
- Do not infer a GI/digestive market from a stray substring. Require explicit GI, gastroenterology, digestive, IBS, IBD, Crohn's, or colitis evidence.
- If the corrected primary segment differs from the current primary segment, set needs_human_review=true unless the evidence is extremely explicit.
- Subsegment tags should belong under the corrected primary market segment. Do not include subsegment tags whose parent segment conflicts with the corrected primary segment.

Key governance:
- Primary market segment = the main buyer/user problem.
- Distribution model = who pays / controls access.
- Product model = how value is delivered.
- Data input layer = what powers the product.
- Consumer labs/biomarkers/longevity can be CONSUMER_HEALTH_OPTIMIZATION.
- CGM/nutrition/metabolic behavior companies generally belong in METABOLIC_NUTRITION_HEALTH, not consumer optimization.
- Oncology support/navigation belongs in SPECIALTY_CONDITION_CARE with oncology_cancer, not broad care navigation, unless the company is broad cross-condition navigation.
"""

    def build_adjudication_context(master_row, audit_row):
        fields = [
            "company",
            "business_model_classification",
            "final_takeaway",
            "commercial_scale_assessment",
            "pmf_scale_assessment",
            "priority_review_note",
            "review_notes",
            "funding_finding",
            "commercial_scale_finding",
            "payer_institutional_finding",
            "outcomes_finding",
            "taxonomy_rationale",
        ]

        context = []
        for field in fields:
            if field in master_row.index:
                value = safe_text(master_row.get(field, ""))
                if value:
                    context.append(f"{field}: {value}")

        context.append("")
        context.append("CURRENT TAXONOMY ASSIGNMENT:")
        for field in [
            "primary_market_segment_code",
            "subsegment_tags",
            "product_model_tags",
            "distribution_model_tags",
            "data_input_tags",
            "taxonomy_confidence",
            "taxonomy_rationale",
        ]:
            context.append(f"{field}: {safe_text(master_row.get(field, ''))}")

        context.append("")
        context.append("DETERMINISTIC QA FLAGS:")
        context.append(safe_text(audit_row.get("qa_flags", "")))

        return "\\n".join(context)

    def parse_json_response(response):
        text = getattr(response, "output_text", None)
        if not text:
            try:
                chunks = []
                for item in response.output:
                    for content in getattr(item, "content", []):
                        value = getattr(content, "text", None)
                        if value:
                            chunks.append(value)
                text = "\\n".join(chunks)
            except Exception:
                text = str(response)

        try:
            return json.loads(text)
        except Exception:
            match = re.search(r"\\{.*\\}", text, flags=re.DOTALL)
            if not match:
                raise
            return json.loads(match.group(0))

    adjudicated_rows = []

    for _, audit_row in audit_df.iterrows():
        idx = int(audit_row["row_index"])
        master_row = master_df.loc[idx]
        company = safe_text(master_row.get("company"))

        print("")
        print("=" * 80)
        print("QA adjudicating:", company)
        print("=" * 80)

        user_prompt = f"""
CONTROLLED TAXONOMY OPTIONS:
{json.dumps(taxonomy_options, indent=2)}

COMPANY AND CURRENT ASSIGNMENT:
{build_adjudication_context(master_row, audit_row)}

Return only the requested JSON object.
"""

        response = client.responses.create(
            model=TAXONOMY_QA_MODEL,
            input=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            text={
                "format": {
                    "type": "json_schema",
                    "name": "taxonomy_qa_adjudication",
                    "strict": True,
                    "schema": adjudication_schema,
                }
            },
        )

        result = parse_json_response(response)

        corrected_primary = normalize_code(
            result["corrected_primary_market_segment_code"],
            allowed_primary,
            label_to_code,
        )

        if not corrected_primary:
            corrected_primary = safe_text(master_row.get("primary_market_segment_code"))

        corrected_subsegments = filter_subsegments_for_primary(
            result["corrected_subsegment_tags"],
            corrected_primary,
        )

        corrected = {
            "row_index": idx,
            "company": company,
            "previous_primary_market_segment_code": safe_text(master_row.get("primary_market_segment_code")),
            "corrected_primary_market_segment_code": corrected_primary,
            "corrected_primary_market_segment": code_to_label.get(corrected_primary, corrected_primary),
            "current_assignment_valid": result["current_assignment_valid"],
            "qa_confidence": int(result["qa_confidence"]),
            "needs_human_review": result["needs_human_review"],
            "human_review_reason": safe_text(result["human_review_reason"]),
            "qa_decision_rationale": safe_text(result["qa_decision_rationale"]),
            "corrected_subsegment_tags": join_tags(corrected_subsegments),
            "corrected_product_model_tags": join_tags(result["corrected_product_model_tags"]),
            "corrected_distribution_model_tags": join_tags(result["corrected_distribution_model_tags"]),
            "corrected_data_input_tags": join_tags(result["corrected_data_input_tags"]),
            "qa_flags": safe_text(audit_row.get("qa_flags")),
        }

        adjudicated_rows.append(corrected)

        print(
            company,
            "previous:",
            corrected["previous_primary_market_segment_code"],
            "corrected:",
            corrected["corrected_primary_market_segment_code"],
            "valid:",
            corrected["current_assignment_valid"],
            "human review:",
            corrected["needs_human_review"],
        )

    adjudicated_df = pd.DataFrame(adjudicated_rows)
    adjudicated_path = OUTPUT_DIR / f"taxonomy_qa_adjudication_{timestamp}.csv"
    adjudicated_df.to_csv(adjudicated_path, index=False)

    print("")
    print("Adjudication complete.")
    print("Adjudication path:", adjudicated_path)
    display(adjudicated_df)

    adjudicated_df["primary_segment_changed"] = (
        adjudicated_df["previous_primary_market_segment_code"].astype(str).str.strip()
        != adjudicated_df["corrected_primary_market_segment_code"].astype(str).str.strip()
    )

    adjudicated_df["has_subsegment_parent_mismatch"] = (
        adjudicated_df["qa_flags"].astype(str).str.contains("SUBSEGMENT_PARENT_MISMATCH", case=False, na=False)
    )

    # Auto-apply is intentionally conservative:
    # - No primary segment changes auto-apply.
    # - Same-primary tag/product cleanup can auto-apply.
    # - Anything asking for human review stays in review.
    apply_df = adjudicated_df[
        (adjudicated_df["primary_segment_changed"] == False)
        & (adjudicated_df["needs_human_review"] == False)
        & (adjudicated_df["qa_confidence"] >= TAXONOMY_QA_CONFIDENCE_AUTO_APPLY_THRESHOLD)
        & (
            (adjudicated_df["current_assignment_valid"] == False)
            | (adjudicated_df["has_subsegment_parent_mismatch"] == True)
        )
    ].copy()

    confirmed_valid_df = adjudicated_df[
        (adjudicated_df["current_assignment_valid"] == True)
        & (adjudicated_df["needs_human_review"] == False)
        & (adjudicated_df["has_subsegment_parent_mismatch"] == False)
    ].copy()

    human_review_df = adjudicated_df[
        (adjudicated_df["needs_human_review"] == True)
        | (adjudicated_df["primary_segment_changed"] == True)
    ].copy()

    print("")
    print("Corrections eligible for auto-apply:", len(apply_df))
    print("QA-flagged rows confirmed valid:", len(confirmed_valid_df))
    print("Rows needing human review:", len(human_review_df))

    if len(apply_df):
        backup_path = MASTER_PATH.with_name(f"{MASTER_PATH.stem}_backup_before_taxonomy_qa_apply_{timestamp}.csv")
        shutil.copy2(MASTER_PATH, backup_path)
        print("Master backup before QA apply:", backup_path)

        for _, correction in apply_df.iterrows():
            idx = int(correction["row_index"])

            master_df.at[idx, "primary_market_segment_code"] = correction["corrected_primary_market_segment_code"]
            master_df.at[idx, "primary_market_segment"] = correction["corrected_primary_market_segment"]
            master_df.at[idx, "market_segment"] = correction["corrected_primary_market_segment"]
            master_df.at[idx, "subsegment_tags"] = correction["corrected_subsegment_tags"]
            master_df.at[idx, "product_model_tags"] = correction["corrected_product_model_tags"]
            master_df.at[idx, "distribution_model_tags"] = correction["corrected_distribution_model_tags"]
            master_df.at[idx, "data_input_tags"] = correction["corrected_data_input_tags"]
            master_df.at[idx, "taxonomy_assignment_method"] = "llm_taxonomy_qa_corrected"
            master_df.at[idx, "taxonomy_assignment_basis"] = correction["qa_decision_rationale"]
            master_df.at[idx, "taxonomy_rationale"] = correction["qa_decision_rationale"]
            master_df.at[idx, "taxonomy_confidence"] = correction["qa_confidence"]
            master_df.at[idx, "taxonomy_needs_review"] = "NO"
            master_df.at[idx, "taxonomy_review_reason"] = ""

        master_df.to_csv(MASTER_PATH, index=False)
        print("Applied taxonomy QA corrections to master:", MASTER_PATH)

    review_path = OUTPUT_DIR / f"taxonomy_qa_human_review_packet_{timestamp}.csv"
    human_review_df.to_csv(review_path, index=False)
    print("Human review packet:", review_path)



# STEP 28 - Taxonomy QA and safe cleanup pass
# Purpose:
# - Mechanically clean taxonomy inconsistencies.
# - Avoid substring false positives like "gi" inside unrelated words.
# - Never auto-apply primary market segment changes.
# - Auto-apply safe same-primary cleanup only: cross-parent subsegment tags and obvious product-model mismatches.
# - Produce a review packet only for true unresolved primary-segment conflicts.

from pathlib import Path
from datetime import datetime
import re
import shutil
import sys
import pandas as pd

from google.colab import drive

drive.mount("/content/drive", force_remount=False)

REPO_DIR = Path("/content/health-tech-research-agent")
SRC_DIR = REPO_DIR / "src"
TAXONOMY_DIR = REPO_DIR / "taxonomy"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from health_tech_research_agent.taxonomy import (
    safe_text,
    split_tags,
    join_tags,
    load_taxonomy_tables,
    code_label_maps,
    allowed_codes,
)

MASTER_PATH = Path("/content/drive/MyDrive/Job Search/Health Tech Research/health_tech_market_research_summary_MASTER.csv")
RESEARCH_DIR = Path("/content/drive/MyDrive/Job Search/Health Tech Research")
OUTPUT_DIR = RESEARCH_DIR / "taxonomy_backfills"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

if not MASTER_PATH.exists():
    raise FileNotFoundError(f"Master not found: {MASTER_PATH}")

master_df = pd.read_csv(MASTER_PATH).fillna("")

tables = load_taxonomy_tables(TAXONOMY_DIR)
code_to_label, label_to_code = code_label_maps(tables)

subsegments = tables["subsegment_tags"]

subsegment_parent = {}
for _, row in subsegments.iterrows():
    tag = safe_text(row.get("tag_code"))
    parent = safe_text(row.get("parent_market_segment"))
    if tag and parent:
        subsegment_parent[tag] = parent

def normalize_tag_list(value):
    tags = []
    for tag in split_tags(value):
        tag = safe_text(tag)
        if tag and tag not in tags:
            tags.append(tag)
    return tags

def filter_subsegments_for_primary(raw_tags, primary_code):
    kept = []
    removed = []

    for tag in normalize_tag_list(raw_tags):
        parent = subsegment_parent.get(tag)
        if parent and parent != primary_code:
            removed.append(tag)
            continue
        kept.append(tag)

    return kept, removed

def clean_product_tags(raw_tags, distribution_tags):
    tags = normalize_tag_list(raw_tags)
    dist = set(normalize_tag_list(distribution_tags))

    removed = []

    # BENEFITS_PLATFORM should usually require an employer/payer/institutional channel.
    institutional_distribution = {
        "EMPLOYER",
        "PAYER",
        "B2B2C",
        "GOV_MEDICAID",
        "HEALTH_SYSTEM",
        "PROVIDER_GROUP",
    }

    if "BENEFITS_PLATFORM" in tags and not (dist & institutional_distribution):
        tags = [tag for tag in tags if tag != "BENEFITS_PLATFORM"]
        removed.append("BENEFITS_PLATFORM")

    return tags, removed

def row_text(row):
    fields = [
        "company",
        "business_model_classification",
        "final_takeaway",
        "taxonomy_rationale",
        "taxonomy_assignment_basis",
        "commercial_scale_assessment",
        "pmf_scale_assessment",
        "review_notes",
        "priority_review_note",
    ]
    return " | ".join(safe_text(row.get(field, "")) for field in fields)

def explicit_condition_signals(text_value):
    text_value = safe_text(text_value)

    patterns = [
        (r"\boncology\b|\bcancer\b", "SPECIALTY_CONDITION_CARE", "oncology/cancer"),
        (r"\bGI\b|\bgastroenterology\b|\bgastrointestinal\b|\bdigestive\b|\bIBS\b|\bIBD\b|\bCrohn'?s\b|\bcolitis\b", "SPECIALTY_CONDITION_CARE", "GI/digestive"),
        (r"\bMSK\b|\bmusculoskeletal\b|\bphysical therapy\b", "SPECIALTY_CONDITION_CARE", "MSK"),
        (r"\bkidney\b|\brenal\b", "SPECIALTY_CONDITION_CARE", "kidney/renal"),
    ]

    matches = []
    for pattern, expected_segment, label in patterns:
        if re.search(pattern, text_value, flags=re.IGNORECASE):
            matches.append((label, expected_segment))

    return matches

audit_rows = []
cleanup_rows = []
human_review_rows = []

for idx, row in master_df.iterrows():
    company = safe_text(row.get("company"))
    primary = safe_text(row.get("primary_market_segment_code"))

    if not primary:
        primary = safe_text(row.get("market_segment"))

    current_subsegments = safe_text(row.get("subsegment_tags"))
    current_products = safe_text(row.get("product_model_tags"))
    current_distribution = safe_text(row.get("distribution_model_tags"))

    filtered_subsegments, removed_subsegments = filter_subsegments_for_primary(
        current_subsegments,
        primary,
    )

    cleaned_products, removed_products = clean_product_tags(
        current_products,
        current_distribution,
    )

    flags = []

    if removed_subsegments:
        flags.append(f"REMOVED_CROSS_PARENT_SUBSEGMENTS: {', '.join(removed_subsegments)}")

    if removed_products:
        flags.append(f"REMOVED_PRODUCT_MODEL_TAGS: {', '.join(removed_products)}")

    # Primary-segment conflict detection is review-only, never auto-apply.
    condition_matches = explicit_condition_signals(row_text(row))

    if primary == "CARE_NAVIGATION_ACCESS" and condition_matches:
        # Only flag when the evidence explicitly says a condition area.
        condition_labels = ", ".join([m[0] for m in condition_matches])
        flags.append(f"REVIEW_ONLY_CONDITION_SIGNAL_IN_CARE_NAV: {condition_labels}")

    if primary == "PAYER_BENEFITS_INFRASTRUCTURE":
        clinical_review_terms = [
            r"\bmaternity\b",
            r"\bmaternal\b",
            r"\bpregnancy\b",
            r"\bpostpartum\b",
            r"\beating disorder\b",
            r"\boncology\b",
            r"\bcancer\b",
            r"\bdiabetes\b",
            r"\bnutrition\b",
            r"\bGI\b",
            r"\bgastroenterology\b",
            r"\bdigestive\b",
            r"\bMSK\b",
        ]

        clinical_hits = [
            term for term in clinical_review_terms
            if re.search(term, row_text(row), flags=re.IGNORECASE)
        ]

        if clinical_hits:
            flags.append("REVIEW_ONLY_CLINICAL_SIGNAL_IN_PAYER_INFRASTRUCTURE")

    if flags:
        audit_rows.append({
            "row_index": idx,
            "company": company,
            "primary_market_segment_code": primary,
            "market_segment": safe_text(row.get("market_segment")),
            "before_subsegment_tags": current_subsegments,
            "after_subsegment_tags": join_tags(filtered_subsegments),
            "before_product_model_tags": current_products,
            "after_product_model_tags": join_tags(cleaned_products),
            "distribution_model_tags": current_distribution,
            "qa_flags": " | ".join(flags),
        })

    # Safe auto-cleanup: same primary only; no primary segment changes.
    if removed_subsegments or removed_products:
        cleanup_rows.append({
            "row_index": idx,
            "company": company,
            "primary_market_segment_code": primary,
            "subsegment_tags": join_tags(filtered_subsegments),
            "product_model_tags": join_tags(cleaned_products),
            "qa_cleanup_reason": " | ".join(flags),
        })

    # Human review packet only for possible primary-segment conflicts.
    if any(flag.startswith("REVIEW_ONLY") for flag in flags):
        human_review_rows.append({
            "row_index": idx,
            "company": company,
            "current_primary_market_segment_code": primary,
            "current_market_segment": safe_text(row.get("market_segment")),
            "qa_flags": " | ".join(flags),
            "taxonomy_rationale": safe_text(row.get("taxonomy_rationale")),
            "business_model_classification": safe_text(row.get("business_model_classification")),
            "final_takeaway": safe_text(row.get("final_takeaway")),
        })

audit_df = pd.DataFrame(audit_rows)
cleanup_df = pd.DataFrame(cleanup_rows)
human_review_df = pd.DataFrame(human_review_rows)

audit_path = OUTPUT_DIR / f"taxonomy_qa_audit_{timestamp}.csv"
cleanup_path = OUTPUT_DIR / f"taxonomy_qa_cleanup_applied_{timestamp}.csv"
human_review_path = OUTPUT_DIR / f"taxonomy_qa_human_review_packet_{timestamp}.csv"

audit_df.to_csv(audit_path, index=False)
cleanup_df.to_csv(cleanup_path, index=False)
human_review_df.to_csv(human_review_path, index=False)

print("Deterministic taxonomy QA audit complete.")
print("Flagged rows:", len(audit_df))
print("Safe cleanup rows:", len(cleanup_df))
print("Rows needing human review:", len(human_review_df))
print("Audit path:", audit_path)
print("Cleanup path:", cleanup_path)
print("Human review packet:", human_review_path)

if len(audit_df):
    display(audit_df)

if len(cleanup_df):
    backup_path = MASTER_PATH.with_name(f"{MASTER_PATH.stem}_backup_before_taxonomy_safe_cleanup_{timestamp}.csv")
    shutil.copy2(MASTER_PATH, backup_path)
    print("Master backup before safe cleanup:", backup_path)

    for _, cleanup in cleanup_df.iterrows():
        idx = int(cleanup["row_index"])

        master_df.at[idx, "subsegment_tags"] = cleanup["subsegment_tags"]
        master_df.at[idx, "product_model_tags"] = cleanup["product_model_tags"]

        previous_method = safe_text(master_df.at[idx, "taxonomy_assignment_method"])
        if previous_method:
            master_df.at[idx, "taxonomy_assignment_method"] = previous_method
        else:
            master_df.at[idx, "taxonomy_assignment_method"] = "llm_taxonomy_safe_cleanup"

        existing_basis = safe_text(master_df.at[idx, "taxonomy_assignment_basis"])
        cleanup_note = safe_text(cleanup["qa_cleanup_reason"])

        if existing_basis:
            master_df.at[idx, "taxonomy_assignment_basis"] = existing_basis + " | Step 28 safe cleanup: " + cleanup_note
        else:
            master_df.at[idx, "taxonomy_assignment_basis"] = "Step 28 safe cleanup: " + cleanup_note

    master_df.to_csv(MASTER_PATH, index=False)
    print("Applied safe taxonomy cleanup to master:", MASTER_PATH)
else:
    print("No safe cleanup needed.")

print("")
print("Step 28 complete.")

